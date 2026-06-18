from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from app.clients.llm_client import LlmClient
from app.models.common import AgentClaim, PredictionResult
from app.services.agent_control import AgentControlService
from app.services.agents.llm_agents import build_llm_agent_claims
from app.services.predictors.shandong_gas92 import ShandongGas92Predictor
from app.services.market_dataset import MarketDatasetService, PredictionContext
from app.services.postgres_snapshot_repository import PostgresSnapshotRepository
from app.services.predictors.advice_engine import build_driver_summary, build_spread_advice
from app.services.predictors.confidence_engine import build_reliability_score
from app.services.predictors.horizons import DEFAULT_HORIZONS, HorizonConfig, resolve_horizon_config
from app.services.predictors.llm_narrative import enrich_prediction_narrative


REGIONAL_FREIGHT_ESTIMATES = {
    "EAST_CHINA": 55.0,
    "NORTH_CHINA": 45.0,
    "SOUTH_CHINA": 95.0,
    "CENTRAL_CHINA": 75.0,
    "NORTHWEST": 110.0,
    "SOUTHWEST": 120.0,
    "NORTHEAST": 95.0,
}


def _zh(text: str) -> str:
    try:
        return text.encode("ascii").decode("unicode_escape")
    except UnicodeEncodeError:
        return text


REGIONAL_FREIGHT_COMPONENT_CONFIGS: list[dict[str, Any]] = [
    {"region_code": "EAST_CHINA", "region_name": _zh("\u534e\u4e1c"), "component_key": "EAST_CHINA_ZHOUSHAN_SHIP", "short_name": _zh("\u6d59\u6c5f\u77f3\u5316"), "route_name": _zh("\u6d59\u6c5f\u7701\u821f\u5c71\u5e02(\u8239\u8fd0\u8d39)"), "freight_value": 100.0},
    {"region_code": "NORTH_CHINA", "region_name": _zh("\u534e\u5317"), "component_key": "NORTH_CHINA_CANGZHOU_TRUCK", "short_name": _zh("\u6cb3\u5317\u946b\u6d77"), "route_name": _zh("\u6cb3\u5317\u7701\u6ca7\u5dde\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 70.0},
    {"region_code": "NORTH_CHINA", "region_name": _zh("\u534e\u5317"), "component_key": "NORTH_CHINA_PUYANG_TRUCK", "short_name": _zh("\u4e30\u5229\u77f3\u5316"), "route_name": _zh("\u6cb3\u5357\u7701\u6fee\u9633\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 130.0},
    {"region_code": "CENTRAL_CHINA", "region_name": _zh("\u534e\u4e2d"), "component_key": "CENTRAL_CHINA_QIANJIANG_TRUCK", "short_name": _zh("\u91d1\u6fb3\u79d1\u6280"), "route_name": _zh("\u6e56\u5317\u7701\u6f5c\u6c5f\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 260.0},
    {"region_code": "CENTRAL_CHINA", "region_name": _zh("\u534e\u4e2d"), "component_key": "CENTRAL_CHINA_NANCHANG_TRUCK", "short_name": _zh("\u5357\u660c\u8d38\u6613"), "route_name": _zh("\u6c5f\u897f\u7701\u5357\u660c\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 300.0},
    {"region_code": "CENTRAL_CHINA", "region_name": _zh("\u534e\u4e2d"), "component_key": "CENTRAL_CHINA_HEFEI_TRUCK", "short_name": _zh("\u5408\u80a5\u8d38\u6613"), "route_name": _zh("\u5b89\u5fbd\u7701\u5408\u80a5\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 240.0},
    {"region_code": "SOUTH_CHINA", "region_name": _zh("\u534e\u5357"), "component_key": "SOUTH_CHINA_HUIZHOU_SHIP", "short_name": _zh("\u60e0\u5dde\u6d77\u6cb9"), "route_name": _zh("\u5e7f\u4e1c\u7701\u60e0\u5dde\u5e02(\u8239\u8fd0\u8d39)"), "freight_value": 120.0},
    {"region_code": "NORTHWEST", "region_name": _zh("\u897f\u5317"), "component_key": "NORTHWEST_YANAN_TRUCK", "short_name": _zh("\u5ef6\u5b89\u70bc\u5382"), "route_name": _zh("\u9655\u897f\u7701\u5ef6\u5b89\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 260.0},
    {"region_code": "NORTHWEST", "region_name": _zh("\u897f\u5317"), "component_key": "NORTHWEST_WUZHONG_TRUCK", "short_name": _zh("\u5b81\u9c81\u77f3\u5316"), "route_name": _zh("\u5b81\u590f\u5434\u5fe0\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 260.0},
    {"region_code": "SOUTHWEST", "region_name": _zh("\u897f\u5357"), "component_key": "SOUTHWEST_SUINING_TRUCK", "short_name": _zh("\u76db\u9a6c\u5316\u5de5"), "route_name": _zh("\u56db\u5ddd\u7701\u9042\u5b81\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 420.0},
    {"region_code": "NORTHEAST", "region_name": _zh("\u4e1c\u5317"), "component_key": "NORTHEAST_PANJIN_TRUCK", "short_name": _zh("\u9526\u57ce\u77f3\u5316"), "route_name": _zh("\u8fbd\u5b81\u7701\u76d8\u9526\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 180.0},
    {"region_code": "NORTHEAST", "region_name": _zh("\u4e1c\u5317"), "component_key": "NORTHEAST_SONGYUAN_TRUCK", "short_name": _zh("\u677e\u539f\u77f3\u5316"), "route_name": _zh("\u5409\u6797\u7701\u677e\u539f\u5e02(\u8f66\u8fd0\u8d39)"), "freight_value": 200.0},
]


REGIONAL_REFINERY_INVENTORY_INDICATORS = {
    "GASOLINE_92": {
        "NORTHEAST": "ganglian_id01374798",
        "NORTH_CHINA": "ganglian_id01374797",
        "EAST_CHINA": "ganglian_id01374812",
        "SOUTH_CHINA": "ganglian_id01374823",
        "CENTRAL_CHINA": "ganglian_id01374792",
        "NORTHWEST": "ganglian_id01374805",
        "SOUTHWEST": "ganglian_id01374789",
    },
    "DIESEL_0": {
        "NORTHEAST": "ganglian_id01374808",
        "NORTH_CHINA": "ganglian_id01374844",
        "EAST_CHINA": "ganglian_id01374814",
        "SOUTH_CHINA": "ganglian_id01374806",
        "CENTRAL_CHINA": "ganglian_id01374821",
        "NORTHWEST": "ganglian_id01374829",
        "SOUTHWEST": "ganglian_id01374807",
    },
}

REGIONAL_STATE_MIN_SAMPLE = 12
REGIONAL_RULE_CORRECTION_WEIGHT = 0.35
REGIONAL_REQUIRED_MARGIN = 40.0
REGIONAL_RISK_BUFFER = 40.0

FREIGHT_INPUT_WORKBOOK = Path(__file__).resolve().parents[3] / "运费录入.xlsx"

REGIONAL_FREIGHT_EXCEL_LABELS = {
    "山东-华东": "EAST_CHINA",
    "山东-华北": "NORTH_CHINA",
    "山东-华南": "SOUTH_CHINA",
    "山东-华中": "CENTRAL_CHINA",
    "山东-西北": "NORTHWEST",
    "山东-西南": "SOUTHWEST",
    "山东-东北": "NORTHEAST",
}



class RegionalFreightWorkbook:
    def __init__(self, path: Path = FREIGHT_INPUT_WORKBOOK) -> None:
        self.path = path

    def available(self) -> bool:
        return self.path.exists()

    def load_settings(self) -> dict[str, dict[str, Any]]:
        if not self.path.exists():
            return {}
        workbook = load_workbook(self.path, data_only=False)
        value_workbook = load_workbook(self.path, data_only=True)
        sheet = workbook["Sheet1"]
        value_sheet = value_workbook["Sheet1"]
        value_row = self._latest_value_row(sheet)
        if value_row is None:
            return {}
        as_of_cell = sheet.cell(value_row, 1).value
        as_of_date = as_of_cell.date().isoformat() if isinstance(as_of_cell, datetime) else str(as_of_cell or "")
        grouped: dict[str, dict[str, Any]] = {}
        pending_components: list[dict[str, Any]] = []
        for column in range(2, sheet.max_column + 1):
            row1 = str(value_sheet.cell(1, column).value or sheet.cell(1, column).value or "").strip()
            row2 = str(value_sheet.cell(2, column).value or sheet.cell(2, column).value or row1).strip()
            is_aggregate = bool(sheet.cell(1, column).font and sheet.cell(1, column).font.bold) or row2 in REGIONAL_FREIGHT_EXCEL_LABELS
            value = self._float_or_none(value_sheet.cell(value_row, column).value)
            if is_aggregate:
                region_code = REGIONAL_FREIGHT_EXCEL_LABELS.get(row2) or REGIONAL_FREIGHT_EXCEL_LABELS.get(row1)
                if region_code:
                    component_values = [item["freight_value"] for item in pending_components if item.get("freight_value") is not None]
                    calculated = round(sum(component_values) / len(component_values), 4) if component_values else value
                    grouped[region_code] = {
                        "region_code": region_code,
                        "region_name": row2.replace("山东-", "") or row1.replace("山东-", ""),
                        "freight_value": calculated if calculated is not None else 0.0,
                        "workbook_value": value,
                        "unit": "元/吨",
                        "source_type": "excel_aggregate",
                        "updated_by": "运费录入.xlsx",
                        "updated_at": self._file_mtime(),
                        "as_of_date": as_of_date,
                        "excel_column": column,
                        "excel_label": row2 or row1,
                        "components": pending_components,
                        "calculation": self._calculation_text(pending_components, calculated),
                    }
                pending_components = []
            else:
                pending_components.append(
                    {
                        "component_key": self._column_letter(column),
                        "short_name": row1,
                        "route_name": row2,
                        "freight_value": value,
                        "unit": "元/吨",
                        "excel_column": column,
                    }
                )
        return grouped

    def update_component(self, *, component_key: str, freight_value: float) -> dict[str, Any]:
        if not self.path.exists():
            raise RuntimeError(f"未找到运费录入文件：{self.path}")
        workbook = load_workbook(self.path, data_only=False)
        sheet = workbook["Sheet1"]
        value_row = self._latest_value_row(sheet)
        if value_row is None:
            value_row = 3
            sheet.cell(value_row, 1).value = datetime.now().date()
        target_column = self._component_column(sheet, component_key)
        if target_column is None:
            raise ValueError(f"未找到运费录入项：{component_key}")
        sheet.cell(value_row, target_column).value = float(freight_value)
        self._recalculate_aggregates(sheet, value_row)
        workbook.save(self.path)
        return self.load_settings()

    def _recalculate_aggregates(self, sheet: Any, value_row: int) -> None:
        pending_columns: list[int] = []
        for column in range(2, sheet.max_column + 1):
            row1 = str(sheet.cell(1, column).value or "").strip()
            row2 = str(sheet.cell(2, column).value or row1).strip()
            is_aggregate = bool(sheet.cell(1, column).font and sheet.cell(1, column).font.bold) or row2 in REGIONAL_FREIGHT_EXCEL_LABELS
            if is_aggregate:
                values = [self._float_or_none(sheet.cell(value_row, col).value) for col in pending_columns]
                values = [value for value in values if value is not None]
                if values:
                    sheet.cell(value_row, column).value = sum(values) / len(values)
                pending_columns = []
            else:
                pending_columns.append(column)

    def _component_column(self, sheet: Any, component_key: str) -> int | None:
        normalized = str(component_key or "").strip()
        if normalized.isalpha():
            column = self._letter_to_column(normalized)
            if 1 <= column <= sheet.max_column:
                return column
        for column in range(2, sheet.max_column + 1):
            labels = {
                str(sheet.cell(1, column).value or "").strip(),
                str(sheet.cell(2, column).value or "").strip(),
                self._column_letter(column),
            }
            is_aggregate = bool(sheet.cell(1, column).font and sheet.cell(1, column).font.bold) or str(sheet.cell(2, column).value or "").strip() in REGIONAL_FREIGHT_EXCEL_LABELS
            if not is_aggregate and normalized in labels:
                return column
        return None

    def _latest_value_row(self, sheet: Any) -> int | None:
        for row in range(sheet.max_row, 2, -1):
            if any(sheet.cell(row, col).value is not None for col in range(1, sheet.max_column + 1)):
                return row
        return None

    def _file_mtime(self) -> datetime | None:
        try:
            return datetime.fromtimestamp(self.path.stat().st_mtime)
        except OSError:
            return None

    def _calculation_text(self, components: list[dict[str, Any]], value: float | None) -> str:
        labels = [item.get("route_name") or item.get("short_name") for item in components]
        labels = [label for label in labels if label]
        if not labels:
            return "未配置明细线路，使用表内区域运费。"
        result = round(float(value), 2) if value is not None else "-"
        return f"({'+'.join(labels)}) ÷ {len(labels)} = {result} 元/吨"
    def _float_or_none(self, value: Any) -> float | None:
        try:
            if value is None or pd.isna(value):
                return None
            return float(value)
        except Exception:
            return None

    def _column_letter(self, column: int) -> str:
        letters = ""
        while column:
            column, remainder = divmod(column - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def _letter_to_column(self, value: str) -> int:
        column = 0
        for char in value.upper():
            if not ("A" <= char <= "Z"):
                return -1
            column = column * 26 + (ord(char) - 64)
        return column


class RegionalSpreadStructureAgent:
    name = "regional_spread_structure_agent"

    def analyze(self, row: pd.Series, extra: dict[str, Any]) -> AgentClaim:
        current_spread = self._num(row.get("target_region_spread"))
        change_1d = self._num(row.get("target_region_spread_change_1d"))
        change_3d = self._num(row.get("target_region_spread_change_3d"))
        score = 0.0
        evidence: list[str] = []

        if change_3d is not None:
            if change_3d >= 1.5:
                score += 2.0
                evidence.append(f"区域价差3日走扩{change_3d:.1f}元/吨，短线动量仍偏强")
            elif change_3d <= -1.5:
                score -= 2.0
                evidence.append(f"区域价差3日收窄{abs(change_3d):.1f}元/吨，短线动量偏弱")
            elif change_3d >= 0.5:
                score += 1.0
                evidence.append(f"区域价差3日小幅走扩{change_3d:.1f}元/吨")
            elif change_3d <= -0.5:
                score -= 1.0
                evidence.append(f"区域价差3日小幅收窄{abs(change_3d):.1f}元/吨")

        if current_spread is not None:
            if current_spread >= 150:
                score -= 1.5
                evidence.append(f"当前价差{current_spread:.1f}元/吨处于高位，存在回落压力")
            elif current_spread <= 30:
                score += 1.5
                evidence.append(f"当前价差{current_spread:.1f}元/吨处于低位，存在修复空间")

        if change_1d is not None and abs(change_1d) >= 0.5:
            score += 1.0 if change_1d > 0 else -1.0
            evidence.append(f"区域价差单日变化{change_1d:.1f}元/吨")

        return self._claim(score, evidence or ["区域价差结构未出现明显单边信号"])

    def _claim(self, score: float, evidence: list[str]) -> AgentClaim:
        direction = "up" if score > 1 else "down" if score < -1 else "flat"
        return AgentClaim(
            agent_name=self.name,
            direction=direction,
            confidence_label="medium" if abs(score) >= 4 else "low",
            confidence_score=min(0.85, 0.45 + abs(score) / 40),
            summary=f"价差结构贡献{score:.1f}元/吨",
            evidence=evidence,
            numeric_signals={"score": round(score, 4)},
            structured_payload={"factor": "spread_structure"},
        )

    def _num(self, value: Any) -> float | None:
        try:
            if value is None or pd.isna(value):
                return None
            return float(value)
        except Exception:
            return None


class RegionalFreightNetbackAgent:
    name = "regional_freight_netback_agent"

    def analyze(self, row: pd.Series, extra: dict[str, Any]) -> AgentClaim:
        spread = self._num(row.get("target_region_spread"))
        freight = self._num(extra.get("freight_estimate"))
        score = 0.0
        evidence: list[str] = []
        if spread is None or freight is None:
            return self._claim(0.0, ["运费或区域价差缺失，净回款不参与打分"])

        netback_spread = spread - freight
        if netback_spread >= 80:
            score -= 6.0
            evidence.append(f"净回款价差{netback_spread:.1f}元/吨，套利窗口过宽，后续有收敛压力")
        elif netback_spread >= 40:
            score -= 3.0
            evidence.append(f"净回款价差{netback_spread:.1f}元/吨，外流利润较好但需防价差回落")
        elif netback_spread >= 0:
            score -= 1.0
            evidence.append(f"净回款价差{netback_spread:.1f}元/吨，价差维持正区间但收敛压力仍在")
        elif netback_spread <= -60:
            score -= 1.0
            evidence.append(f"净回款价差{netback_spread:.1f}元/吨，目标区域显著弱于山东")
        elif netback_spread <= -20:
            score -= 0.5
            evidence.append(f"净回款价差{netback_spread:.1f}元/吨，目标区域偏弱")
        else:
            evidence.append(f"净回款价差{netback_spread:.1f}元/吨，运费后价差处于中性区间")

        return self._claim(score, evidence)

    def _claim(self, score: float, evidence: list[str]) -> AgentClaim:
        direction = "up" if score > 1 else "down" if score < -1 else "flat"
        return AgentClaim(
            agent_name=self.name,
            direction=direction,
            confidence_label="medium" if abs(score) >= 4 else "low",
            confidence_score=min(0.85, 0.45 + abs(score) / 40),
            summary=f"运费净回款贡献{score:.1f}元/吨",
            evidence=evidence,
            numeric_signals={"score": round(score, 4)},
            structured_payload={"factor": "freight_netback"},
        )

    def _num(self, value: Any) -> float | None:
        try:
            if value is None or pd.isna(value):
                return None
            return float(value)
        except Exception:
            return None


class RegionalInventoryPressureAgent:
    name = "regional_inventory_pressure_agent"

    def analyze(self, row: pd.Series, extra: dict[str, Any]) -> AgentClaim:
        snapshot = extra.get("regional_inventory") or {}
        if not snapshot.get("available"):
            return self._claim(0.0, [str(snapshot.get("reason") or "区域库存数据缺失，不参与打分")], "low")

        ratio = snapshot.get("ratio_to_median")
        wow_change = snapshot.get("wow_change")
        latest_date = snapshot.get("latest_date")
        subject = snapshot.get("subject") or "贸易商"
        score = 0.0
        evidence: list[str] = []

        if ratio is not None:
            if ratio >= 1.25:
                score -= 4.0
                evidence.append(f"{subject}库存较区域中位数高{(ratio - 1) * 100:.1f}%，目标区域承压")
            elif ratio >= 1.12:
                score -= 2.0
                evidence.append(f"{subject}库存较区域中位数高{(ratio - 1) * 100:.1f}%，偏弱")
            elif ratio <= 0.75:
                score += 4.0
                evidence.append(f"{subject}库存较区域中位数低{(1 - ratio) * 100:.1f}%，目标区域抗跌")
            elif ratio <= 0.88:
                score += 2.0
                evidence.append(f"{subject}库存较区域中位数低{(1 - ratio) * 100:.1f}%，偏强")

        if wow_change is not None:
            if wow_change >= 5:
                score -= 2.0
                evidence.append(f"{subject}库存环比增加{wow_change:.2f}万吨")
            elif wow_change <= -5:
                score += 2.0
                evidence.append(f"{subject}库存环比下降{abs(wow_change):.2f}万吨")

        if not evidence:
            evidence.append(f"{latest_date}区域库存处于中性位置")
        return self._claim(score, evidence, "medium" if abs(score) >= 5 else "low")

    def _claim(self, score: float, evidence: list[str], confidence_label: str) -> AgentClaim:
        direction = "up" if score > 1 else "down" if score < -1 else "flat"
        return AgentClaim(
            agent_name=self.name,
            direction=direction,
            confidence_label=confidence_label,
            confidence_score=min(0.9, 0.45 + abs(score) / 35),
            summary=f"区域库存压力贡献{score:.1f}元/吨",
            evidence=evidence,
            numeric_signals={"score": round(score, 4)},
            structured_payload={"factor": "regional_inventory_pressure"},
        )


@dataclass(frozen=True)
class RegionalSpreadConfig:
    region_code: str
    region_name: str
    price_column: str
    spread_column: str
    shandong_price_column: str = "sd_gas92_market"
    product_code: str = "GASOLINE_92_SPREAD"
    product_label: str = "92#"

    @property
    def spread_change_1d_column(self) -> str:
        return f"{self.spread_column}_change_1d"

    @property
    def spread_change_3d_column(self) -> str:
        return f"{self.spread_column}_change_3d"

    @property
    def entity_code(self) -> str:
        suffix = "DIESEL0" if self.product_code == "DIESEL_0_SPREAD" else "GAS92"
        return f"{self.region_code}_VS_SD_{suffix}_SPREAD"

    @property
    def output_region_code(self) -> str:
        return f"{self.region_code}_VS_SHANDONG"


REGIONAL_SPREAD_CONFIGS: dict[str, RegionalSpreadConfig] = {
    "EAST_CHINA": RegionalSpreadConfig("EAST_CHINA", "华东", "east_china_gas92_market", "sd_vs_east_china_spread"),
    "NORTH_CHINA": RegionalSpreadConfig("NORTH_CHINA", "华北", "north_china_gas92_market", "sd_vs_north_china_spread"),
    "SOUTH_CHINA": RegionalSpreadConfig("SOUTH_CHINA", "华南", "south_china_gas92_market", "sd_vs_south_china_spread"),
    "CENTRAL_CHINA": RegionalSpreadConfig(
        "CENTRAL_CHINA", "华中", "central_china_gas92_market", "sd_vs_central_china_spread"
    ),
    "NORTHWEST": RegionalSpreadConfig("NORTHWEST", "西北", "northwest_gas92_market", "sd_vs_northwest_spread"),
    "SOUTHWEST": RegionalSpreadConfig("SOUTHWEST", "西南", "southwest_gas92_market", "sd_vs_southwest_spread"),
    "NORTHEAST": RegionalSpreadConfig("NORTHEAST", "东北", "northeast_gas92_market", "sd_vs_northeast_spread"),
}

DIESEL_REGIONAL_SPREAD_CONFIGS: dict[str, RegionalSpreadConfig] = {
    "EAST_CHINA": RegionalSpreadConfig(
        "EAST_CHINA", "华东", "east_china_diesel0_market", "sd_vs_east_china_diesel_spread", "sd_diesel0_market", "DIESEL_0_SPREAD", "0#柴油"
    ),
    "NORTH_CHINA": RegionalSpreadConfig(
        "NORTH_CHINA", "华北", "north_china_diesel0_market", "sd_vs_north_china_diesel_spread", "sd_diesel0_market", "DIESEL_0_SPREAD", "0#柴油"
    ),
    "SOUTH_CHINA": RegionalSpreadConfig(
        "SOUTH_CHINA", "华南", "south_china_diesel0_market", "sd_vs_south_china_diesel_spread", "sd_diesel0_market", "DIESEL_0_SPREAD", "0#柴油"
    ),
    "CENTRAL_CHINA": RegionalSpreadConfig(
        "CENTRAL_CHINA", "华中", "central_china_diesel0_market", "sd_vs_central_china_diesel_spread", "sd_diesel0_market", "DIESEL_0_SPREAD", "0#柴油"
    ),
    "NORTHWEST": RegionalSpreadConfig(
        "NORTHWEST", "西北", "northwest_diesel0_market", "sd_vs_northwest_diesel_spread", "sd_diesel0_market", "DIESEL_0_SPREAD", "0#柴油"
    ),
    "SOUTHWEST": RegionalSpreadConfig(
        "SOUTHWEST", "西南", "southwest_diesel0_market", "sd_vs_southwest_diesel_spread", "sd_diesel0_market", "DIESEL_0_SPREAD", "0#柴油"
    ),
    "NORTHEAST": RegionalSpreadConfig(
        "NORTHEAST", "东北", "northeast_diesel0_market", "sd_vs_northeast_diesel_spread", "sd_diesel0_market", "DIESEL_0_SPREAD", "0#柴油"
    ),
}


def attach_regional_price_forecasts(
    regional_predictions: list[PredictionResult],
    outright_predictions: list[PredictionResult],
) -> list[PredictionResult]:
    outright_by_horizon = {item.horizon: item for item in outright_predictions}
    fallback_outright = outright_predictions[0] if outright_predictions else None
    for regional in regional_predictions:
        outright = outright_by_horizon.get(regional.horizon) or fallback_outright
        if outright is None:
            continue
        raw_context = dict(regional.raw_context or {})
        current_region_price = _float_or_none(raw_context.get("current_counter_region_price"))
        current_shandong_price = _float_or_none(raw_context.get("current_shandong_price"))
        actual_region_minus_shandong = _float_or_none(raw_context.get("current_spread"))
        if current_region_price is None:
            if current_shandong_price is not None and actual_region_minus_shandong is not None:
                current_region_price = current_shandong_price + actual_region_minus_shandong
            else:
                current_region_price = float(regional.point_value)
        if current_shandong_price is None and actual_region_minus_shandong is not None:
            current_shandong_price = current_region_price - actual_region_minus_shandong
        predicted_region_price = float(regional.point_value)
        predicted_shandong_price = float(outright.point_value)
        raw_context.update(
            {
                "predicted_shandong_price": round(predicted_shandong_price, 2),
                "predicted_region_price": round(predicted_region_price, 2),
                "predicted_region_price_range_lower": round(float(regional.range_lower), 2),
                "predicted_region_price_range_upper": round(float(regional.range_upper), 2),
                "predicted_region_minus_shandong_spread": round(predicted_region_price - predicted_shandong_price, 2),
                "predicted_region_minus_shandong_spread_range_lower": round(float(regional.range_lower) - predicted_shandong_price, 2),
                "predicted_region_minus_shandong_spread_range_upper": round(float(regional.range_upper) - predicted_shandong_price, 2),
                "actual_region_minus_shandong_spread": actual_region_minus_shandong,
                "predicted_shandong_minus_region_spread": round(predicted_shandong_price - predicted_region_price, 2),
                "predicted_shandong_minus_region_spread_range_lower": round(predicted_shandong_price - float(regional.range_upper), 2),
                "predicted_shandong_minus_region_spread_range_upper": round(predicted_shandong_price - float(regional.range_lower), 2),
                "actual_shandong_minus_region_spread": round(-actual_region_minus_shandong, 2)
                if actual_region_minus_shandong is not None
                else None,
                "regional_price_prediction_mode": raw_context.get("regional_price_prediction_mode") or "regionalized_shandong_market_logic",
                "regional_price_prediction_formula": raw_context.get("regional_price_prediction_formula")
                or "\u9884\u6d4b\u533a\u57df\u5355\u4ef7=\u533a\u57df\u5f53\u524d\u4ef7+\u5c71\u4e1c\u540c\u6b3e\u9884\u6d4b\u903b\u8f91\u5728\u533a\u57df\u6570\u636e\u4e0a\u91cd\u7b97\u7684\u53d8\u5316\u3002",
                "predicted_spread_formula": "\u9884\u6d4b\u533a\u57df\u4ef7\u5dee=\u533a\u57df\u9884\u6d4b\u4ef7-\u5c71\u4e1c\u9884\u6d4b\u4ef7",
                "actual_spread_formula": "\u5f53\u524d\u533a\u57df\u4ef7\u5dee=\u5f53\u524d\u533a\u57df\u4ef7-\u5f53\u524d\u5c71\u4e1c\u4ef7",
                "shandong_prediction_run_id": outright.run_id,
            }
        )
        composite_variant = raw_context.get("regional_composite_prediction")
        if not isinstance(composite_variant, dict):
            composite_variant = _regional_variant_from_outright_prediction(
                prediction=regional,
                current_region_price=float(current_region_price),
                current_shandong_price=current_shandong_price,
                model_name="\u533a\u57df\u667a\u80fd\u4f53\u9884\u6d4b",
                prediction_type="regional_composite",
                shandong_prediction_source="\u533a\u57df\u5316\u5c71\u4e1c\u667a\u80fd\u4f53\u903b\u8f91",
            )
        raw_context["regional_composite_prediction"] = _attach_region_price_fields_to_variant(
            composite_variant,
            predicted_shandong_price=predicted_shandong_price,
        )
        baseline_variant = raw_context.get("regional_baseline_prediction")
        if isinstance(baseline_variant, dict):
            raw_context["regional_baseline_prediction"] = _attach_region_price_fields_to_variant(
                baseline_variant,
                predicted_shandong_price=_float_or_none(baseline_variant.get("predicted_shandong_price")) or predicted_shandong_price,
            )
        raw_context["regional_prediction_variants"] = [
            raw_context["regional_composite_prediction"],
            *(
                [raw_context["regional_baseline_prediction"]]
                if isinstance(raw_context.get("regional_baseline_prediction"), dict)
                else []
            ),
        ]
        regional.raw_context = raw_context
    return regional_predictions


def _attach_region_price_fields_to_variant(variant: dict[str, Any], *, predicted_shandong_price: float) -> dict[str, Any]:
    result = dict(variant)
    point = _float_or_none(result.get("predicted_region_price")) or _float_or_none(result.get("point_value"))
    if point is None:
        point = _float_or_none(result.get("predicted_region_minus_shandong_spread"))
        if point is not None:
            point = predicted_shandong_price + point
    if point is None:
        return result
    lower = _float_or_none(result.get("predicted_region_price_range_lower")) or _float_or_none(result.get("range_lower")) or point
    upper = _float_or_none(result.get("predicted_region_price_range_upper")) or _float_or_none(result.get("range_upper")) or point
    result.update(
        {
            "predicted_shandong_price": round(float(predicted_shandong_price), 2),
            "predicted_region_price": round(float(point), 2),
            "predicted_region_price_range_lower": round(float(lower), 2),
            "predicted_region_price_range_upper": round(float(upper), 2),
            "predicted_region_minus_shandong_spread": round(float(point) - float(predicted_shandong_price), 2),
            "predicted_region_minus_shandong_spread_range_lower": round(float(lower) - float(predicted_shandong_price), 2),
            "predicted_region_minus_shandong_spread_range_upper": round(float(upper) - float(predicted_shandong_price), 2),
            "predicted_shandong_minus_region_spread": round(float(predicted_shandong_price) - float(point), 2),
            "predicted_shandong_minus_region_spread_range_lower": round(float(predicted_shandong_price) - float(upper), 2),
            "predicted_shandong_minus_region_spread_range_upper": round(float(predicted_shandong_price) - float(lower), 2),
        }
    )
    return result


def _regional_variant_from_outright_prediction(
    *,
    prediction: PredictionResult,
    current_region_price: float,
    current_shandong_price: float | None,
    model_name: str,
    prediction_type: str,
    shandong_prediction_source: str,
) -> dict[str, Any]:
    shandong_anchor = float((prediction.raw_context or {}).get("current_price") or current_shandong_price or prediction.point_value)
    predicted_delta = float(prediction.point_value) - shandong_anchor
    lower_delta = float(prediction.range_lower) - shandong_anchor
    upper_delta = float(prediction.range_upper) - shandong_anchor
    predicted_region_price = current_region_price + predicted_delta
    range_lower = current_region_price + lower_delta
    range_upper = current_region_price + upper_delta
    predicted_shandong_price = float(prediction.point_value)
    return {
        "model_name": model_name,
        "prediction_type": prediction_type,
        "direction_label": prediction.direction_label,
        "score": round(float(prediction.score_value), 4),
        "predicted_delta": round(predicted_delta, 4),
        "predicted_region_price": round(predicted_region_price, 2),
        "predicted_region_price_range_lower": round(range_lower, 2),
        "predicted_region_price_range_upper": round(range_upper, 2),
        "predicted_shandong_price": round(predicted_shandong_price, 2),
        "predicted_region_minus_shandong_spread": round(predicted_region_price - predicted_shandong_price, 2),
        "predicted_region_minus_shandong_spread_range_lower": round(range_lower - predicted_shandong_price, 2),
        "predicted_region_minus_shandong_spread_range_upper": round(range_upper - predicted_shandong_price, 2),
        "predicted_shandong_minus_region_spread": round(predicted_shandong_price - predicted_region_price, 2),
        "predicted_shandong_minus_region_spread_range_lower": round(predicted_shandong_price - range_upper, 2),
        "predicted_shandong_minus_region_spread_range_upper": round(predicted_shandong_price - range_lower, 2),
        "range_half_width": round(max(abs(predicted_region_price - range_lower), abs(range_upper - predicted_region_price)), 4),
        "method": "shandong_market_logic_delta_replay",
        "basis": "\u533a\u57df\u5355\u4ef7\u4e0d\u518d\u7528\u533a\u57df\u4ef7\u5dee\u56e0\u5b50\u5355\u72ec\u9884\u6d4b\uff1b\u6539\u4e3a\u5b8c\u5168\u590d\u7528\u5c71\u4e1c\u5e02\u573a\u4ef7\u683c\u540c\u5468\u671f\u9884\u6d4b\u903b\u8f91\uff0c\u628a\u5c71\u4e1c\u9884\u6d4b\u53d8\u5316\u5e73\u79fb\u5230\u533a\u57df\u5f53\u524d\u4ef7\u3002",
        "calculation": f"\u533a\u57df\u5f53\u524d\u4ef7{round(current_region_price, 2)} + \u5c71\u4e1c\u9884\u6d4b\u53d8\u5316{round(predicted_delta, 4)} = {round(predicted_region_price, 2)}",
        "factor_breakdown": list(prediction.factor_breakdown or []),
        "scorecard": (prediction.raw_context or {}).get("business_scorecard"),
        "business_scorecard": (prediction.raw_context or {}).get("business_scorecard"),
        "agent_claims": (prediction.raw_context or {}).get("agent_claims"),
        "shandong_prediction_source": shandong_prediction_source,
        "regional_price_prediction_formula": "\u9884\u6d4b\u533a\u57df\u5355\u4ef7=\u5f53\u524d\u533a\u57df\u5355\u4ef7+\u5c71\u4e1c\u540c\u5468\u671f\u9884\u6d4b\u53d8\u5316",
    }


def _regional_variant_from_business_prediction(
    *,
    business_prediction: dict[str, Any],
    scorecard: dict[str, Any] | None,
    current_region_price: float,
    current_shandong_price: float | None,
    shandong_prediction_source: str,
) -> dict[str, Any]:
    shandong_point = _float_or_none(business_prediction.get("point_value"))
    if shandong_point is None:
        return {}
    anchor = _float_or_none(business_prediction.get("current_price")) or current_shandong_price or shandong_point
    predicted_delta = shandong_point - anchor
    range_lower_delta = (_float_or_none(business_prediction.get("range_lower")) or shandong_point) - anchor
    range_upper_delta = (_float_or_none(business_prediction.get("range_upper")) or shandong_point) - anchor
    predicted_region_price = current_region_price + predicted_delta
    range_lower = current_region_price + range_lower_delta
    range_upper = current_region_price + range_upper_delta
    return {
        "model_name": "\u533a\u57df\u4e1a\u52a1\u903b\u8f91\u9884\u6d4b\uff08\u590d\u523b\u5c71\u4e1c\u903b\u8f91\uff09",
        "prediction_type": "regional_baseline",
        "direction_label": business_prediction.get("direction_label") or "flat",
        "score": round(float(business_prediction.get("score") or 0.0), 4),
        "predicted_delta": round(predicted_delta, 4),
        "predicted_region_price": round(predicted_region_price, 2),
        "predicted_region_price_range_lower": round(range_lower, 2),
        "predicted_region_price_range_upper": round(range_upper, 2),
        "predicted_shandong_price": round(shandong_point, 2),
        "predicted_region_minus_shandong_spread": round(predicted_region_price - shandong_point, 2),
        "predicted_region_minus_shandong_spread_range_lower": round(range_lower - shandong_point, 2),
        "predicted_region_minus_shandong_spread_range_upper": round(range_upper - shandong_point, 2),
        "predicted_shandong_minus_region_spread": round(shandong_point - predicted_region_price, 2),
        "predicted_shandong_minus_region_spread_range_lower": round(shandong_point - range_upper, 2),
        "predicted_shandong_minus_region_spread_range_upper": round(shandong_point - range_lower, 2),
        "range_half_width": business_prediction.get("range_half_width"),
        "method": "shandong_business_scorecard_delta_replay",
        "basis": "\u533a\u57df\u4e1a\u52a1\u903b\u8f91\u5b8c\u5168\u590d\u7528\u5c71\u4e1c\u4e1a\u52a1\u6253\u5206\u903b\u8f91\uff0c\u4ee5\u5c71\u4e1c\u4e1a\u52a1\u9884\u6d4b\u53d8\u5316\u5e73\u79fb\u5230\u533a\u57df\u5f53\u524d\u4ef7\u3002",
        "calculation": f"\u533a\u57df\u5f53\u524d\u4ef7{round(current_region_price, 2)} + \u5c71\u4e1c\u4e1a\u52a1\u9884\u6d4b\u53d8\u5316{round(predicted_delta, 4)} = {round(predicted_region_price, 2)}",
        "scorecard": scorecard,
        "factor_breakdown": [],
        "shandong_prediction_source": shandong_prediction_source,
        "regional_price_prediction_formula": "\u9884\u6d4b\u533a\u57df\u5355\u4ef7=\u5f53\u524d\u533a\u57df\u5355\u4ef7+\u5c71\u4e1c\u4e1a\u52a1\u903b\u8f91\u9884\u6d4b\u53d8\u5316",
    }


def _regional_variant_from_legacy(
    variant: Any,
    *,
    model_name: str,
    prediction_type: str,
    direction_label: str,
    point_value: float,
    range_lower: float,
    range_upper: float,
    predicted_delta: Any,
    method: Any,
    basis: str,
) -> dict[str, Any]:
    if isinstance(variant, dict):
        return dict(variant)
    return {
        "model_name": model_name,
        "prediction_type": prediction_type,
        "direction_label": direction_label,
        "predicted_delta": _float_or_none(predicted_delta),
        "predicted_region_minus_shandong_spread": round(float(point_value), 2),
        "predicted_region_minus_shandong_spread_range_lower": round(float(range_lower), 2),
        "predicted_region_minus_shandong_spread_range_upper": round(float(range_upper), 2),
        "method": method,
        "basis": basis,
    }


def _attach_regional_variant_prices(
    variant: dict[str, Any],
    *,
    predicted_shandong_price: float,
    shandong_prediction_source: str,
) -> dict[str, Any]:
    result = dict(variant)
    spread = _float_or_none(
        result.get("predicted_region_minus_shandong_spread")
        if result.get("predicted_region_minus_shandong_spread") is not None
        else result.get("point_value")
    )
    if spread is None:
        return result
    range_lower = _float_or_none(result.get("predicted_region_minus_shandong_spread_range_lower"))
    range_upper = _float_or_none(result.get("predicted_region_minus_shandong_spread_range_upper"))
    if range_lower is None:
        range_lower = spread
    if range_upper is None:
        range_upper = spread
    result.update(
        {
            "predicted_shandong_price": round(float(predicted_shandong_price), 2),
            "shandong_prediction_source": shandong_prediction_source,
            "predicted_region_price": round(float(predicted_shandong_price) + spread, 2),
            "predicted_region_price_range_lower": round(float(predicted_shandong_price) + range_lower, 2),
            "predicted_region_price_range_upper": round(float(predicted_shandong_price) + range_upper, 2),
            "predicted_region_minus_shandong_spread": round(spread, 2),
            "predicted_region_minus_shandong_spread_range_lower": round(range_lower, 2),
            "predicted_region_minus_shandong_spread_range_upper": round(range_upper, 2),
            "predicted_shandong_minus_region_spread": round(-spread, 2),
            "predicted_shandong_minus_region_spread_range_lower": round(-range_upper, 2),
            "predicted_shandong_minus_region_spread_range_upper": round(-range_lower, 2),
            "regional_price_prediction_formula": "预测区域单价=对应山东92#预测价+区域预测价差",
            "predicted_spread_formula": "预测展示价差=山东92#预测价-预测区域单价",
        }
    )
    return result


def _float_or_none(value: Any) -> float | None:
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None



def build_freight_settings_from_components(snapshot_repository: PostgresSnapshotRepository | None) -> list[dict[str, Any]]:
    if snapshot_repository is not None:
        components = snapshot_repository.list_freight_components(REGIONAL_FREIGHT_COMPONENT_CONFIGS)
    else:
        components = [
            dict(item, unit="元/吨", display_order=index, updated_by=None, updated_at=None)
            for index, item in enumerate(REGIONAL_FREIGHT_COMPONENT_CONFIGS, start=1)
        ]
    grouped: dict[str, list[dict[str, Any]]] = {}
    for component in components:
        grouped.setdefault(component["region_code"], []).append(component)
    items: list[dict[str, Any]] = []
    for config in REGIONAL_SPREAD_CONFIGS.values():
        region_components = sorted(grouped.get(config.region_code, []), key=lambda item: item.get("display_order") or 0)
        if not region_components:
            items.append(
                {
                    "region_code": config.region_code,
                    "region_name": config.region_name,
                    "freight_value": REGIONAL_FREIGHT_ESTIMATES.get(config.region_code, 80.0),
                    "unit": "元/吨",
                    "source_type": "default",
                    "updated_by": None,
                    "updated_at": None,
                    "components": [],
                    "calculation": "未配置运费明细，使用默认估算。",
                }
            )
            continue
        values = [float(item["freight_value"]) for item in region_components]
        average = round(sum(values) / len(values), 4)
        latest_updated_at = max((item.get("updated_at") for item in region_components if item.get("updated_at")), default=None)
        labels = [item.get("route_name") or item.get("short_name") or item.get("component_key") for item in region_components]
        items.append(
            {
                "region_code": config.region_code,
                "region_name": config.region_name,
                "freight_value": average,
                "unit": "元/吨",
                "source_type": "database_components",
                "updated_by": next((item.get("updated_by") for item in region_components if item.get("updated_by")), None),
                "updated_at": latest_updated_at,
                "components": [
                    {
                        "component_key": item["component_key"],
                        "short_name": item.get("short_name"),
                        "route_name": item.get("route_name"),
                        "freight_value": float(item["freight_value"]),
                        "unit": "元/吨",
                        "excel_column": None,
                    }
                    for item in region_components
                ],
                "calculation": f"({'+'.join(labels)}) ÷ {len(labels)} = {round(average, 2)} 元/吨",
            }
        )
    return items

class ShandongRegionalSpreadPredictor:
    def __init__(
        self,
        dataset_service: MarketDatasetService,
        llm_client: LlmClient,
        agent_control_service: AgentControlService,
        snapshot_repository: PostgresSnapshotRepository | None = None,
        outright_predictor: ShandongGas92Predictor | None = None,
    ) -> None:
        self.dataset_service = dataset_service
        self.llm_client = llm_client
        self.agent_control_service = agent_control_service
        self.snapshot_repository = snapshot_repository
        self.outright_predictor = outright_predictor
        self.scope_key = "regional_spread"
        self._multi_prediction_cache: dict[str, tuple[datetime, dict[str, list[PredictionResult]]]] = {}
        self._multi_prediction_cache_seconds = 600
        self._regional_inventory_cache: dict[date, list[dict[str, Any]]] = {}
        self._regional_inventory_context_cache: dict[tuple[str, date], dict[str, Any]] = {}
        self.agents = [
            RegionalSpreadStructureAgent(),
            RegionalFreightNetbackAgent(),
            RegionalInventoryPressureAgent(),
        ]

    def list_regions(self, product_code: str = "GASOLINE_92") -> list[dict[str, str]]:
        return [{"region_code": item.region_code, "region_name": item.region_name} for item in self._config_map(product_code).values()]

    def run_prediction(
        self,
        region_code: str,
        as_of_date: date | None = None,
        horizon: str = "D1",
        use_llm_explainer: bool = True,
        scenario_text: str | None = None,
        enable_refined_news: bool = True,
        enable_event_risk: bool = True,
        product_code: str = "GASOLINE_92",
    ) -> PredictionResult:
        as_of_date = as_of_date or date.today()
        context = self.dataset_service.build_context(as_of_date)
        return self.run_prediction_from_context(
            context=context,
            region_code=region_code,
            as_of_date=as_of_date,
            horizon=horizon,
            use_llm_explainer=use_llm_explainer,
            scenario_text=scenario_text,
            enable_refined_news=enable_refined_news,
            enable_event_risk=enable_event_risk,
            product_code=product_code,
        )

    def run_prediction_from_context(
        self,
        context: PredictionContext,
        region_code: str,
        as_of_date: date,
        horizon: str = "D1",
        use_llm_explainer: bool = True,
        scenario_text: str | None = None,
        enable_refined_news: bool = True,
        enable_event_risk: bool = True,
        product_code: str = "GASOLINE_92",
    ) -> PredictionResult:
        config = self._resolve_config(region_code, product_code=product_code)
        horizon_config = resolve_horizon_config(horizon)
        return self._predict_from_context(
            context=context,
            config=config,
            horizon_config=horizon_config,
            as_of_date=as_of_date,
            use_llm_explainer=use_llm_explainer,
            scenario_text=scenario_text,
            enable_refined_news=enable_refined_news,
            enable_event_risk=enable_event_risk,
            context_metadata=context.metadata,
            product_code=product_code,
        )

    def run_all_predictions(
        self,
        as_of_date: date | None = None,
        horizon: str = "D1",
        use_llm_explainer: bool = True,
        scenario_text: str | None = None,
        enable_refined_news: bool = True,
        enable_event_risk: bool = True,
        region_codes: list[str] | None = None,
        product_code: str = "GASOLINE_92",
    ) -> list[PredictionResult]:
        as_of_date = as_of_date or date.today()
        context = self.dataset_service.build_context(as_of_date)
        return self.run_all_predictions_from_context(
            context=context,
            as_of_date=as_of_date,
            horizon=horizon,
            use_llm_explainer=use_llm_explainer,
            scenario_text=scenario_text,
            enable_refined_news=enable_refined_news,
            enable_event_risk=enable_event_risk,
            region_codes=region_codes,
            product_code=product_code,
        )

    def run_all_predictions_from_context(
        self,
        context: PredictionContext,
        as_of_date: date,
        horizon: str = "D1",
        use_llm_explainer: bool = True,
        scenario_text: str | None = None,
        enable_refined_news: bool = True,
        enable_event_risk: bool = True,
        region_codes: list[str] | None = None,
        product_code: str = "GASOLINE_92",
    ) -> list[PredictionResult]:
        horizon_config = resolve_horizon_config(horizon)
        configs = self._resolve_configs(region_codes, product_code=product_code)
        return [
            self._predict_from_context(
                context=context,
                config=config,
                horizon_config=horizon_config,
                as_of_date=as_of_date,
                use_llm_explainer=use_llm_explainer,
                scenario_text=scenario_text,
                enable_refined_news=enable_refined_news,
                enable_event_risk=enable_event_risk,
                context_metadata=context.metadata,
            )
            for config in configs
        ]

    def run_multi_horizon_predictions_from_context(
        self,
        context: PredictionContext,
        as_of_date: date,
        region_codes: list[str] | None = None,
        horizons: list[str] | None = None,
        use_llm_explainer: bool = True,
        scenario_text: str | None = None,
        enable_refined_news: bool = True,
        enable_event_risk: bool = True,
        product_code: str = "GASOLINE_92",
    ) -> dict[str, list[PredictionResult]]:
        selected_horizons = horizons or DEFAULT_HORIZONS
        configs = self._resolve_configs(region_codes, product_code=product_code)
        cache_key = self._multi_prediction_cache_key(
            as_of_date=as_of_date,
            region_codes=[config.region_code for config in configs],
            horizons=selected_horizons,
            use_llm_explainer=use_llm_explainer,
            scenario_text=scenario_text,
            enable_refined_news=enable_refined_news,
            enable_event_risk=enable_event_risk,
            context_metadata=context.metadata,
            product_code=product_code,
        )
        now = datetime.now()
        cached = self._multi_prediction_cache.get(cache_key)
        if cached and (now - cached[0]).total_seconds() <= self._multi_prediction_cache_seconds:
            return {horizon: list(items) for horizon, items in cached[1].items()}
        result: dict[str, list[PredictionResult]] = {horizon: [] for horizon in selected_horizons}
        freight_settings = {item["region_code"]: item for item in self.list_freight_settings()}
        regional_cache = {
            (config.region_code, float((freight_settings.get(config.region_code) or {}).get("freight_value", REGIONAL_FREIGHT_ESTIMATES.get(config.region_code, 80.0)))): self._build_regional_history_cache(
                frame=context.feature_frame,
                config=config,
                as_of_date=as_of_date,
                freight_estimate=float((freight_settings.get(config.region_code) or {}).get("freight_value", REGIONAL_FREIGHT_ESTIMATES.get(config.region_code, 80.0))),
            )
            for config in configs
        }
        tasks: list[tuple[str, RegionalSpreadConfig, HorizonConfig]] = [
            (horizon, config, resolve_horizon_config(horizon))
            for horizon in selected_horizons
            for config in configs
        ]
        if not tasks:
            return result
        for horizon, config, horizon_config in tasks:
            result[horizon].append(
                self._predict_from_context(
                    context=context,
                    config=config,
                    horizon_config=horizon_config,
                    as_of_date=as_of_date,
                    use_llm_explainer=use_llm_explainer,
                    scenario_text=scenario_text,
                    enable_refined_news=enable_refined_news,
                    enable_event_risk=enable_event_risk,
                    context_metadata=context.metadata,
                    freight_settings=freight_settings,
                    regional_cache=regional_cache,
                )
            )
        order = {config.region_code: index for index, config in enumerate(configs)}
        for horizon in result:
            result[horizon].sort(key=lambda item: order.get((item.raw_context or {}).get("counter_region_code") or item.region_code, 999))
        self._multi_prediction_cache[cache_key] = (now, {horizon: list(items) for horizon, items in result.items()})
        if len(self._multi_prediction_cache) > 12:
            oldest_key = min(self._multi_prediction_cache, key=lambda key: self._multi_prediction_cache[key][0])
            self._multi_prediction_cache.pop(oldest_key, None)
        return result

    def _multi_prediction_cache_key(
        self,
        *,
        as_of_date: date,
        region_codes: list[str],
        horizons: list[str],
        use_llm_explainer: bool,
        scenario_text: str | None,
        enable_refined_news: bool,
        enable_event_risk: bool,
        context_metadata: dict[str, Any] | None,
        product_code: str,
    ) -> str:
        payload = {
            "as_of_date": as_of_date.isoformat(),
            "region_codes": region_codes,
            "horizons": horizons,
            "product_code": product_code,
            "use_llm_explainer": use_llm_explainer,
            "scenario_text": scenario_text or "",
            "enable_refined_news": enable_refined_news,
            "enable_event_risk": enable_event_risk,
            "market_data_mode": (context_metadata or {}).get("market_data_mode"),
            "market_data_reason": (context_metadata or {}).get("market_data_reason"),
            "regional_logic_version": "regionalized_shandong_required_v2",
        }
        return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()

    def _try_region_market_price_prediction(
        self,
        *,
        context: PredictionContext,
        config: RegionalSpreadConfig,
        as_of_date: date,
        horizon_config: HorizonConfig,
        use_llm_explainer: bool,
        scenario_text: str | None,
        enable_refined_news: bool,
        enable_event_risk: bool,
        context_metadata: dict[str, Any] | None,
        current_row: pd.Series,
        current_region_price: float,
        current_shandong_price: float,
        current_spread_value: float,
        freight_estimate: float,
        freight_setting: dict[str, Any],
        regional_inventory: dict[str, Any],
    ) -> PredictionResult | None:
        if self.outright_predictor is None:
            return None
        region_context, data_quality = self._build_region_market_prediction_context(
            context=context,
            config=config,
            as_of_date=as_of_date,
        )
        try:
            if config.product_code == "DIESEL_0_SPREAD":
                prediction = self.outright_predictor.run_diesel0_prediction_from_context(
                    context=region_context,
                    as_of_date=as_of_date,
                    horizon=horizon_config.code,
                    use_llm_explainer=use_llm_explainer,
                    scenario_text=scenario_text,
                    enable_refined_news=enable_refined_news,
                    enable_event_risk=enable_event_risk,
                )
            else:
                prediction = self.outright_predictor.run_prediction_from_context(
                    context=region_context,
                    as_of_date=as_of_date,
                    horizon=horizon_config.code,
                    use_llm_explainer=use_llm_explainer,
                    scenario_text=scenario_text,
                    enable_refined_news=enable_refined_news,
                    enable_event_risk=enable_event_risk,
                )
        except Exception as exc:
            raise RuntimeError(f"regionalized_shandong_market_logic_failed: {exc}") from exc
        raw_context = dict(prediction.raw_context or {})
        business_scorecard = raw_context.get("business_scorecard")
        if isinstance(business_scorecard, dict):
            business_scorecard = self._augment_scorecard_with_region_quality(business_scorecard, data_quality)
            raw_context["business_scorecard"] = business_scorecard
            if isinstance(raw_context.get("business_scorecard_prediction"), dict):
                raw_context["business_scorecard_prediction"]["scorecard"] = business_scorecard
        raw_context.update(
            {
                "current_spread": round(current_spread_value, 2),
                "spread_formula": f"\u76ee\u6807\u533a\u57df{config.product_label}\u4ef7\u683c-\u5c71\u4e1c{config.product_label}\u4ef7\u683c",
                "current_shandong_price": round(current_shandong_price, 2),
                "current_counter_region_price": round(current_region_price, 2),
                "counter_region_code": config.region_code,
                "counter_region_name": config.region_name,
                "freight_estimate": round(freight_estimate, 2),
                "freight_source": freight_setting.get("source_type", "default"),
                "freight_components": freight_setting.get("components", []),
                "regional_inventory": regional_inventory,
                "regional_price_prediction_mode": "regionalized_shandong_market_logic",
                "regional_price_prediction_formula": "\u590d\u523b\u5c71\u4e1c\u5e02\u573a\u4ef7\u9884\u6d4b\u903b\u8f91\uff1a\u4ef7\u683c/\u5f00\u5de5\u7387/\u5e93\u5b58/\u68c0\u4fee\u7b49\u5c71\u4e1c\u4e13\u5c5e\u5b57\u6bb5\u4f18\u5148\u66ff\u6362\u4e3a\u76ee\u6807\u533a\u57df\u5bf9\u5e94\u6570\u636e\uff1b\u7f3a\u5931\u5219\u5728\u8be6\u60c5\u4e2d\u663e\u793a\u7f3a\u5931\uff0c\u4e0d\u7528\u5c71\u4e1c\u6570\u636e\u66ff\u4ee3\u3002",
                "regional_data_quality": data_quality,
                "market_data_mode": (context_metadata or {}).get("market_data_mode"),
                "market_data_reason": (context_metadata or {}).get("market_data_reason"),
            }
        )
        business_prediction = raw_context.get("business_scorecard_prediction") or {}
        raw_context["regional_composite_prediction"] = _regional_variant_from_outright_prediction(
            prediction=prediction,
            current_region_price=current_region_price,
            current_shandong_price=current_region_price,
            model_name="\u533a\u57df\u667a\u80fd\u4f53\u9884\u6d4b\uff08\u533a\u57df\u5316\u5c71\u4e1c\u903b\u8f91\uff09",
            prediction_type="regional_composite",
            shandong_prediction_source="\u533a\u57df\u5316\u5c71\u4e1c\u667a\u80fd\u4f53\u903b\u8f91",
        )
        if isinstance(business_prediction, dict) and business_prediction:
            raw_context["regional_baseline_prediction"] = _regional_variant_from_business_prediction(
                business_prediction=business_prediction,
                scorecard=raw_context.get("business_scorecard"),
                current_region_price=current_region_price,
                current_shandong_price=current_region_price,
                shandong_prediction_source="\u533a\u57df\u5316\u5c71\u4e1c\u4e1a\u52a1\u6253\u5206\u903b\u8f91",
            )
        raw_context["regional_prediction_variants"] = [
            raw_context["regional_composite_prediction"],
            *([raw_context["regional_baseline_prediction"]] if isinstance(raw_context.get("regional_baseline_prediction"), dict) else []),
        ]
        return prediction.model_copy(
            update={
                "run_id": f"regional-{config.region_code.lower()}-{prediction.run_id}",
                "entity_code": config.entity_code,
                "region_code": config.output_region_code,
                "product_code": config.product_code,
                "point_value": prediction.point_value,
                "range_lower": prediction.range_lower,
                "range_upper": prediction.range_upper,
                "explanation": f"{config.region_name}{config.product_label}\u6309\u533a\u57df\u5316\u5c71\u4e1c\u903b\u8f91\u91cd\u7b97\uff1a\u5f53\u524d\u533a\u57df\u4ef7{current_region_price:.2f}\uff0c\u9884\u6d4b\u70b9\u4f4d{prediction.point_value:.2f}\u3002",
                "raw_context": raw_context,
            }
        )

    def _augment_scorecard_with_region_quality(self, scorecard: dict[str, Any], data_quality: dict[str, Any]) -> dict[str, Any]:
        result = dict(scorecard)
        result["groups"] = [
            group
            for group in list(result.get("groups") or [])
            if str(group.get("group_code") or "") not in {"regional_data_mapping", "regional_agent_missing"}
        ]
        quality = dict(result.get("data_quality") or {})
        quality["regional_available_count"] = data_quality.get("available_count")
        quality["regional_missing_count"] = data_quality.get("missing_count")
        quality["regional_missing_items"] = list(data_quality.get("missing_items") or [])
        result["data_quality"] = quality
        result["unresolved_items"] = [
            item
            for item in list(result.get("unresolved_items") or [])
            if str(item.get("group_code") or "") not in {"regional_data_mapping", "regional_agent_missing"}
        ]
        return result

    def _build_region_market_prediction_context(
        self,
        *,
        context: PredictionContext,
        config: RegionalSpreadConfig,
        as_of_date: date,
    ) -> tuple[PredictionContext, dict[str, Any]]:
        frame = context.feature_frame.copy()
        replacements: list[dict[str, Any]] = []
        missing: list[dict[str, Any]] = []
        price_target = "sd_diesel0_market" if config.product_code == "DIESEL_0_SPREAD" else "sd_gas92_market"
        self._copy_series_for_region(frame, source=config.price_column, target=price_target, replacements=replacements, missing=missing, label="\u533a\u57df\u5e02\u573a\u4ef7")
        region_prefix = self._region_feature_prefix(config.region_code)
        cdu_col = f"{region_prefix}_cdu_utilization_weekly" if region_prefix else None
        if cdu_col:
            self._copy_series_for_region(frame, source=cdu_col, target="shandong_cdu_utilization_weekly", replacements=replacements, missing=missing, label="\u533a\u57df\u5e38\u51cf\u538b\u5f00\u5de5\u7387")
            if cdu_col in frame.columns:
                frame["shandong_cdu_utilization_percentile_weekly"] = self._expanding_percentile_since_frame(frame, frame[cdu_col])
                frame["shandong_cdu_utilization_percentile_monthly"] = frame["shandong_cdu_utilization_percentile_weekly"]
                frame["regional_cdu_utilization_weekly"] = frame[cdu_col]
                frame["regional_cdu_utilization_change_weekly"] = self._change_from_previous_observation_local(frame[cdu_col])
                frame["regional_cdu_utilization_percentile_weekly"] = frame["shandong_cdu_utilization_percentile_weekly"]
                replacements.append({"field": "shandong_cdu_utilization_percentile_weekly", "source": cdu_col, "label": "\u533a\u57df\u5f00\u5de5\u7387\u5206\u4f4d"})
        shipment_col = self._regional_shipments_column(config=config, region_prefix=region_prefix)
        if shipment_col:
            self._copy_series_for_region(frame, source=shipment_col, target="regional_shipments_weekly", replacements=replacements, missing=missing, label="\u533a\u57df\u51fa\u8d27\u91cf")
            if shipment_col in frame.columns:
                frame["regional_shipments_change_weekly"] = self._change_from_previous_observation_local(frame[shipment_col])
                replacements.append({"field": "regional_shipments_change_weekly", "source": shipment_col, "label": "\u533a\u57df\u51fa\u8d27\u91cf\u73af\u6bd4"})
        self._apply_region_inventory_features(
            frame=frame,
            config=config,
            as_of_date=as_of_date,
            replacements=replacements,
            missing=missing,
        )
        regional_plan = self._regional_maintenance_plan(config=config, as_of_date=as_of_date)
        metadata = dict(context.metadata or {})
        if regional_plan:
            metadata["oilchem_maintenance_plan"] = regional_plan
            replacements.append({"field": "oilchem_maintenance_plan", "source": regional_plan.get("source"), "label": "\u533a\u57df\u68c0\u4fee\u8ba1\u5212"})
        else:
            missing.append({"field": "oilchem_maintenance_plan", "reason": "\u672a\u627e\u5230\u76ee\u6807\u533a\u57df\u5730\u65b9/\u4e3b\u8425\u70bc\u5382\u68c0\u4fee\u8ba1\u5212"})
        current_frame = frame[frame["date"] <= as_of_date]
        current_row = current_frame.iloc[-1]
        data_quality = {
            "mode": "regionalized_shandong_market_logic",
            "region_code": config.region_code,
            "region_name": config.region_name,
            "replacements": replacements,
            "missing_items": missing,
            "available_count": len(replacements),
            "missing_count": len(missing),
            "note": "\u5c71\u4e1c\u4e13\u5c5e\u5b57\u6bb5\u4f18\u5148\u66ff\u6362\u4e3a\u76ee\u6807\u533a\u57df\u5bf9\u5e94\u6570\u636e\uff1b\u627e\u4e0d\u5230\u5219\u663e\u793a\u7f3a\u5931\uff0c\u4e0d\u7528\u5c71\u4e1c\u6570\u636e\u66ff\u4ee3\u3002",
        }
        metadata["regionalized_data_quality"] = data_quality
        return PredictionContext(
            feature_frame=frame,
            current_row=current_row,
            report_payload=context.report_payload,
            news_items=context.news_items,
            refined_news_items=context.refined_news_items,
            policy_items=context.policy_items,
            metadata=metadata,
        ), data_quality

    def _copy_series_for_region(self, frame: pd.DataFrame, *, source: str | None, target: str, replacements: list[dict[str, Any]], missing: list[dict[str, Any]], label: str) -> None:
        if source and source in frame.columns and pd.to_numeric(frame[source], errors="coerce").notna().any():
            frame[target] = frame[source]
            replacements.append({"field": target, "source": source, "label": label})
        else:
            missing.append({"field": target, "source": source, "reason": f"{label}\u7f3a\u5931"})

    def _region_feature_prefix(self, region_code: str) -> str | None:
        return {
            "EAST_CHINA": "east_china",
            "NORTH_CHINA": "north_china",
            "SOUTH_CHINA": "south_china",
            "CENTRAL_CHINA": "central_china",
            "NORTHWEST": "northwest",
            "SOUTHWEST": "southwest",
            "NORTHEAST": "northeast",
        }.get(region_code)

    def _regional_shipments_column(self, *, config: RegionalSpreadConfig, region_prefix: str | None) -> str | None:
        if not region_prefix:
            return None
        product_prefix = "diesel" if config.product_code == "DIESEL_0_SPREAD" else "gasoline"
        return f"{region_prefix}_{product_prefix}_shipments_weekly"

    def _expanding_percentile_since_frame(self, frame: pd.DataFrame, series: pd.Series, *, min_periods: int = 5) -> pd.Series:
        source = pd.to_numeric(series, errors="coerce")
        result = pd.Series(np.nan, index=frame.index, dtype="float64")
        dates = pd.to_datetime(frame["date"], errors="coerce").dt.date
        mask = dates >= date(2024, 1, 1)
        if not bool(mask.any()):
            return result
        values = []
        for idx, value in source.loc[mask].items():
            if pd.isna(value):
                result.loc[idx] = np.nan
                continue
            values.append(float(value))
            if len(values) >= min_periods:
                result.loc[idx] = 100.0 * sum(item <= float(value) for item in values) / len(values)
        return result

    def _change_from_previous_observation_local(self, series: pd.Series) -> pd.Series:
        numeric = pd.to_numeric(series, errors="coerce")
        observed_mask = numeric.notna() & numeric.ne(numeric.shift())
        observed = numeric.where(observed_mask)
        previous_observed = observed.ffill().shift()
        return (observed - previous_observed).where(observed_mask)

    def _apply_region_inventory_features(self, *, frame: pd.DataFrame, config: RegionalSpreadConfig, as_of_date: date, replacements: list[dict[str, Any]], missing: list[dict[str, Any]]) -> None:
        rows = self._load_regional_inventory_rows(as_of_date=as_of_date)
        product = "\u67f4\u6cb9" if config.product_code == "DIESEL_0_SPREAD" else "\u6c7d\u6cb9"
        region_token = f"{config.region_name}\u5730\u533a"
        if config.region_name == "\u4e1c\u5317":
            region_token = "东北地区"
        series_by_project: dict[int, pd.Series] = {}
        for project_id in self._regional_inventory_project_ids(product=product):
            series = self._regional_inventory_series(rows=rows, project_id=project_id, region_token=region_token)
            if series is not None:
                series_by_project[project_id] = series.reindex(pd.to_datetime(frame["date"]).dt.date).ffill().astype(float)
        main_ids = (12891,) if product == "柴油" else (12887,)
        trader_ids = (12981, 12945) if product == "柴油" else (12975, 12944)
        main = next((series_by_project[pid] for pid in main_ids if pid in series_by_project), None)
        trader = next((series_by_project[pid] for pid in trader_ids if pid in series_by_project), None)
        components = []
        if main is not None:
            frame["shandong_main_company_inventory"] = main.values
            frame["shandong_main_company_inventory_change_weekly"] = self._change_from_previous_observation_local(frame["shandong_main_company_inventory"])
            frame["shandong_main_company_inventory_percentile_monthly"] = self._expanding_percentile_since_frame(frame, frame["shandong_main_company_inventory"])
            replacements.append({"field": "shandong_main_company_inventory", "source": "oilchem regional main inventory", "label": "\u533a\u57df\u4e3b\u8425\u5e93\u5b58"})
            components.append(frame["shandong_main_company_inventory"])
        else:
            missing.append({"field": "shandong_main_company_inventory", "reason": "\u672a\u627e\u5230\u76ee\u6807\u533a\u57df\u4e3b\u8425\u5e93\u5b58"})
        if trader is not None:
            frame["shandong_trade_company_inventory"] = trader.values
            replacements.append({"field": "shandong_trade_company_inventory", "source": "oilchem regional trader inventory", "label": "\u533a\u57df\u8d38\u6613\u5546\u5e93\u5b58"})
            components.append(frame["shandong_trade_company_inventory"])
        else:
            missing.append({"field": "shandong_trade_company_inventory", "reason": "\u672a\u627e\u5230\u76ee\u6807\u533a\u57df\u8d38\u6613\u5546\u5e93\u5b58"})
        independent = self._regional_refinery_inventory_series(frame=frame, config=config, as_of_date=as_of_date)
        if independent is not None:
            frame["shandong_independent_refinery_inventory"] = independent.values
            frame["refinery_inventory_monthly"] = independent.values
            frame["shandong_refinery_inventory_change_weekly"] = self._change_from_previous_observation_local(frame["shandong_independent_refinery_inventory"])
            frame["shandong_refinery_inventory_percentile_monthly"] = self._expanding_percentile_since_frame(frame, frame["shandong_independent_refinery_inventory"])
            if config.product_code == "DIESEL_0_SPREAD":
                frame["shandong_diesel_inventory_change_mom"] = frame["shandong_refinery_inventory_change_weekly"]
                frame["shandong_diesel_refinery_inventory_percentile_monthly"] = frame["shandong_refinery_inventory_percentile_monthly"]
            else:
                frame["shandong_gasoline_inventory_change_mom"] = frame["shandong_refinery_inventory_change_weekly"]
            replacements.append({"field": "shandong_independent_refinery_inventory", "source": "ganglian regional refinery inventory", "label": "\u533a\u57df\u72ec\u7acb\u70bc\u5382\u5382\u5185\u5e93\u5b58"})
            components.append(frame["shandong_independent_refinery_inventory"])
        else:
            independent_missing_reason = "\u672a\u627e\u5230\u76ee\u6807\u533a\u57df\u5382\u5185\u5e93\u5b58\uff1a\u72ec\u7acb\u70bc\u5382\uff08\u5468\uff09\uff0c\u672c\u9879\u4e0d\u7528\u5c71\u4e1c\u6216\u4e3b\u8425\u5e93\u5b58\u66ff\u4ee3"
            for field in (
                "shandong_independent_refinery_inventory",
                "shandong_refinery_inventory_change_weekly",
                "shandong_refinery_inventory_percentile_monthly",
                "shandong_gasoline_inventory_change_mom",
                "shandong_diesel_inventory_change_mom",
                "refinery_inventory_monthly",
            ):
                frame[field] = np.nan
            missing.append({"field": "shandong_independent_refinery_inventory", "reason": independent_missing_reason})
        if components:
            total = pd.concat(components, axis=1).sum(axis=1, min_count=1)
            frame["shandong_product_inventory_total_formal"] = total
            frame["shandong_product_inventory_change_weekly"] = self._change_from_previous_observation_local(total)
            frame["shandong_product_inventory_percentile_weekly"] = self._expanding_percentile_since_frame(frame, total)
            replacements.append({"field": "shandong_product_inventory_percentile_weekly", "source": "regional available inventory sum", "label": "\u533a\u57df\u53ef\u7528\u5e93\u5b58\u5408\u8ba1\u5206\u4f4d"})

    def _regional_refinery_inventory_series(self, *, frame: pd.DataFrame, config: RegionalSpreadConfig, as_of_date: date) -> pd.Series | None:
        product_key = "DIESEL_0" if config.product_code == "DIESEL_0_SPREAD" else "GASOLINE_92"
        indicator_code = REGIONAL_REFINERY_INVENTORY_INDICATORS.get(product_key, {}).get(config.region_code)
        if not indicator_code or not self.snapshot_repository:
            return None
        try:
            rows = self.snapshot_repository.load_market_timeseries_values(
                source_code="ganglian_excel_import",
                indicator_codes=[indicator_code],
                start_date=date(2024, 1, 1),
                end_date=as_of_date,
            )
        except Exception:
            return None
        values: dict[date, float] = {}
        for row in rows:
            row_date = row.get("dt")
            value = row.get("value_num")
            if row_date is None or value is None:
                continue
            if not isinstance(row_date, date):
                try:
                    row_date = pd.Timestamp(row_date).date()
                except Exception:
                    continue
            values[row_date] = float(value)
        if not values:
            return None
        series = pd.Series(values).sort_index()
        frame_dates = pd.to_datetime(frame["date"], errors="coerce").dt.date
        aligned = series.reindex(frame_dates).ffill()
        if not aligned.notna().any():
            return None
        return pd.Series(aligned.to_numpy(dtype="float64"), index=frame.index)

    def _regional_inventory_project_ids(self, *, product: str) -> tuple[int, ...]:
        return (12891, 12981, 12945) if product == "\u67f4\u6cb9" else (12887, 12975, 12944)

    def _regional_inventory_series(self, *, rows: list[dict[str, Any]], project_id: int, region_token: str) -> pd.Series | None:
        values: dict[date, float] = {}
        for row in rows:
            if row.get("project_quota_id") != project_id:
                continue
            if region_token not in str(row.get("region") or ""):
                continue
            parsed = self._parse_date_text(row.get("date"))
            value = self._float_or_none(row.get("value"))
            if parsed is not None and value is not None:
                values[parsed] = value
        if not values:
            return None
        return pd.Series(values).sort_index()

    def _regional_maintenance_plan(self, *, config: RegionalSpreadConfig, as_of_date: date) -> dict[str, Any] | None:
        try:
            records = self.dataset_service._load_archived_oilchem_records(
                source_codes=["oilchem_local_refinery_maintenance_plan", "oilchem_main_refinery_maintenance_plan"],
                end_date=as_of_date,
                limit_per_source=10,
            )
            return self.dataset_service._aggregate_maintenance_plan(records=records, as_of_date=as_of_date, target_region=config.region_name)
        except Exception:
            return None

    def _config_map(self, product_code: str = "GASOLINE_92") -> dict[str, RegionalSpreadConfig]:
        normalized = (product_code or "GASOLINE_92").strip().upper()
        if normalized in {"DIESEL_0", "DIESEL_0_SPREAD", "DIESEL0"}:
            return DIESEL_REGIONAL_SPREAD_CONFIGS
        return REGIONAL_SPREAD_CONFIGS

    def _resolve_config(self, region_code: str, product_code: str = "GASOLINE_92") -> RegionalSpreadConfig:
        normalized = region_code.strip().upper()
        config_map = self._config_map(product_code)
        if normalized not in config_map:
            supported = ", ".join(sorted(config_map))
            raise ValueError(f"Unsupported region_code={region_code}. Supported: {supported}")
        return config_map[normalized]

    def _resolve_configs(self, region_codes: list[str] | None, product_code: str = "GASOLINE_92") -> list[RegionalSpreadConfig]:
        config_map = self._config_map(product_code)
        if not region_codes:
            return list(config_map.values())
        return [self._resolve_config(region_code, product_code=product_code) for region_code in region_codes]

    def list_freight_settings(self) -> list[dict[str, Any]]:
        return self._component_settings_grouped()

    def update_freight_setting(self, *, region_code: str, freight_value: float, updated_by: str | None) -> dict[str, Any]:
        config = self._resolve_config(region_code)
        components = [item for item in REGIONAL_FREIGHT_COMPONENT_CONFIGS if item["region_code"] == config.region_code]
        if not components:
            raise ValueError(f"未配置区域运费明细：{config.region_name}")
        per_component = float(freight_value)
        for component in components:
            self.update_freight_component(
                component_key=component["component_key"],
                freight_value=per_component,
                updated_by=updated_by,
            )
        return self._freight_setting_for(config)

    def update_freight_component(
        self,
        *,
        component_key: str,
        freight_value: float,
        updated_by: str | None = None,
    ) -> list[dict[str, Any]]:
        if self.snapshot_repository is None:
            raise RuntimeError("PostgreSQL 未配置，无法保存运费明细。")
        self.snapshot_repository.upsert_freight_component(
            defaults=REGIONAL_FREIGHT_COMPONENT_CONFIGS,
            component_key=component_key,
            freight_value=freight_value,
            updated_by=updated_by,
        )
        return self.list_freight_settings()

    def _component_settings_grouped(self) -> list[dict[str, Any]]:
        return build_freight_settings_from_components(self.snapshot_repository)

    def _freight_setting_for(self, config: RegionalSpreadConfig) -> dict[str, Any]:
        settings = {item["region_code"]: item for item in self.list_freight_settings()}
        return settings.get(
            config.region_code,
            {
                "region_code": config.region_code,
                "region_name": config.region_name,
                "freight_value": REGIONAL_FREIGHT_ESTIMATES.get(config.region_code, 80.0),
                "source_type": "default",
                "updated_by": None,
                "updated_at": None,
                "components": [],
                "calculation": "未配置运费明细，使用默认估算。",
            },
        )

    def _predict_from_context(
        self,
        context: PredictionContext,
        config: RegionalSpreadConfig,
        horizon_config: HorizonConfig,
        as_of_date: date,
        use_llm_explainer: bool,
        scenario_text: str | None,
        enable_refined_news: bool,
        enable_event_risk: bool,
        context_metadata: dict[str, Any] | None,
        freight_settings: dict[str, dict[str, Any]] | None = None,
        regional_cache: dict[tuple[str, float], dict[str, Any]] | None = None,
    ) -> PredictionResult:
        feature_frame = context.feature_frame.copy()
        current_frame = feature_frame[feature_frame["date"] <= as_of_date]
        if current_frame.empty:
            raise RuntimeError(f"No feature row found for as_of_date={as_of_date}")

        current_row = self._enrich_row(current_frame.iloc[-1], config)
        current_spread = current_row.get("target_region_spread")
        counter_region_price = current_row.get("target_region_price")
        if pd.isna(current_spread) or pd.isna(counter_region_price):
            raise RuntimeError(f"Missing region price history for {config.region_name} on or before {as_of_date}.")
        freight_setting = (freight_settings or {}).get(config.region_code) or self._freight_setting_for(config)
        freight_estimate = float(freight_setting.get("freight_value", REGIONAL_FREIGHT_ESTIMATES.get(config.region_code, 80.0)))
        regional_inventory = self._regional_inventory_context(config=config, as_of_date=as_of_date)
        if self.outright_predictor is not None:
            regionalized = self._try_region_market_price_prediction(
                context=context,
                config=config,
                as_of_date=as_of_date,
                horizon_config=horizon_config,
                use_llm_explainer=use_llm_explainer,
                scenario_text=scenario_text,
                enable_refined_news=enable_refined_news,
                enable_event_risk=enable_event_risk,
                context_metadata=context_metadata,
                current_row=current_row,
                current_region_price=float(counter_region_price),
                current_shandong_price=float(current_row[config.shandong_price_column]),
                current_spread_value=float(current_spread),
                freight_estimate=freight_estimate,
                freight_setting=freight_setting,
                regional_inventory=regional_inventory,
            )
            if regionalized is not None:
                return regionalized
        raise RuntimeError(
            "regionalized_shandong_market_logic_required: region price prediction must reuse Shandong market price logic; old regional spread fallback is disabled"
        )

        # Region price prediction must use the regionalized Shandong outright logic above.
        # The previous regional-spread fallback is intentionally disabled so no custom
        # spread/netback/inventory score can diverge from the Shandong scoring logic.

    def _enrich_row(self, row: pd.Series, config: RegionalSpreadConfig) -> pd.Series:
        enriched = row.copy()
        target_price = row.get(config.price_column)
        shandong_price = row.get(config.shandong_price_column)
        enriched["target_region_price"] = target_price
        if pd.notna(target_price) and pd.notna(shandong_price):
            enriched["target_region_spread"] = float(target_price) - float(shandong_price)
        else:
            enriched["target_region_spread"] = np.nan
        enriched["target_region_spread_change_1d"] = self._invert_number(row.get(config.spread_change_1d_column))
        enriched["target_region_spread_change_3d"] = self._invert_number(row.get(config.spread_change_3d_column))
        return enriched

    def _score_row(
        self,
        row: pd.Series,
        extra: dict[str, Any],
        excluded_agent_names: set[str] | None = None,
    ) -> tuple[list[AgentClaim], float]:
        claims: list[AgentClaim] = []
        total_score = 0.0
        controls = self.agent_control_service.get_runtime_controls(self.scope_key)
        excluded_agent_names = excluded_agent_names or set()
        for agent in self.agents:
            if agent.name in excluded_agent_names:
                continue
            claim = agent.analyze(row, extra)
            control = controls.get(agent.name, {"enabled": True, "weight": 1.0})
            raw_score = float(claim.numeric_signals.get("score", 0.0))
            enabled = bool(control.get("enabled", True))
            weight = float(control.get("weight", 1.0)) if enabled else 0.0
            weighted_score = raw_score * weight
            claim.numeric_signals = {
                **claim.numeric_signals,
                "score": round(weighted_score, 4),
                "raw_score": round(raw_score, 4),
                "weight": round(weight, 4),
                "weighted_score": round(weighted_score, 4),
            }
            claim.structured_payload = {
                **claim.structured_payload,
                "runtime_control": {
                    "scope_key": self.scope_key,
                    "enabled": enabled,
                    "weight": round(weight, 4),
                },
            }
            claims.append(claim)
            total_score += weighted_score
        return claims, round(total_score, 4)

    def _bounded_delta_from_score(self, score_value: float, *, horizon_config: HorizonConfig) -> float:
        horizon_multiplier = {
            "D1": 1.0,
            "D3": 1.25,
            "W1": 1.6,
            "M1": 2.2,
        }.get(horizon_config.code, 1.0)
        return round(float(score_value) * horizon_multiplier, 4)

    def _regional_baseline_range_half_width(
        self,
        *,
        rule_delta: dict[str, Any],
        regional_inventory: dict[str, Any],
        horizon_config: HorizonConfig,
        predicted_delta: float,
    ) -> float:
        base = {
            "D1": 14.0,
            "D3": 26.0,
            "W1": 40.0,
            "M1": 58.0,
        }.get(horizon_config.code, 14.0)
        contributions = [
            float(value)
            for value in (rule_delta.get("contributions") or {}).values()
            if abs(float(value)) > 0
        ]
        has_positive = any(value > 0 for value in contributions)
        has_negative = any(value < 0 for value in contributions)
        conflict_addon = 5.0 if has_positive and has_negative else 0.0
        inventory_addon = 5.0 if not regional_inventory.get("available") else 0.0
        weak_signal_addon = (
            4.0 if abs(float(predicted_delta or 0.0)) < self._regional_direction_threshold(horizon_config) else 0.0
        )
        return min(65.0, base + conflict_addon + inventory_addon + weak_signal_addon)


    def _build_regional_history_cache(
        self,
        *,
        frame: pd.DataFrame,
        config: RegionalSpreadConfig,
        as_of_date: date,
        freight_estimate: float,
    ) -> dict[str, Any]:
        rows: list[dict[str, Any]] = []
        sorted_frame = frame[frame["date"] <= as_of_date].sort_values("date").reset_index(drop=True)
        for _, source_row in sorted_frame.iterrows():
            source_date = self._parse_date_text(source_row.get("date"))
            if source_date is None or source_date > as_of_date:
                continue
            enriched = self._enrich_row(source_row, config)
            spread = self._float_or_none(enriched.get("target_region_spread"))
            if spread is None:
                continue
            rows.append(
                {
                    "date": source_date,
                    "spread": spread,
                    "change_3d": self._float_or_none(enriched.get("target_region_spread_change_3d")) or 0.0,
                    "netback": spread - freight_estimate,
                }
            )
        return {
            "rows": rows,
            "spreads": [item["spread"] for item in rows],
        }

    def _regional_state_table_delta(
        self,
        *,
        frame: pd.DataFrame,
        config: RegionalSpreadConfig,
        current_row: pd.Series,
        as_of_date: date,
        freight_estimate: float,
        horizon_config: HorizonConfig,
        regional_cache: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        cache_rows = list((regional_cache or {}).get("rows") or [])
        history_rows: list[dict[str, Any]] = []
        if cache_rows:
            for idx in range(len(cache_rows) - horizon_config.steps):
                source_item = cache_rows[idx]
                target_item = cache_rows[idx + horizon_config.steps]
                if target_item["date"] > as_of_date:
                    continue
                source_spread = source_item.get("spread")
                target_spread = target_item.get("spread")
                if source_spread is None or target_spread is None:
                    continue
                history_rows.append(
                    {
                        "date": source_item["date"],
                        "spread": source_spread,
                        "change_3d": source_item.get("change_3d") or 0.0,
                        "netback": source_item.get("netback", source_spread - freight_estimate),
                        "future_delta": target_spread - source_spread,
                    }
                )
        else:
            sorted_frame = frame[frame["date"] <= as_of_date].sort_values("date").reset_index(drop=True)
            for idx in range(len(sorted_frame) - horizon_config.steps):
                source_row = sorted_frame.iloc[idx]
                target_row = sorted_frame.iloc[idx + horizon_config.steps]
                source_date = self._parse_date_text(source_row.get("date"))
                target_date = self._parse_date_text(target_row.get("date"))
                if source_date is None or target_date is None or target_date > as_of_date:
                    continue
                enriched_source = self._enrich_row(source_row, config)
                enriched_target = self._enrich_row(target_row, config)
                source_spread = self._float_or_none(enriched_source.get("target_region_spread"))
                target_spread = self._float_or_none(enriched_target.get("target_region_spread"))
                if source_spread is None or target_spread is None:
                    continue
                history_rows.append(
                    {
                        "date": source_date,
                        "spread": source_spread,
                        "change_3d": self._float_or_none(enriched_source.get("target_region_spread_change_3d")) or 0.0,
                        "netback": source_spread - freight_estimate,
                        "future_delta": target_spread - source_spread,
                    }
                )
        if len(history_rows) < REGIONAL_STATE_MIN_SAMPLE:
            return {
                "status": "fallback",
                "reason": f"有效历史状态样本{len(history_rows)}条，低于{REGIONAL_STATE_MIN_SAMPLE}条",
                "sample_size": len(history_rows),
                "median_delta": 0.0,
            }

        spreads = np.array([item["spread"] for item in history_rows], dtype=float)
        quantiles = {
            "q10": float(np.quantile(spreads, 0.10)),
            "q30": float(np.quantile(spreads, 0.30)),
            "q70": float(np.quantile(spreads, 0.70)),
            "q90": float(np.quantile(spreads, 0.90)),
        }
        current_state = self._regional_state_key(
            spread=float(current_row["target_region_spread"]),
            change_3d=self._float_or_none(current_row.get("target_region_spread_change_3d")) or 0.0,
            netback=float(current_row["target_region_spread"]) - freight_estimate,
            quantiles=quantiles,
        )
        for level in ("full", "position_momentum", "position"):
            matched = [
                item["future_delta"]
                for item in history_rows
                if self._state_matches(
                    current_state,
                    self._regional_state_key(
                        spread=item["spread"],
                        change_3d=item["change_3d"],
                        netback=item["netback"],
                        quantiles=quantiles,
                    ),
                    level=level,
                )
            ]
            if len(matched) >= REGIONAL_STATE_MIN_SAMPLE:
                lower = float(np.quantile(matched, 0.20))
                upper = float(np.quantile(matched, 0.80))
                return {
                    "status": "enabled",
                    "match_level": level,
                    "state_key": current_state,
                    "sample_size": len(matched),
                    "median_delta": round(float(np.median(matched)), 4),
                    "p20_delta": round(lower, 4),
                    "p80_delta": round(upper, 4),
                    "quantiles": {key: round(value, 4) for key, value in quantiles.items()},
                    "reason": "使用预测日前已完成样本的同状态价差变化中位数",
                }
        return {
            "status": "fallback",
            "reason": f"状态{current_state}匹配样本不足，降级为当日规则修正",
            "state_key": current_state,
            "sample_size": len(history_rows),
            "median_delta": 0.0,
            "quantiles": {key: round(value, 4) for key, value in quantiles.items()},
        }

    def _regional_state_key(
        self,
        *,
        spread: float,
        change_3d: float,
        netback: float,
        quantiles: dict[str, float],
    ) -> dict[str, str]:
        if spread <= quantiles["q10"]:
            position = "extreme_low"
        elif spread <= quantiles["q30"]:
            position = "low"
        elif spread >= quantiles["q90"]:
            position = "extreme_high"
        elif spread >= quantiles["q70"]:
            position = "high"
        else:
            position = "normal"
        if change_3d >= 2.0:
            momentum = "expanding"
        elif change_3d <= -2.0:
            momentum = "narrowing"
        else:
            momentum = "stable"
        if netback >= 80.0:
            netback_state = "strong_open"
        elif netback >= 30.0:
            netback_state = "open"
        elif netback >= 0.0:
            netback_state = "weak_open"
        else:
            netback_state = "closed"
        return {"position": position, "momentum": momentum, "netback": netback_state}

    def _state_matches(self, current: dict[str, str], candidate: dict[str, str], *, level: str) -> bool:
        if level == "full":
            return current == candidate
        if level == "position_momentum":
            return current["position"] == candidate["position"] and current["momentum"] == candidate["momentum"]
        if level == "position":
            return current["position"] == candidate["position"]
        return False

    def _regional_operating_bounds(
        self,
        *,
        frame: pd.DataFrame,
        config: RegionalSpreadConfig,
        as_of_date: date,
        freight_estimate: float,
        regional_cache: dict[str, Any] | None = None,
    ) -> dict[str, float]:
        cached_spreads = (regional_cache or {}).get("spreads")
        if cached_spreads is not None:
            spreads = list(cached_spreads)
        else:
            spreads: list[float] = []
            for _, source_row in frame[frame["date"] <= as_of_date].iterrows():
                enriched = self._enrich_row(source_row, config)
                spread = self._float_or_none(enriched.get("target_region_spread"))
                if spread is not None:
                    spreads.append(spread)
        if len(spreads) < 20:
            lower = -0.35 * freight_estimate - REGIONAL_RISK_BUFFER
            upper = freight_estimate + REGIONAL_REQUIRED_MARGIN + REGIONAL_RISK_BUFFER
            return {"lower": round(lower, 4), "upper": round(upper, 4), "basis": "route_cost_fallback"}
        values = np.array(spreads, dtype=float)
        median = float(np.median(values))
        robust_sigma = 1.4826 * float(np.median(np.abs(values - median)))
        distribution_lower = float(np.quantile(values, 0.05)) - 0.5 * robust_sigma
        distribution_upper = float(np.quantile(values, 0.95)) + 0.5 * robust_sigma
        route_lower = -0.35 * freight_estimate - REGIONAL_RISK_BUFFER
        route_upper = freight_estimate + REGIONAL_REQUIRED_MARGIN + REGIONAL_RISK_BUFFER
        lower = max(distribution_lower, route_lower)
        upper = min(distribution_upper, route_upper) if distribution_upper > lower else route_upper
        return {
            "lower": round(lower, 4),
            "upper": round(upper, 4),
            "basis": "distribution_and_route_cost",
            "median": round(median, 4),
            "robust_sigma": round(robust_sigma, 4),
            "distribution_lower": round(distribution_lower, 4),
            "distribution_upper": round(distribution_upper, 4),
            "route_lower": round(route_lower, 4),
            "route_upper": round(route_upper, 4),
        }

    def _regional_expert_range_half_width(
        self,
        *,
        expert_delta: dict[str, Any],
        regional_inventory: dict[str, Any],
        horizon_config: HorizonConfig,
    ) -> float:
        base = {
            "D1": 12.0,
            "D3": 22.0,
            "W1": 35.0,
            "M1": 50.0,
        }.get(horizon_config.code, 12.0)
        rule_delta = expert_delta.get("rule_delta") or expert_delta
        contributions = [
            float(value)
            for value in (rule_delta.get("contributions") or {}).values()
            if abs(float(value)) > 0
        ]
        has_positive = any(value > 0 for value in contributions)
        has_negative = any(value < 0 for value in contributions)
        conflict_addon = 4.0 if has_positive and has_negative else 0.0
        inventory_addon = 4.0 if not regional_inventory.get("available") else 0.0
        state_addon = 0.0 if (expert_delta.get("state_delta") or {}).get("status") == "enabled" else 6.0
        weak_signal_addon = 3.0 if abs(float(expert_delta.get("predicted_delta") or 0.0)) < self._regional_direction_threshold(horizon_config) else 0.0
        return min(50.0, base + conflict_addon + inventory_addon + state_addon + weak_signal_addon)

    def _regional_direction_threshold(self, horizon_config: HorizonConfig) -> float:
        return {
            "D1": 0.5,
            "D3": 1.0,
            "W1": 1.5,
            "M1": 2.5,
        }.get(horizon_config.code, 0.5)

    def _clip(self, value: float, lower: float, upper: float) -> float:
        return max(lower, min(upper, float(value)))

    def _regional_inventory_context(self, *, config: RegionalSpreadConfig, as_of_date: date) -> dict[str, Any]:
        cache_key = (config.region_code, as_of_date)
        cached = self._regional_inventory_context_cache.get(cache_key)
        if cached is not None:
            return cached

        rows = self._load_regional_inventory_rows(as_of_date=as_of_date)
        if not rows:
            result = {"available": False, "reason": "未找到区域库存归档数据"}
            self._regional_inventory_context_cache[cache_key] = result
            return result

        region_token = config.region_name
        selected_project_id = None
        selected_rows: list[dict[str, Any]] = []
        for project_id in (12975, 12944):
            candidates = [
                row
                for row in rows
                if row.get("project_quota_id") == project_id
                and "汽油" in str(row.get("product") or row.get("project_label") or "")
            ]
            if candidates:
                selected_project_id = project_id
                selected_rows = candidates
                break
        if not selected_rows or selected_project_id is None:
            result = {"available": False, "reason": "未找到汽油贸易商区域库存数据"}
            self._regional_inventory_context_cache[cache_key] = result
            return result

        latest_date = max((self._parse_date_text(row.get("date")) for row in selected_rows), default=None)
        if latest_date is None:
            result = {"available": False, "reason": "区域库存日期缺失"}
            self._regional_inventory_context_cache[cache_key] = result
            return result

        latest_rows = [row for row in selected_rows if self._parse_date_text(row.get("date")) == latest_date]
        target_rows = [row for row in latest_rows if region_token in str(row.get("region") or "")]
        if not target_rows:
            result = {"available": False, "reason": f"未找到{config.region_name}库存样本"}
            self._regional_inventory_context_cache[cache_key] = result
            return result

        target_value = self._float_or_none(target_rows[0].get("value"))
        peer_values = [self._float_or_none(row.get("value")) for row in latest_rows]
        peer_values = [value for value in peer_values if value is not None]
        if target_value is None or not peer_values:
            result = {"available": False, "reason": f"{config.region_name}库存数值缺失"}
            self._regional_inventory_context_cache[cache_key] = result
            return result

        prior_dates = sorted(
            {
                parsed
                for row in selected_rows
                for parsed in [self._parse_date_text(row.get("date"))]
                if parsed is not None and parsed < latest_date
            },
            reverse=True,
        )
        prior_value = None
        if prior_dates:
            prior_date = prior_dates[0]
            prior_target = [
                row
                for row in selected_rows
                if self._parse_date_text(row.get("date")) == prior_date and region_token in str(row.get("region") or "")
            ]
            if prior_target:
                prior_value = self._float_or_none(prior_target[0].get("value"))
        median_value = float(np.median(peer_values))
        wow_change = target_value - prior_value if prior_value is not None else None
        stale_days = (as_of_date - latest_date).days
        subject = "贸易商" if selected_project_id == 12975 else "部分社会油库"
        score_ready = stale_days == 1 and wow_change is not None and abs(float(wow_change)) >= 1e-9
        result = {
            "available": stale_days <= 21,
            "score_ready": bool(score_ready),
            "score_note": None if score_ready else "沿用上一期周度库存或本期无变化，本次不计分",
            "reason": None if stale_days <= 21 else f"区域库存最新日期{latest_date.isoformat()}，距预测日超过21天",
            "project_quota_id": selected_project_id,
            "subject": subject,
            "latest_date": latest_date.isoformat(),
            "target_region": config.region_name,
            "target_value": round(target_value, 4),
            "peer_median": round(median_value, 4),
            "ratio_to_median": round(target_value / median_value, 4) if median_value else None,
            "wow_change": round(wow_change, 4) if wow_change is not None else None,
            "unit": target_rows[0].get("unit") or "万吨",
        }
        self._regional_inventory_context_cache[cache_key] = result
        return result

    def _load_regional_inventory_rows(self, *, as_of_date: date) -> list[dict[str, Any]]:
        cached = self._regional_inventory_cache.get(as_of_date)
        if cached is not None:
            return cached
        try:
            payload = self.dataset_service.get_oilchem_openapi_inventory(
                start_date=as_of_date - timedelta(days=90),
                end_date=as_of_date,
            )
            rows = list(payload.get("items") or [])
        except Exception:
            rows = []
        self._regional_inventory_cache[as_of_date] = rows
        return rows

    def _parse_date_text(self, value: Any) -> date | None:
        if isinstance(value, date):
            return value
        try:
            text = str(value or "").strip()
            if not text:
                return None
            return datetime.strptime(text[:10], "%Y-%m-%d").date()
        except Exception:
            return None

    def _float_or_none(self, value: Any) -> float | None:
        try:
            if value is None or pd.isna(value):
                return None
            return float(value)
        except Exception:
            return None

    def _build_explanation(
        self,
        *,
        config: RegionalSpreadConfig,
        claims: list[AgentClaim],
        current_spread: float,
        score_value: float,
        point_value: float,
        range_lower: float,
        range_upper: float,
        direction_label: str,
        horizon_config: HorizonConfig,
    ) -> str:
        direction_text = {"up": "走扩", "down": "收敛", "flat": "震荡"}[direction_label]
        top_claims = sorted(
            claims,
            key=lambda item: abs(float(item.numeric_signals.get("weighted_score", 0.0))),
            reverse=True,
        )[:3]
        details = "；".join(item.evidence[0] if item.evidence else item.summary for item in top_claims if item.summary)
        return (
            f"{config.region_name}-山东92#价差 {horizon_config.code} 判断为{direction_text}，当前价差 {current_spread:.2f}，"
            f"综合分 {score_value:.2f}，预测点位 {point_value:.2f}，区间 {range_lower:.2f}~{range_upper:.2f}。"
            + (f" 主要依据包括：{details}。" if details else "")
        )

    def _invert_number(self, value: Any) -> float | None:
        try:
            if value is None or pd.isna(value):
                return None
            return -float(value)
        except Exception:
            return None

    def _direction_from_delta(self, predicted_delta: float, threshold: float = 2.0) -> str:
        if predicted_delta > threshold:
            return "up"
        if predicted_delta < -threshold:
            return "down"
        return "flat"

    def _probabilities_from_score(self, score_value: float) -> dict[str, float]:
        up_raw = max(score_value, 0.0) + 10.0
        down_raw = max(-score_value, 0.0) + 10.0
        flat_raw = max(5.0, 50.0 - abs(score_value))
        total = up_raw + down_raw + flat_raw
        return {
            "up": round(up_raw / total, 4),
            "flat": round(flat_raw / total, 4),
            "down": round(down_raw / total, 4),
        }

    def _freight_review_required(self, freight_setting: dict[str, Any]) -> bool:
        if freight_setting.get("source_type") != "manual":
            return True
        updated_at = freight_setting.get("updated_at")
        if not updated_at:
            return True
        if isinstance(updated_at, datetime):
            updated_dt = updated_at
        else:
            try:
                updated_dt = datetime.fromisoformat(str(updated_at).replace("Z", "+00:00"))
            except ValueError:
                return True
        if updated_dt.tzinfo is None:
            updated_dt = updated_dt.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - updated_dt.astimezone(timezone.utc)).total_seconds() > 24 * 3600

    def _trade_action_from_netback_spread(
        self,
        *,
        netback_spread: float,
        freight_review_required: bool,
    ) -> dict[str, Any]:
        if freight_review_required:
            review_note = "运费超过24小时未确认或仍为默认估算，执行前必须复核。"
        else:
            review_note = "运费已人工确认。"
        if netback_spread > 60.0:
            return {
                "level": "expand",
                "label": "扩大外发",
                "action": "连续两次刷新保持为正后，可扩大外发。",
                "trigger": "预测净回款价差 > 60 元/吨",
                "review_note": review_note,
            }
        if netback_spread >= 30.0:
            return {
                "level": "trial",
                "label": "小批量试发",
                "action": "成交活跃且客户账期可控时，小批量试发。",
                "trigger": "预测净回款价差 30-60 元/吨",
                "review_note": review_note,
            }
        if netback_spread >= 0.0:
            return {
                "level": "maintenance",
                "label": "客户维护",
                "action": "仅做客户维护或锁价订单，不主动放量。",
                "trigger": "预测净回款价差 0-30 元/吨",
                "review_note": review_note,
            }
        return {
            "level": "stop",
            "label": "原则停发",
            "action": "原则上停止跨区外发，优先本地或近距离成交。",
            "trigger": "预测净回款价差 < 0 元/吨",
            "review_note": review_note,
        }

    def _round_or_none(self, value: Any) -> float | None:
        if value is None or pd.isna(value):
            return None
        return round(float(value), 4)

    def _build_input_hash(self, payload: dict[str, Any]) -> str:
        serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
        return hashlib.sha256(serialized.encode("utf-8")).hexdigest()

