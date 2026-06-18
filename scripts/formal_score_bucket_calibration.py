from __future__ import annotations

import json
import math
import platform
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Any

platform._wmi_query = lambda *a, **k: ['10.0.0', '1', 'Multiprocessor Free', '0', '0']
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.container import get_dataset_service, get_predictor
from app.services.predictors.horizons import DEFAULT_HORIZONS, resolve_horizon_config

START_DATE = date(2024, 1, 1)
AS_OF = date(2026, 6, 12)
OUTPUT_JSON = Path('artifacts/formal_score_bucket_calibration_20240101_20260612.json')
OUTPUT_XLSX = Path('outputs/formal_score_bucket_calibration_20240101_20260612.xlsx')


def q(values: list[float], quantile: float) -> float | None:
    values = [float(value) for value in values if value == value]
    if not values:
        return None
    return round(float(np.quantile(values, quantile)), 2)


def up_rate(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(1 for value in values if value > 0) / len(values), 3)


def evaluate_bucket(predictor: Any, scored: pd.DataFrame, score_column: str) -> list[dict[str, Any]]:
    bucket_defs = predictor._score_bucket_defs(score_column)
    scored = scored.copy()
    scored['_score_points'] = scored[score_column].map(
        lambda value: predictor._score_points(score_column=score_column, score_value=float(value))
    )
    scored['_bucket_index'] = scored['_score_points'].map(
        lambda value: predictor._score_bucket_index(float(value), bucket_defs=bucket_defs)
    )
    rows: list[dict[str, Any]] = []
    for index, bucket in enumerate(bucket_defs):
        bucket_rows = scored[scored['_bucket_index'] == index]
        deltas = bucket_rows['target_delta'].astype(float).tolist()
        rows.append(
            {
                '状态桶': bucket['label'],
                '分数区间': bucket['range_label'],
                '样本数': len(deltas),
                '历史P25': q(deltas, 0.25),
                '历史P50': q(deltas, 0.5),
                '历史P75': q(deltas, 0.75),
                '上涨概率': up_rate(deltas),
                '是否样本充足': len(deltas) >= 12,
                '业务解释': explain_bucket(bucket['label'], deltas),
            }
        )
    return rows


def explain_bucket(label: str, deltas: list[float]) -> str:
    if not deltas:
        return '该桶当前无历史样本，不能单独解释，系统应合并相邻桶或提示谨慎。'
    median = q(deltas, 0.5)
    rate = up_rate(deltas)
    if len(deltas) < 12:
        return f'样本{len(deltas)}条，低于12条最低解释门槛；只作参考，需合并相邻桶。'
    return f'历史同桶样本{len(deltas)}条，中位涨跌{median}元/吨，上涨概率{rate}；用于说明该分数段通常对应的价格变化。'


def build() -> dict[str, Any]:
    dataset_service = get_dataset_service()
    predictor = get_predictor()
    # Need future target price after AS_OF for W1/M1 target shift, but calibration samples only start from 2024-01-01.
    frame_start = START_DATE - timedelta(days=180)
    frame_end = AS_OF
    frame = dataset_service.build_feature_frame(start_date=frame_start, end_date=frame_end).sort_values('date').copy()
    frame['date'] = pd.to_datetime(frame['date'])
    frame = frame[frame['date'] >= pd.Timestamp(START_DATE)].copy()
    summary: dict[str, Any] = {
        'as_of': str(AS_OF),
        'start_date': str(START_DATE),
        'frame_rows': int(len(frame)),
        'date_min': str(frame['date'].min().date()) if not frame.empty else None,
        'date_max': str(frame['date'].max().date()) if not frame.empty else None,
        'horizons': {},
    }
    for horizon in DEFAULT_HORIZONS:
        horizon_config = resolve_horizon_config(horizon)
        work = frame.copy()
        work['target_date'] = work['date'].shift(-horizon_config.steps)
        work['target_price'] = work['sd_gas92_market'].shift(-horizon_config.steps)
        work['target_delta'] = work['target_price'] - work['sd_gas92_market']
        history = work[(work['date'] < pd.Timestamp(AS_OF)) & (work['target_date'] <= pd.Timestamp(AS_OF))]
        history = history.dropna(subset=['target_delta']).copy()
        scored = predictor.score_frame_for_backtest(
            history,
            enable_refined_news=True,
            enable_event_risk=True,
            horizon=horizon,
        ).dropna(subset=['agent_score', 'business_scorecard_score', 'target_delta'])
        summary['horizons'][horizon] = {
            '样本数': int(len(scored)),
            '智能体状态桶': evaluate_bucket(predictor, scored, 'agent_score'),
            '业务状态桶': evaluate_bucket(predictor, scored, 'business_scorecard_score'),
        }
    return summary


def write_excel(summary: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = '正式状态桶校准'
    ws.append(['说明', f"本表用{summary['start_date']}至{summary['as_of']}的历史特征回放，校准智能体和业务状态桶。样本数少于12条的桶不能单独作为强结论，应合并相邻桶或提示谨慎。"])
    ws.append([])
    ws.append(['模型', '周期', '状态桶', '分数区间', '样本数', '历史P25', '历史P50', '历史P75', '上涨概率', '是否样本充足', '业务解释'])
    for horizon, payload in summary['horizons'].items():
        for model_name, key in [('智能体', '智能体状态桶'), ('业务', '业务状态桶')]:
            for item in payload[key]:
                ws.append([
                    model_name,
                    horizon,
                    item['状态桶'],
                    item['分数区间'],
                    item['样本数'],
                    item['历史P25'],
                    item['历史P50'],
                    item['历史P75'],
                    item['上涨概率'],
                    '是' if item['是否样本充足'] else '否',
                    item['业务解释'],
                ])
    thin = Side(style='thin', color='D9E2EC')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(name='微软雅黑', size=10)
            if cell.row in {1, 3}:
                cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF' if cell.row == 3 else '000000')
                if cell.row == 3:
                    cell.fill = header_fill
    widths = [10, 8, 10, 18, 10, 12, 12, 12, 12, 14, 56]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = 'A4'
    wb.save(OUTPUT_XLSX)


if __name__ == '__main__':
    summary = build()
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    write_excel(summary)
    print(json.dumps({
        'json': str(OUTPUT_JSON),
        'xlsx': str(OUTPUT_XLSX),
        'start_date': summary['start_date'],
        'as_of': summary['as_of'],
        'frame_rows': summary['frame_rows'],
        'horizon_samples': {key: value['样本数'] for key, value in summary['horizons'].items()},
    }, ensure_ascii=False, indent=2))
