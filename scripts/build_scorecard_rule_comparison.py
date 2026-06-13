from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "数据清单及打分逻辑" / "数据清单及打分逻辑.xlsx"
YAML_PATH = ROOT / "configs" / "scorecards" / "shandong_scorecards_v1.yaml"
OUTPUT_XLSX = ROOT / "数据清单及打分逻辑" / "当前规则与Excel打分逻辑对比表.xlsx"


FEATURE_LABELS = {
    "brent_change_usd_d1": "Brent 原油 T+1 变动",
    "brent_change_usd_d3": "Brent 原油 T+3 变动",
    "brent_change_usd_w1": "Brent 原油 W+1 变动",
    "brent_change_usd_mom": "Brent 原油 M+1 变动",
    "shandong_cdu_utilization_percentile_weekly": "山东地炼常减压开工率分位数",
    "shandong_cdu_utilization_percentile_monthly": "山东地炼常减压开工率分位数",
    "shandong_refinery_load_news_adjustment_d1": "炼厂临时降负荷/复工资讯修正",
    "shandong_refinery_load_news_adjustment_d3": "炼厂临时降负荷/复工资讯修正",
    "shandong_refinery_load_news_adjustment_w1": "炼厂临时降负荷/复工资讯修正",
    "sales_production_ratio_d1": "汽油产销率",
    "sales_production_ratio_d3_avg": "连续3日产销率均值",
    "sales_production_ratio_w1_avg": "W+1产销率均值",
    "trader_sentiment_label_d1": "成交活跃度、贸易商心态",
    "trader_sentiment_label_d3": "成交活跃度、贸易商心态",
    "shandong_product_inventory_percentile_weekly": "山东成品油库存分位数",
    "price_window_expectation_weekly": "发改委调价窗口预期",
    "price_window_expectation_monthly": "发改委调价窗口预期",
    "next_month_maintenance_plan": "主营/地炼次月检修及开工计划",
    "monthly_seasonality_phase": "月度季节性周期",
    "restocking_rhythm_monthly": "备货节奏",
    "holiday_demand_delta_monthly": "阶段性需求/节假日",
    "refinery_inventory_monthly": "山东独立炼厂库存分位数",
    "main_company_inventory_monthly": "山东主营销售公司库存分位数",
    "market_sentiment_monthly": "市场月度整体情绪",
}


IMPLEMENTATION_NOTES = {
    "brent_change_usd_d1": "按用户指定保留当前口径：优先取 Brent 日报当日预测点位 - 昨日 settlement；不使用实时价作为打分输入。",
    "brent_change_usd_d3": "优先取 Brent 日报 horizon_forecasts.D3.change_usd；缺 D3 时回退 W1.change_usd。",
    "brent_change_usd_w1": "优先取 Brent 日报 horizon_forecasts.W1.change_usd。",
    "brent_change_usd_mom": "优先取 Brent 日报 horizon_forecasts.M1 或 W4 的 change_usd。",
    "shandong_cdu_utilization_percentile_weekly": "代码门控：若 shandong_cdu_utilization_wow_pct 或 crude_run_change_1w 为 0，则本项强制 0 分；否则按分位数打分。",
    "shandong_cdu_utilization_percentile_monthly": "代码门控：若 shandong_cdu_utilization_wow_pct 或 crude_run_change_1w 为 0，则本项强制 0 分；否则按分位数打分。",
    "shandong_refinery_load_news_adjustment_d1": "资讯修正项：识别停工、检修、降负荷、限产、复工、提负、开工提升、供应增加等文字，范围 [-5,+5]。",
    "shandong_refinery_load_news_adjustment_d3": "资讯修正项：识别停工、检修、降负荷、限产、复工、提负、开工提升、供应增加等文字，范围 [-5,+5]。",
    "shandong_refinery_load_news_adjustment_w1": "资讯修正项：识别停工、检修、降负荷、限产、复工、提负、开工提升、供应增加等文字，范围 [-5,+5]。",
    "sales_production_ratio_d1": "使用产销率，不再用销量替代；YAML 标识缺失时可用 T-1。",
    "sales_production_ratio_d3_avg": "当前代码 rolling(3, min_periods=1) 均值。",
    "sales_production_ratio_w1_avg": "当前代码 rolling(7, min_periods=1) 均值，已按 Excel 连续7日产销率口径调整。",
    "trader_sentiment_label_d1": "LLM 只判断成交活跃度、贸易商心态，输出 bullish_active/neutral_flat/bearish_selling；代码按固定分值映射。",
    "trader_sentiment_label_d3": "LLM 只判断成交活跃度、贸易商心态，输出 bullish_active/neutral_flat/bearish_selling；代码按固定分值映射。",
    "shandong_product_inventory_percentile_weekly": "按山东贸易商库存+山东主营库存+山东独立炼厂库存三项求和后做分位数；缺任一组件则 missing=0；若总库存变化为0则强制0分。",
    "refinery_inventory_monthly": "当前读取山东独立炼厂/汽油库存分位数字段；若库存变化字段为 0，则强制 0 分。",
    "main_company_inventory_monthly": "字段已打通，主营库存缺失时 missing=0；若变化字段为 0，则强制 0 分。",
    "price_window_expectation_weekly": "读取调价预测金额字段，>50 元/吨为上调预期，<-50 元/吨为下调预期，缺失时 neutral。",
    "price_window_expectation_monthly": "读取调价预测金额字段，>50 元/吨为上调预期，<-50 元/吨为下调预期，缺失时 neutral。",
    "monthly_seasonality_phase": "按预测目标下月的汽油淡旺季月份判断：旺季+5、次旺+2、次淡-2、淡季-5。",
    "restocking_rhythm_monthly": "用月度产销率均值相对上一月变化判断：>=5个百分点 active_restocking，<=-5个百分点 reduced_restocking。",
    "holiday_demand_delta_monthly": "优先读取落库节假日变化字段；缺失时按2026官方节假日安排统计当前月与下月天数差。",
    "market_sentiment_monthly": "LLM 只判断月度备货和市场心态标签，输出 peak_season_bullish/neutral/bearish；代码按固定分值映射。",
}


COMPARISON_ROWS = [
    ["D1", "成本端", "Brent 原油", "分档基本一致", "Excel 写“原油用实时数据”；当前按用户指定使用“日报预测点位-昨日 settlement”", "有意保留差异", "不改 T+1 Brent 数据口径"],
    ["D1", "成本端", "裂解价差", "已废除打分项", "现系统不再从ETA/手工模板取预计算裂解价差；按 山东92#市场价/1.13-2109.76-Brent*6.77*人民币汇率中间价 公式计算，仅展示和备查", "已调整", "不再作为智能体或业务基准打分项"],
    ["D1", "供给端", "山东地炼常减压开工率", "一致", "Excel 明确“当前数据与前一条一致则0分”；行业报告可修正 [-5,+5]", "已实现", "已加无变化强制0分和资讯修正"],
    ["D1", "需求端", "产销率", "一致", "Excel 明确汽油产销率分档，T日未更新用T-1", "基本一致", "数据链路需保证 T-1 fallback 实际生效"],
    ["D1", "情绪端", "成交活跃度、贸易商心态", "一致", "Excel 不是泛化资讯情绪，而是抢货/推涨、成交平淡、抛货/看跌", "已实现", "LLM 只判标签，代码固定映射分值"],
    ["D3", "成本端", "Brent 原油", "一致", "Excel 写原油预测结果；当前优先 D3 预测，缺失回退 W1", "基本一致", "需保证 Brent 日报能稳定给 D3 或 W1"],
    ["D3", "成本端", "裂解价差", "已废除打分项", "现系统不再从ETA/手工模板取预计算裂解价差；按 山东92#市场价/1.13-2109.76-Brent*6.77*人民币汇率中间价 公式计算，仅展示和备查", "已调整", "不再作为智能体或业务基准打分项"],
    ["D3", "供给端", "山东地炼常减压开工率", "一致", "Excel 明确无变化则0分，行业报告修正", "已实现", ""],
    ["D3", "需求端", "连续3日产销率均值", "一致", "Excel 明确连续3日均值", "一致", "当前 rolling(3)"],
    ["D3", "情绪端", "成交活跃度、贸易商心态", "一致", "文字判定同 D1，但分值 ±10", "已实现", ""],
    ["W1", "成本端", "Brent 原油", "一致", "Excel 写原油预测结果；当前用 W1 预测变动", "基本一致", ""],
    ["W1", "成本端", "裂解价差", "已废除打分项", "现系统不再从ETA/手工模板取预计算裂解价差；按 山东92#市场价/1.13-2109.76-Brent*6.77*人民币汇率中间价 公式计算，仅展示和备查", "已调整", "不再作为智能体或业务基准打分项"],
    ["W1", "供给端", "山东地炼常减压开工率", "一致", "无变化则0分；行业报告修正 [-5,+5]", "已实现", ""],
    ["W1", "需求端", "产销率均值", "一致", "Excel 文字写“连续7日产销率均值”；当前代码已改为 rolling(7)", "已实现", "继续保证产销率可回看7天"],
    ["W1", "库存端", "山东成品油库存", "分档一致", "Excel 要求贸易商库存+主营库存+独立炼厂库存求和后分位数；当前代码已按三项求和字段等待数据", "字段已预留", "补齐三类库存字段后自动生效"],
    ["W1", "调价端", "油价调价预期", "一致", "Excel 明确上调超50 +5、下调超50 -5；当前代码已改为读取调价预测金额字段", "字段已预留", "补齐调价预测金额字段后自动生效"],
    ["M1", "成本端", "Brent 原油", "一致", "Excel 已修订为 Brent ±20；当前采用 Brent ±20 + 裂解 ±10 的总成本30", "已实现", ""],
    ["M1", "成本端", "裂解价差", "已废除打分项", "现系统不再从ETA/手工模板取预计算裂解价差；按 山东92#市场价/1.13-2109.76-Brent*6.77*人民币汇率中间价 公式计算，仅展示和备查", "已调整", "不再作为智能体或业务基准打分项"],
    ["M1", "供给端", "山东地炼常减压开工率", "一致", "无变化则0分", "已实现", ""],
    ["M1", "供给端", "次月检修/开工计划", "一致", "集中检修减量 +15、平稳0、集中复工过剩 -15", "基本一致", "依赖检修计划结构化质量"],
    ["M1", "需求端", "备货节奏", "一致", "Excel 写按T月份产销与T-1月份比较判断补库节奏；当前用月度产销率变化字段判断", "已实现", "阈值暂定±5个百分点，可由业务继续校准"],
    ["M1", "需求端", "月度季节性周期", "一致", "Excel 已修订为旺季+5、次旺+2、次淡-2、淡季-5；当前按下月汽油季节带执行", "已实现", ""],
    ["M1", "需求端", "阶段性需求", "一致", "Excel 要看 T+1 月节假日是否增加；当前已支持落库字段，缺失时按2026官方节假日统计", "已实现", "后续按年度维护或接入节假日数据源"],
    ["M1", "库存端", "独立炼厂库存", "分档一致", "Excel 明确无变化则0分，按分位数；当前读取山东独立炼厂/汽油库存字段", "基本一致", "继续确认字段是否与Excel定义完全同口径"],
    ["M1", "库存端", "主营销售公司库存", "分档一致", "Excel 明确无变化则0分，按分位数", "字段已预留", "主营库存数据缺失时仍为0分"],
    ["M1", "政策/情绪", "调价预期", "一致", "Excel 仍按上调/下调超50判断；当前代码已读取调价预测金额字段", "字段已预留", "补齐字段后自动生效"],
    ["M1", "政策/情绪", "市场月度整体情绪", "一致", "Excel 需要判断月度市场情绪；当前已接入 LLM 标签，代码固定打分", "已实现", "继续补充月度午评/点评文本质量"],
]


def rules_to_text(feature: dict[str, Any]) -> str:
    method = feature.get("method")
    if method == "bucket_score":
        parts = []
        for rule in feature.get("rules", []):
            lo = "-∞" if rule.get("min") is None else rule.get("min")
            hi = "+∞" if rule.get("max") is None else rule.get("max")
            parts.append(f"[{lo},{hi}) => {rule.get('score')} ({rule.get('label', '')})")
        return "\n".join(parts)
    if method == "enum_score":
        return "\n".join(f"{k} => {v}" for k, v in (feature.get("rules") or {}).items())
    if method == "bounded_numeric":
        return f"bounded numeric: [{feature.get('min')},{feature.get('max')}]"
    if method == "calendar_month_band":
        return (
            f"peak={feature.get('peak_months')} => {feature.get('scores', {}).get('peak')}; "
            f"secondary_peak={feature.get('secondary_peak_months')} => {feature.get('scores', {}).get('secondary_peak')}; "
            f"secondary_off={feature.get('secondary_off_months')} => {feature.get('scores', {}).get('secondary_off')}; "
            f"off={feature.get('off_months')} => {feature.get('scores', {}).get('off')}"
        )
    return str(method or "")


def collect_current_rules() -> list[list[Any]]:
    payload = yaml.safe_load(YAML_PATH.read_text(encoding="utf-8"))
    card = next(c for c in payload["scorecards"] if c["scorecard_code"] == "sd_gas92")
    rows = []
    for horizon, h_payload in card["horizons"].items():
        for group in h_payload.get("factor_groups", []):
            entries = [(False, item) for item in group.get("features", [])]
            entries += [(True, item) for item in group.get("adjustments", [])]
            for is_adjustment, feature in entries:
                name = feature.get("feature_name")
                rows.append(
                    [
                        horizon,
                        group.get("group_code"),
                        group.get("display_name"),
                        group.get("weight_pct"),
                        group.get("score_cap"),
                        FEATURE_LABELS.get(name, name),
                        name,
                        "修正项" if is_adjustment else "主项",
                        feature.get("method"),
                        feature.get("weight_pct_inside_group", ""),
                        feature.get("score_cap", ""),
                        rules_to_text(feature),
                        feature.get("source_hint") or feature.get("fallback") or feature.get("note") or "",
                        IMPLEMENTATION_NOTES.get(name, ""),
                    ]
                )
    return rows


def collect_excel_rules() -> list[list[Any]]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    rows = []
    for sheet_name, horizon in [
        ("打分逻辑T+1", "D1"),
        ("打分逻辑T+3", "D3"),
        ("打分逻辑W+1", "W1"),
        ("打分逻辑M+1", "M1"),
    ]:
        ws = wb[sheet_name]
        current_dimension = ""
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in r):
                continue
            dimension, weight, score_range, indicator, scoring_rule, data_desc = r[:6]
            if dimension:
                current_dimension = dimension
            rows.append(
                [
                    horizon,
                    sheet_name,
                    current_dimension,
                    weight,
                    score_range,
                    indicator or "",
                    scoring_rule or "",
                    data_desc or "",
                ]
            )
    return rows


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "A": 10,
        "B": 16,
        "C": 18,
        "D": 12,
        "E": 12,
        "F": 26,
        "G": 34,
        "H": 12,
        "I": 18,
        "J": 14,
        "K": 14,
        "L": 46,
        "M": 38,
        "N": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 54


def add_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_current = wb.create_sheet("当前系统规则清单")
    add_table(
        ws_current,
        [
            "周期",
            "维度编码",
            "维度名称",
            "维度权重",
            "维度封顶",
            "指标名称",
            "系统字段",
            "项目类型",
            "方法",
            "组内权重",
            "单项封顶",
            "当前分数规则",
            "YAML说明/数据口径",
            "代码实现口径/文字判定",
        ],
        collect_current_rules(),
    )
    style_sheet(ws_current)

    ws_excel = wb.create_sheet("Excel原始规则清单")
    add_table(
        ws_excel,
        ["周期", "Sheet", "一级维度", "权重", "打分", "二级量化指标", "双向打分规则原文", "数据描述原文"],
        collect_excel_rules(),
    )
    style_sheet(ws_excel)
    ws_excel.column_dimensions["G"].width = 72
    ws_excel.column_dimensions["H"].width = 60

    ws_cmp = wb.create_sheet("当前规则与Excel对比")
    add_table(
        ws_cmp,
        ["周期", "维度", "指标", "分数区间对比", "文字描述/判定口径对比", "当前结论", "建议动作"],
        COMPARISON_ROWS,
    )
    style_sheet(ws_cmp)
    ws_cmp.column_dimensions["D"].width = 24
    ws_cmp.column_dimensions["E"].width = 72
    ws_cmp.column_dimensions["F"].width = 18
    ws_cmp.column_dimensions["G"].width = 52
    fills = {
        "不一致": PatternFill("solid", fgColor="F8CBAD"),
        "不完整": PatternFill("solid", fgColor="FCE4D6"),
        "部分一致": PatternFill("solid", fgColor="FFF2CC"),
        "有意保留差异": PatternFill("solid", fgColor="DDEBF7"),
    }
    for row in ws_cmp.iter_rows(min_row=2):
        status = str(row[5].value or "")
        if status in fills:
            for cell in row:
                cell.fill = fills[status]

    ws_text = wb.create_sheet("文字判定重点")
    add_table(
        ws_text,
        ["序号", "Excel关键文字", "为什么重要", "当前系统是否支持", "处理建议"],
        [
            [1, "原油用实时数据", "会决定 Brent 打分输入值；但用户已指定 T+1 保留日报预测点位-昨日 settlement", "T+1 有意不按 Excel 实时价", "表内保留说明，防止后续误改"],
            [2, "若当前数据与前一条数据一致则0分", "直接改变开工率、库存项分数，不能只看分位数区间", "开工率和库存已加代码门控", "继续保证采集数据有变化字段"],
            [3, "炼厂临时降负荷等消息可修正 [-5,+5]", "供给端不仅是开工率数值，还要看新闻事件", "已做关键词修正", "后续可改成 LLM 提取事件标签再固定打分"],
            [4, "若T日数据未更新则用T-1", "影响日报早间预测能否稳定出分", "D1 YAML 有 fallback 标识，数据链路仍需核对", "建议在数据层记录实际取数日期"],
            [5, "连续7日产销率均值", "影响W+1需求端平滑程度和分数敏感度", "已改为 rolling(7)", "继续保证产销率可回看7天"],
            [6, "上调超50/下调超50", "调价端必须按预测调价金额，不能用 Brent 变动近似", "已改为读取调价预测金额字段", "补齐字段后自动生效"],
            [7, "转入旺季/转入淡季", "季节性直接影响M+1需求端", "已按新版Excel改为旺季+5、次旺+2、次淡-2、淡季-5", "后续可继续确认是否需要严格判断“转入”"],
            [8, "T+1月节假日是否增加", "节假日天数变化影响阶段性需求", "已支持字段优先，2026官方节假日兜底统计", "建议后续接入年度节假日表"],
            [9, "成交活跃度、贸易商心态/月度市场情绪", "情绪端不是宏观新闻情绪，而是交易行为、备货和心态", "已用 LLM 判标签、代码固定打分", "补充隆众午评/每日点评稳定来源"],
        ],
    )
    style_sheet(ws_text)
    ws_text.column_dimensions["B"].width = 44
    ws_text.column_dimensions["C"].width = 52
    ws_text.column_dimensions["D"].width = 42
    ws_text.column_dimensions["E"].width = 48

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
