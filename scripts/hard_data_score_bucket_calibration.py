
from __future__ import annotations

import argparse
import json
import math
import platform
import sys
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any

platform._wmi_query = lambda *a, **k: ['10.0.0', '1', 'Multiprocessor Free', '0', '0']
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.container import get_dataset_service, get_predictor
from app.services.predictors.horizons import DEFAULT_HORIZONS, resolve_horizon_config
from app.services.predictors.shandong_regional_spreads import REGIONAL_SPREAD_CONFIGS, DIESEL_REGIONAL_SPREAD_CONFIGS

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_START_DATE = date(2024, 1, 1)
DEFAULT_AS_OF = date(2026, 6, 12)
MIN_BUCKET_SAMPLE_SIZE = 12
HARD_DATA_EXCEL = ROOT / '\u4ea7\u9500\u7387\u7ec8\u7248.xlsx'

# Hard-data calibration excludes non-replayable soft factors: refined news, event gate, policy sentiment, Brent forecast points.
# Brent uses actual historical settlement changes that can be reconstructed.
HARD_REQUIRED_COLUMNS = {
    'D1': ['brent_change_1d', 'sd_gas_crack', 'shandong_cdu_utilization_weekly', 'sales_production_ratio_d1', 'shandong_product_inventory_percentile_weekly'],
    'W1': ['brent_change_5d', 'sd_gas_crack', 'shandong_cdu_utilization_weekly', 'sales_production_ratio_w1_avg', 'shandong_product_inventory_percentile_weekly'],
    'M1': ['brent_change_20d', 'sd_gas_crack', 'shandong_cdu_utilization_weekly', 'sales_production_ratio_monthly_avg', 'shandong_refinery_inventory_percentile_monthly', 'shandong_main_company_inventory_percentile_monthly'],
}
DIESEL_REQUIRED_COLUMN_ALIASES = {
    'sales_production_ratio_d1': 'diesel_sales_production_ratio_d1',
    'sales_production_ratio_w1_avg': 'diesel_sales_production_ratio_w1_avg',
    'sales_production_ratio_monthly_avg': 'diesel_sales_production_ratio_monthly_avg',
    'shandong_product_inventory_percentile_weekly': 'shandong_diesel_product_inventory_percentile_weekly',
    'shandong_refinery_inventory_percentile_monthly': 'shandong_diesel_refinery_inventory_percentile_monthly',
    'shandong_main_company_inventory_percentile_monthly': 'shandong_diesel_main_company_inventory_percentile_monthly',
}

GAS_PRICE_TARGETS = {
    'SHANDONG_GAS92': ('山东汽油', 'sd_gas92_market'),
    'EAST_CHINA_GAS92': ('华东汽油', 'east_china_gas92_market'),
    'NORTH_CHINA_GAS92': ('华北汽油', 'north_china_gas92_market'),
    'SOUTH_CHINA_GAS92': ('华南汽油', 'south_china_gas92_market'),
    'CENTRAL_CHINA_GAS92': ('华中汽油', 'central_china_gas92_market'),
    'NORTHWEST_GAS92': ('西北汽油', 'northwest_gas92_market'),
    'SOUTHWEST_GAS92': ('西南汽油', 'southwest_gas92_market'),
    'NORTHEAST_GAS92': ('东北汽油', 'northeast_gas92_market'),
}
DIESEL_PRICE_TARGETS = {
    'SHANDONG_DIESEL0': ('山东柴油', 'sd_diesel0_market'),
    'EAST_CHINA_DIESEL0': ('华东柴油', 'east_china_diesel0_market'),
    'NORTH_CHINA_DIESEL0': ('华北柴油', 'north_china_diesel0_market'),
    'SOUTH_CHINA_DIESEL0': ('华南柴油', 'south_china_diesel0_market'),
    'CENTRAL_CHINA_DIESEL0': ('华中柴油', 'central_china_diesel0_market'),
    'NORTHWEST_DIESEL0': ('西北柴油', 'northwest_diesel0_market'),
    'SOUTHWEST_DIESEL0': ('西南柴油', 'southwest_diesel0_market'),
    'NORTHEAST_DIESEL0': ('东北柴油', 'northeast_diesel0_market'),
}


def parse_iso_date(value: str, fallback: date) -> date:
    return date.fromisoformat(value) if value else fallback


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return isinstance(value, str) and not value.strip()


def q(values: list[float], quantile: float) -> float | None:
    values = [float(v) for v in values if v == v]
    if not values:
        return None
    return round(float(np.quantile(values, quantile)), 2)


def up_rate(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(1 for v in values if v > 0) / len(values), 3)


def load_sales_ratio_excel(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        dt = pd.to_datetime(row[0], errors='coerce')
        if pd.isna(dt):
            continue
        gas = pd.to_numeric(row[1], errors='coerce') if len(row) > 1 else np.nan
        diesel = pd.to_numeric(row[2], errors='coerce') if len(row) > 2 else np.nan
        rows.append({'date': pd.Timestamp(dt).normalize(), 'sales_production_ratio_d1': float(gas) * 100.0 if pd.notna(gas) else np.nan, 'diesel_sales_production_ratio_d1': float(diesel) * 100.0 if pd.notna(diesel) else np.nan})
    frame = pd.DataFrame(rows).drop_duplicates(subset=['date'], keep='last').sort_values('date')
    return frame


def apply_sales_ratio_override(frame: pd.DataFrame, ratio_frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    out['date'] = pd.to_datetime(out['date']).dt.normalize()
    out = out.merge(ratio_frame, on='date', how='left', suffixes=('', '_excel'))
    for col in ['sales_production_ratio_d1', 'diesel_sales_production_ratio_d1']:
        excel_col = f'{col}_excel'
        if excel_col in out.columns:
            out[col] = out[excel_col].combine_first(out.get(col))
            out = out.drop(columns=[excel_col])
    out['sales_production_ratio_d3_avg'] = out['sales_production_ratio_d1'].rolling(3, min_periods=3).mean()
    out['sales_production_ratio_w1_avg'] = out['sales_production_ratio_d1'].rolling(7, min_periods=7).mean()
    out['sales_production_ratio_monthly_avg'] = out['sales_production_ratio_d1'].rolling(30, min_periods=1).mean()
    out['diesel_sales_production_ratio_d3_avg'] = out['diesel_sales_production_ratio_d1'].rolling(3, min_periods=3).mean()
    out['diesel_sales_production_ratio_w1_avg'] = out['diesel_sales_production_ratio_d1'].rolling(7, min_periods=7).mean()
    out['diesel_sales_production_ratio_monthly_avg'] = out['diesel_sales_production_ratio_d1'].rolling(30, min_periods=1).mean()
    return out


def load_actual_price_dates(dataset_service: Any, *, start_date: date, end_date: date, columns: list[str]) -> dict[str, set[pd.Timestamp]]:
    loader = getattr(dataset_service, '_load_price_history_rows_from_archive', None)
    if loader is None:
        return {}
    try:
        archive_frame = loader(requested=columns, start_date=start_date, end_date=end_date)
    except Exception:
        return {}
    if archive_frame is None or archive_frame.empty or 'date' not in archive_frame.columns:
        return {}
    archive_frame = archive_frame.copy()
    archive_frame['date'] = pd.to_datetime(archive_frame['date'], errors='coerce').dt.normalize()
    result: dict[str, set[pd.Timestamp]] = {}
    for column in columns:
        if column in archive_frame.columns:
            mask = archive_frame[column].notna()
            result[column] = set(archive_frame.loc[mask, 'date'].dropna().tolist())
    return result


def hard_extra(row_date: date, horizon: str, product_code: str) -> dict[str, Any]:
    return {
        'as_of_date': row_date,
        'mode': 'hard_data_bucket_calibration',
        'report_payload': None,
        'news_items': [],
        'refined_news_items': [],
        'policy_items': [],
        'prediction_subject': 'outright',
        'enable_refined_news': False,
        'enable_event_risk': False,
        'horizon': horizon,
        'product_code': product_code,
        'refined_news_labels': {},
        'trade_sentiment': {},
        'monthly_market_sentiment': {},
    }


def required_columns_for(horizon: str, product_code: str) -> list[str]:
    columns = list(HARD_REQUIRED_COLUMNS.get(horizon, []))
    if str(product_code).upper() == 'DIESEL_0':
        return [DIESEL_REQUIRED_COLUMN_ALIASES.get(col, col) for col in columns]
    return columns


def ignored_missing_columns_for(*, target_key: str, horizon: str, product_code: str) -> set[str]:
    # Shandong gasoline has daily local prices, including weekends. If weekend rows only lack
    # Brent trading data, keep the row in the calibration sample and let the score replay proceed
    # without treating Brent as a hard completeness gate.
    if target_key == 'SHANDONG_GAS92' and str(product_code).upper() == 'GASOLINE_92':
        brent_col = {
            'D1': 'brent_change_1d',
            'W1': 'brent_change_5d',
            'M1': 'brent_change_20d',
        }.get(horizon)
        return {brent_col} if brent_col else set()
    return set()


def score_hard_frame(
    predictor: Any,
    frame: pd.DataFrame,
    *,
    target_key: str,
    horizon: str,
    product_code: str,
    product_spec: Any | None,
) -> pd.DataFrame:
    work = frame.copy()
    if product_spec is not None:
        current = work.iloc[-1] if not work.empty else pd.Series(dtype=object)
        work, _, _quality = predictor._product_feature_view(feature_frame=work, current_row=current, product_spec=product_spec)
    rows = []
    required_columns = required_columns_for(horizon, product_code)
    ignored_missing_columns = ignored_missing_columns_for(
        target_key=target_key,
        horizon=horizon,
        product_code=product_code,
    )
    for _, row in work.iterrows():
        row_date = pd.Timestamp(row['date']).date()
        missing_all = [col for col in required_columns if col not in row.index or is_missing(row.get(col))]
        missing = [col for col in missing_all if col not in ignored_missing_columns]
        extra = hard_extra(row_date, horizon, product_code)
        business_claim = predictor._score_business_scorecard(row, extra=extra)
        claims, agent_score = predictor._score_row(row, extra=extra)
        item = row.to_dict()
        item['agent_score'] = float(agent_score)
        item['business_scorecard_score'] = float(business_claim.numeric_signals.get('standalone_score', 0.0))
        item['strict_missing_fields'] = sorted(set(missing))
        item['strict_missing_count'] = len(set(missing))
        item['ignored_missing_fields'] = sorted(set(missing_all) - set(missing))
        item['strict_complete'] = len(set(missing)) == 0
        rows.append(item)
    return pd.DataFrame(rows)


def score_points(predictor: Any, score_column: str, value: float) -> float:
    return predictor._score_points(score_column=score_column, score_value=float(value))


def bucket_index(score: float, defs: list[dict[str, Any]]) -> int:
    for idx, item in enumerate(defs):
        if float(item['lower']) <= float(score) < float(item['upper']):
            return idx
    return len(defs) - 1


def evaluate_buckets(predictor: Any, scored: pd.DataFrame, score_column: str) -> list[dict[str, Any]]:
    defs = predictor._score_bucket_defs(score_column)
    out = []
    if scored.empty:
        for item in defs:
            out.append({'bucket': item['label'], 'score_range': item['range_label'], 'sample_size': 0, 'p25': None, 'p50': None, 'p75': None, 'up_rate': None, 'sample_sufficient': False})
        return out
    work = scored.copy()
    work['_score_points'] = work[score_column].map(lambda v: score_points(predictor, score_column, float(v)))
    work['_bucket_index'] = work['_score_points'].map(lambda v: bucket_index(float(v), defs))
    for idx, item in enumerate(defs):
        vals = work[work['_bucket_index'] == idx]['target_delta'].astype(float).tolist()
        out.append({'bucket': item['label'], 'score_range': item['range_label'], 'sample_size': len(vals), 'p25': q(vals, .25), 'p50': q(vals, .5), 'p75': q(vals, .75), 'up_rate': up_rate(vals), 'sample_sufficient': len(vals) >= MIN_BUCKET_SAMPLE_SIZE})
    return out


BUCKET_LABELS = ['\u5f3a\u7a7a', '\u504f\u7a7a', '\u5f31\u7a7a', '\u9707\u8361', '\u5f31\u591a', '\u504f\u591a', '\u5f3a\u591a']
BUCKET_POLARITIES = [-1, -1, -1, 0, 1, 1, 1]
AGENT_CANDIDATES = [
    [-30, -18, -8, 8, 18, 30],
    [-25, -15, -6, 6, 15, 25],
    [-20, -12, -5, 5, 15, 25],
    [-15, -9, -3, 3, 10, 20],
    [-15, -9, 0, 15, 25, 35],
    [-12, -6, 0, 18, 28, 38],
    [-20, -10, 0, 20, 30, 40],
    [-25, -15, -5, 10, 25, 35],
]
BUSINESS_CANDIDATES = [
    [-40, -25, -10, 10, 25, 40],
    [-35, -20, -8, 8, 20, 35],
    [-30, -15, -5, 5, 15, 30],
    [-25, -15, -5, 5, 10, 15],
    [-20, -12, -4, 4, 10, 18],
    [-20, -10, -1, 1, 8, 15],
    [-15, -8, -1, 1, 8, 15],
]


def make_bucket_defs(thresholds: list[float]) -> list[dict[str, Any]]:
    bounds = [-math.inf, *[float(v) for v in thresholds], math.inf]
    defs = []
    for idx, label in enumerate(BUCKET_LABELS):
        lower = bounds[idx]
        upper = bounds[idx + 1]
        lower_text = '-\u221e' if math.isinf(lower) and lower < 0 else f'{lower:g}'
        upper_text = '+\u221e' if math.isinf(upper) and upper > 0 else f'{upper:g}'
        defs.append({'label': label, 'lower': lower, 'upper': upper, 'range_label': f'[{lower_text}, {upper_text})', 'polarity': BUCKET_POLARITIES[idx]})
    return defs


def evaluate_thresholds(predictor: Any, scored: pd.DataFrame, score_column: str, thresholds: list[float]) -> dict[str, Any]:
    defs = make_bucket_defs(thresholds)
    if scored.empty:
        return {'thresholds': thresholds, 'sample_count': 0, 'mae': None, 'rmse': None, 'penalty': None, 'objective': None, 'bucket_stats': []}
    rows = []
    by_bucket = [[] for _ in defs]
    for _, row in scored.iterrows():
        points = score_points(predictor, score_column, float(row[score_column]))
        delta = float(row['target_delta'])
        idx = bucket_index(points, defs)
        rows.append((points, delta, idx))
        by_bucket[idx].append(delta)
    preds = []
    for _points, delta, idx in rows:
        sample = list(by_bucket[idx])
        if len(sample) < MIN_BUCKET_SAMPLE_SIZE:
            sample = []
            for near_idx in [idx - 1, idx, idx + 1]:
                if 0 <= near_idx < len(by_bucket):
                    sample.extend(by_bucket[near_idx])
        pred = q(sample, 0.5) if sample else 0.0
        preds.append((float(pred), delta))
    mae = sum(abs(pred - actual) for pred, actual in preds) / len(preds)
    rmse = math.sqrt(sum((pred - actual) ** 2 for pred, actual in preds) / len(preds))
    penalty = 0.0
    stats = []
    for idx, item in enumerate(defs):
        vals = by_bucket[idx]
        med = q(vals, 0.5)
        n = len(vals)
        polarity = int(item['polarity'])
        if 0 < n < MIN_BUCKET_SAMPLE_SIZE:
            penalty += (MIN_BUCKET_SAMPLE_SIZE - n) * 5
        if n >= MIN_BUCKET_SAMPLE_SIZE and med is not None:
            if polarity < 0 and med > 0:
                penalty += abs(med) * 2 + 50
            elif polarity > 0 and med < 0:
                penalty += abs(med) * 2 + 50
            elif polarity == 0 and abs(med) > 30:
                penalty += abs(med)
        stats.append({'bucket': item['label'], 'score_range': item['range_label'], 'sample_size': n, 'p25': q(vals, 0.25), 'p50': med, 'p75': q(vals, 0.75), 'up_rate': up_rate(vals), 'sample_sufficient': n >= MIN_BUCKET_SAMPLE_SIZE})
    return {
        'thresholds': [float(v) for v in thresholds],
        'sample_count': int(len(scored)),
        'mae': round(float(mae), 2),
        'rmse': round(float(rmse), 2),
        'penalty': round(float(penalty), 2),
        'objective': round(float(mae + penalty), 2),
        'bucket_stats': stats,
    }


def candidate_results(predictor: Any, strict: pd.DataFrame, score_column: str, candidates: list[list[float]]) -> list[dict[str, Any]]:
    results = [evaluate_thresholds(predictor, strict, score_column, thresholds) for thresholds in candidates]
    return sorted(results, key=lambda item: (float('inf') if item['objective'] is None else item['objective'], float('inf') if item['mae'] is None else item['mae']))


def _date_membership(series: pd.Series, valid_dates: set[pd.Timestamp]) -> pd.Series:
    normalized = pd.to_datetime(series, errors='coerce').dt.normalize()
    return normalized.isin(valid_dates)


def calibrate_target(
    predictor: Any,
    frame: pd.DataFrame,
    *,
    target_key: str,
    target_label: str,
    target_col: str,
    product_code: str,
    product_spec: Any | None,
    start_date: date,
    as_of: date,
    actual_price_dates: dict[str, set[pd.Timestamp]] | None = None,
) -> dict[str, Any]:
    result = {'target_key': target_key, 'target_label': target_label, 'target_column': target_col, 'product_code': product_code, 'horizons': {}}
    for horizon in DEFAULT_HORIZONS:
        hc = resolve_horizon_config(horizon)
        work = frame.copy()
        if target_col not in work.columns:
            result['horizons'][horizon] = {'raw_samples': 0, 'complete_samples': 0, 'excluded_samples': 0, 'missing_reason': f'missing target column {target_col}', 'agent_buckets': [], 'business_buckets': []}
            continue
        price_valid_dates = (actual_price_dates or {}).get(target_col) or set()
        if price_valid_dates:
            work = work[_date_membership(work['date'], price_valid_dates)].copy()
        work['target_date'] = work['date'].shift(-hc.steps)
        work['target_price'] = work[target_col].shift(-hc.steps)
        work['target_delta'] = work['target_price'] - work[target_col]
        history = work[(work['date'] >= pd.Timestamp(start_date)) & (work['date'] < pd.Timestamp(as_of)) & (work['target_date'] <= pd.Timestamp(as_of))].dropna(subset=['target_delta']).copy()
        if price_valid_dates:
            price_mask = _date_membership(history['date'], price_valid_dates) & _date_membership(history['target_date'], price_valid_dates)
            history = history[price_mask].copy()
        scored_all = score_hard_frame(predictor, history, target_key=target_key, horizon=horizon, product_code=product_code, product_spec=product_spec) if not history.empty else pd.DataFrame()
        strict = scored_all[scored_all['strict_complete']].dropna(subset=['agent_score', 'business_scorecard_score', 'target_delta']).copy() if not scored_all.empty else scored_all
        missing_counts = scored_all['strict_missing_count'].value_counts().sort_index().to_dict() if not scored_all.empty else {}
        result['horizons'][horizon] = {
            'raw_samples': int(len(history)),
            'complete_samples': int(len(strict)),
            'excluded_samples': int(len(scored_all) - len(strict)),
            'missing_count_distribution': {str(k): int(v) for k, v in missing_counts.items()},
            'threshold_update_allowed': bool(len(strict) >= MIN_BUCKET_SAMPLE_SIZE * 7),
            'actual_price_date_filter': bool(price_valid_dates),
            'agent_buckets': evaluate_buckets(predictor, strict, 'agent_score'),
            'business_buckets': evaluate_buckets(predictor, strict, 'business_scorecard_score'),
            'agent_threshold_candidates': candidate_results(predictor, strict, 'agent_score', AGENT_CANDIDATES) if len(strict) >= MIN_BUCKET_SAMPLE_SIZE * 7 else [],
            'business_threshold_candidates': candidate_results(predictor, strict, 'business_scorecard_score', BUSINESS_CANDIDATES) if len(strict) >= MIN_BUCKET_SAMPLE_SIZE * 7 else [],
        }
    return result


def write_xlsx(summary: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    ws.append(['target_key', 'target_label', 'product', 'horizon', 'raw_samples', 'complete_samples', 'excluded_samples', 'threshold_update_allowed', 'missing_distribution', 'best_agent_thresholds', 'best_business_thresholds', 'calibration_note'])
    note = '\u786c\u6570\u636e\u6821\u51c6\uff1a\u4ea7\u9500\u7387\u4f7f\u7528\u4ea7\u9500\u7387\u7ec8\u7248.xlsx\uff1bBrent\u4f7f\u7528\u5386\u53f2\u5df2\u53d1\u751f\u7684\u771f\u5b9e\u53d8\u52a8\uff1b\u8d44\u8baf/\u4e8b\u4ef6/\u653f\u7b56\u60c5\u7eea\u6682\u4e0d\u53c2\u4e0e\uff1b\u4e0d\u5141\u8bb8\u7f3a\u5931\u515c\u5e95\u3002'
    for target in summary['targets']:
        for horizon, item in target['horizons'].items():
            best_agent = (item.get('agent_threshold_candidates') or [{}])[0].get('thresholds')
            best_business = (item.get('business_threshold_candidates') or [{}])[0].get('thresholds')
            ws.append([target['target_key'], target['target_label'], target['product_code'], horizon, item.get('raw_samples'), item.get('complete_samples'), item.get('excluded_samples'), 'yes' if item.get('threshold_update_allowed') else 'no', json.dumps(item.get('missing_count_distribution') or {}, ensure_ascii=False), json.dumps(best_agent, ensure_ascii=False) if best_agent else '', json.dumps(best_business, ensure_ascii=False) if best_business else '', note])

    ws_best = wb.create_sheet('RecommendedThresholds')
    ws_best.append(['target_key', 'target_label', 'product', 'horizon', 'model', 'recommended_thresholds', 'sample_count', 'MAE', 'RMSE', 'penalty', 'objective', 'update_allowed', 'remark'])
    for target in summary['targets']:
        for horizon, item in target['horizons'].items():
            for model, key in [('\u667a\u80fd\u4f53', 'agent_threshold_candidates'), ('\u4e1a\u52a1', 'business_threshold_candidates')]:
                best = (item.get(key) or [{}])[0]
                thresholds = best.get('thresholds')
                remark = '\u6837\u672c\u5145\u8db3\uff0c\u53ef\u4f5c\u4e3a\u786c\u6570\u636e\u53e3\u5f84\u5019\u9009\u9608\u503c' if thresholds else '\u6837\u672c\u4e0d\u8db3\u6216\u786c\u6570\u636e\u7f3a\u5931\uff0c\u4e0d\u5efa\u8bae\u66f4\u65b0\u9608\u503c'
                ws_best.append([target['target_key'], target['target_label'], target['product_code'], horizon, model, json.dumps(thresholds, ensure_ascii=False) if thresholds else '', best.get('sample_count'), best.get('mae'), best.get('rmse'), best.get('penalty'), best.get('objective'), 'yes' if item.get('threshold_update_allowed') else 'no', remark])

    for model_key, title in [('agent_buckets', 'CurrentAgentBuckets'), ('business_buckets', 'CurrentBusinessBuckets')]:
        ws2 = wb.create_sheet(title)
        ws2.append(['target_key', 'target_label', 'product', 'horizon', 'bucket', 'score_range', 'sample_size', 'P25', 'P50', 'P75', 'up_rate', 'sample_sufficient'])
        for target in summary['targets']:
            for horizon, item in target['horizons'].items():
                for row in item.get(model_key) or []:
                    ws2.append([target['target_key'], target['target_label'], target['product_code'], horizon, row['bucket'], row['score_range'], row['sample_size'], row['p25'], row['p50'], row['p75'], row['up_rate'], 'yes' if row['sample_sufficient'] else 'no'])

    ws3 = wb.create_sheet('CandidateDetails')
    ws3.append(['target_key', 'target_label', 'product', 'horizon', 'model', 'rank', 'thresholds', 'sample_count', 'MAE', 'RMSE', 'penalty', 'objective'])
    for target in summary['targets']:
        for horizon, item in target['horizons'].items():
            for model, key in [('\u667a\u80fd\u4f53', 'agent_threshold_candidates'), ('\u4e1a\u52a1', 'business_threshold_candidates')]:
                for rank, cand in enumerate(item.get(key) or [], 1):
                    ws3.append([target['target_key'], target['target_label'], target['product_code'], horizon, model, rank, json.dumps(cand.get('thresholds'), ensure_ascii=False), cand.get('sample_count'), cand.get('mae'), cand.get('rmse'), cand.get('penalty'), cand.get('objective')])

    ws4 = wb.create_sheet('RecommendedBucketStats')
    ws4.append(['target_key', 'target_label', 'product', 'horizon', 'model', 'thresholds', 'bucket', 'score_range', 'sample_size', 'P25', 'P50', 'P75', 'up_rate', 'sample_sufficient'])
    for target in summary['targets']:
        for horizon, item in target['horizons'].items():
            for model, key in [('\u667a\u80fd\u4f53', 'agent_threshold_candidates'), ('\u4e1a\u52a1', 'business_threshold_candidates')]:
                best = (item.get(key) or [{}])[0]
                thresholds = best.get('thresholds')
                for row in best.get('bucket_stats') or []:
                    ws4.append([target['target_key'], target['target_label'], target['product_code'], horizon, model, json.dumps(thresholds, ensure_ascii=False), row['bucket'], row['score_range'], row['sample_size'], row['p25'], row['p50'], row['p75'], row['up_rate'], 'yes' if row['sample_sufficient'] else 'no'])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='FCE4D6')
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20
        sheet.freeze_panes = 'A2'
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description='Hard-data score bucket calibration for Shandong and regional targets.')
    parser.add_argument('--start-date', default=DEFAULT_START_DATE.isoformat())
    parser.add_argument('--as-of', default=DEFAULT_AS_OF.isoformat())
    parser.add_argument('--sales-ratio-excel', default=str(HARD_DATA_EXCEL))
    parser.add_argument('--output-json', default='artifacts/hard_data_score_bucket_calibration_20240101_20260612.json')
    parser.add_argument('--output-xlsx', default='outputs/hard_data_score_bucket_calibration_20240101_20260612.xlsx')
    args = parser.parse_args()
    start_date = date.fromisoformat(args.start_date)
    as_of = date.fromisoformat(args.as_of)
    dataset_service = get_dataset_service()
    predictor = get_predictor()
    frame = dataset_service.build_feature_frame(start_date=start_date - timedelta(days=90), end_date=as_of).sort_values('date').copy()
    frame['date'] = pd.to_datetime(frame['date']).dt.normalize()
    frame = apply_sales_ratio_override(frame, load_sales_ratio_excel(Path(args.sales_ratio_excel)))
    all_target_columns = [col for _key, (_label, col) in [*GAS_PRICE_TARGETS.items(), *DIESEL_PRICE_TARGETS.items()]]
    actual_price_dates = load_actual_price_dates(
        dataset_service,
        start_date=start_date - timedelta(days=90),
        end_date=as_of,
        columns=all_target_columns,
    )
    targets = []
    for key, (label, col) in GAS_PRICE_TARGETS.items():
        targets.append(calibrate_target(predictor, frame, target_key=key, target_label=label, target_col=col, product_code='GASOLINE_92', product_spec=None, start_date=start_date, as_of=as_of, actual_price_dates=actual_price_dates))
    diesel_spec = getattr(predictor, 'DIESEL_0_SPEC', None)
    # module-level constant is not an instance attribute; import fallback through function globals
    from app.services.predictors.shandong_gas92 import DIESEL_0_SPEC
    for key, (label, col) in DIESEL_PRICE_TARGETS.items():
        targets.append(calibrate_target(predictor, frame, target_key=key, target_label=label, target_col=col, product_code='DIESEL_0', product_spec=DIESEL_0_SPEC, start_date=start_date, as_of=as_of, actual_price_dates=actual_price_dates))
    summary = {
        'policy': 'hard_data_only; production-sales from 产销率终版.xlsx; Brent uses realized historical changes; M1 inventory uses month-end/latest statistical inventory percentile; refined news/event/policy factors excluded from calibration; no missing fallback',
        'start_date': start_date.isoformat(),
        'as_of': as_of.isoformat(),
        'min_bucket_sample_size': MIN_BUCKET_SAMPLE_SIZE,
        'targets': targets,
    }
    output_json = Path(args.output_json)
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
    write_xlsx(summary, Path(args.output_xlsx))
    print(json.dumps({'json': str(output_json), 'xlsx': args.output_xlsx, 'target_count': len(targets), 'samples': {t['target_key']: {h: v['complete_samples'] for h, v in t['horizons'].items()} for t in targets}}, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
