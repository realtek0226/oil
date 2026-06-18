
from pathlib import Path
from openpyxl import load_workbook
paths=[x for x in Path('outputs').glob('*已更新待替换.xlsx') if not x.name.startswith('~')]
path=paths[0] if paths else Path('outputs') / '逻辑说明.xlsx'

def fail(msg, failures): failures.append(msg)
def check(round_no):
    wb=load_workbook(path,data_only=True)
    failures=[]
    for sheet in ['AgentScoring','BusinessScoring']:
        ws=wb[sheet]
        headers=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
        if '打分参照规则' not in headers: fail(f'{sheet} missing 打分参照规则', failures)
        if '原始规则' not in headers: fail(f'{sheet} missing 原始规则', failures)
        else:
            raw_idx=headers.index('原始规则')
            ref_idx=headers.index('打分参照规则')
            if raw_idx != ref_idx + 1: fail(f'{sheet} 原始规则 not right of 打分参照规则', failures)
    ws=wb['BusinessScoring']
    rows=list(ws.iter_rows(min_row=4,values_only=True))
    keys=[(r[0],r[1]) for r in rows if r and r[0]]
    if len(keys)!=len(set(keys)): fail('BusinessScoring still has duplicate 周期+模块 rows', failures)
    for required in [('D1','成本侧'),('D1','供给侧'),('W1','库存端'),('M1','需求侧'),('M1','库存端'),('M1','政策与情绪侧')]:
        if required not in keys: fail(f'missing merged row {required}', failures)
    text='\\n'.join(str(c.value or '') for ws in wb.worksheets for row in ws.iter_rows() for c in row)
    for phrase in ['原始规则：模块得分','原始规则：raw_score','库存=贸易商+主营+独立炼厂可用项求和','山东M1备货节奏','调价预测来自卓创成品油调价专栏','理论有效区间修正为-82~89.5']:
        if phrase not in text: fail(f'missing phrase {phrase}', failures)
    for bad in ['三项齐全','Policy and sentiment']:
        if bad in text: fail(f'bad phrase {bad}', failures)
    # value consistency checks
    for r in rows:
        if r[0]=='W1' and r[1]=='库存端':
            if 'shandong_product_inventory_percentile_weekly=21.36752136752137' not in str(r[4]): fail('W1 inventory value mismatch', failures)
            if '15' not in str(r[6]): fail('W1 inventory score missing', failures)
        if r[0]=='M1' and r[1]=='需求侧':
            if 'restocking_rhythm_monthly=active_restocking' not in str(r[4]): fail('M1 restocking value mismatch', failures)
        if r[0]=='M1' and r[1]=='政策与情绪侧':
            if 'price_window_expectation_monthly=down_adjustment_expected' not in str(r[4]): fail('M1 price window value mismatch', failures)
    print('round', round_no, 'failures', failures)
    return failures
for i in range(1,4):
    failures=check(i)
    if failures:
        raise SystemExit(1)
print('all checks passed', path)
