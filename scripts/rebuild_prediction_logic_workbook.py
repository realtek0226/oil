from __future__ import annotations

import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INPUT_JSON = ROOT / "artifacts" / "prediction_20260608_current_logic.json"
OUTPUT = ROOT / "outputs" / "两种预测逻辑用数及换算说明_预测20260608_智能体有效分修正版.xlsx"

ACTUAL_20260608_PRICE = 8003.0

HORIZON_LABELS = {
    "D1": "下一交易日",
    "D3": "三日",
    "W1": "一周",
    "M1": "一月",
}

DIRECTION_LABELS = {
    "up": "上涨",
    "down": "下跌",
    "flat": "震荡",
}

AGENT_LABELS = {
    "business_scorecard_agent": "业务基准打分模型",
    "crude_cost_agent": "原油成本智能体",
    "market_structure_agent": "市场结构智能体",
    "supply_inventory_agent": "供给库存智能体",
    "demand_seasonality_agent": "需求季节智能体",
    "refined_oil_news_agent": "成品油资讯智能体",
    "shandong_spot_jump_agent": "山东现货跳变识别智能体",
    "policy_cycle_agent": "政策周期智能体",
    "event_risk_agent": "事件风险智能体",
    "agent_judge_agent": "智能体裁判",
}

AGENT_RULES = {
    "crude_cost_agent": "Brent预测/结算变化按分档给原油成本分；汽油裂解价差分位仅展示不计分；MTBE与石脑油三日变化作为调油成本补充，按变化合计/60*10并限幅±10。",
    "market_structure_agent": "区域价差取各目标区域价格-山东价格，区域平均价差越高越利多山东；山东-全国价差与山东92#短期动量作为辅助项。",
    "supply_inventory_agent": "开工率分位越低代表供应偏紧越利多；开工率环比下降利多、上升利空；炼油利润越低越抑制供应，偏利多。",
    "demand_seasonality_agent": "产销率是主评分项，按D1/D3/W1/M1选择当日、3日均值、7日均值或月度均值；再叠加季节性、节假日和月度补库节奏。",
    "refined_oil_news_agent": "优先使用LLM从成品油日评/资讯抽取成交活跃度、贸易商心态、报价行为标签，再由代码固定映射分值；无标签时才用新闻文本兜底。",
    "shandong_spot_jump_agent": "只在炼厂报价、成交重心、低价资源扫空、封单惜售、抢货/抛货、出货节奏等硬信号出现时触发；产销率只做确认或否决，不单独重复打分。",
    "policy_cycle_agent": "发改委调价预测金额按上调/下调分档，临近调价窗口放大影响；上轮调价残余效应按幅度折算。",
    "event_risk_agent": "LLM或规则识别地缘、黑天鹅、突发事件等级和方向；高等级事件作为风险门控，不作为日常价格主驱动。",
    "agent_judge_agent": "裁判不参与综合分加权，负责检查硬数据、软信号和业务基准是否冲突；必要时调整点位、扩宽区间或降低置信度。",
}

FEATURE_LABELS = {
    "brent_change_usd_d1": "Brent预测涨跌D1",
    "brent_change_usd_d3": "Brent预测涨跌D3",
    "brent_change_usd_w1": "Brent预测涨跌W1",
    "brent_change_usd_m1": "Brent预测涨跌M1",
    "shandong_cdu_utilization_weekly": "山东地炼常减压开工率",
    "shandong_refinery_load_news_adjustment_d1": "山东炼厂负荷新闻修正",
    "sales_production_ratio_d1": "山东地炼汽油产销率D1",
    "sales_production_ratio_d3_avg": "山东地炼汽油产销率3日均值",
    "sales_production_ratio_w1_avg": "山东地炼汽油产销率7日均值",
    "sales_production_ratio_monthly_avg": "山东地炼汽油产销率月度均值",
    "trader_sentiment_label_d1": "贸易商/成交情绪D1",
    "trader_sentiment_label_d3": "贸易商/成交情绪D3",
    "market_sentiment_monthly": "月度市场情绪",
    "monthly_seasonality_phase": "下月淡旺季阶段",
    "restocking_rhythm_monthly": "补库节奏",
    "holiday_demand_delta_monthly": "下月节假日需求变化",
    "next_month_maintenance_plan": "次月检修计划",
    "price_window_expectation_weekly": "调价窗口预测金额W1",
    "price_window_expectation_monthly": "调价窗口预测金额M1",
    "shandong_product_inventory_percentile_weekly": "山东库存合计分位",
    "refinery_inventory_monthly": "山东独立炼厂库存分位",
    "main_company_inventory_monthly": "主营销售公司库存分位",
}

GROUP_LABELS = {
    "cost": "成本端",
    "supply": "供给端",
    "demand": "需求端",
    "sentiment": "情绪端",
    "inventory": "库存端",
    "policy": "政策端",
    "seasonality": "季节性",
}

SOURCE_HINTS = {
    "brent_change_usd": "Brent预测日报；取预测点位/预测涨跌与Brent特征结算价形成特征",
    "shandong_cdu_utilization": "手工/钢联Excel导入或隆众周度数据；当前按原始开工率百分比打分，不用分位",
    "sales_production_ratio": "山东独立炼厂汽油产销率；D1取当日/上一可用日，D3取3日均值，W1取7日均值，M1取月度均值",
    "trader_sentiment": "成品油日评/资讯经LLM抽取成交活跃度、贸易商心态、报价行为，再由代码固定打分",
    "market_sentiment": "月度市场情绪标签，LLM抽取标签，代码固定映射分值",
    "monthly_seasonality": "系统日历规则：按下月进入旺季/次旺季/次淡季/淡季打分",
    "holiday": "节假日统计：判断下月节假日数量是否增加",
    "maintenance": "隆众检修计划归档，结构化后判断集中检修/复工",
    "price_window": "发改委调价预测金额/调价窗口历史",
    "inventory": "贸易商+主营+独立炼厂库存合计或相应库存分位；缺失时按0分",
}

BUSINESS_BUCKET_RULES = {
    "business_scorecard": [
        ("强空", "<=-70"),
        ("偏空", "-70~-35"),
        ("弱空", "-35~-10"),
        ("震荡", "-10~10"),
        ("弱多", "10~35"),
        ("偏多", "35~70"),
        ("强多", ">=70"),
    ],
    "agent_composite": [
        ("强空", "<=-12"),
        ("偏空", "-12~-6"),
        ("弱空", "-6~-2"),
        ("震荡", "-2~2"),
        ("弱多", "2~6"),
        ("偏多", "6~12"),
        ("强多", ">=12"),
    ],
}


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def short_json(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def round_number(value: Any, digits: int = 2) -> Any:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return clean(value)
    if math.isnan(numeric) or math.isinf(numeric):
        return ""
    return round(numeric, digits)


def source_hint(feature_name: str) -> str:
    for key, text in SOURCE_HINTS.items():
        if key in feature_name:
            return text
    return "系统特征表/规则计算；缺失不做方向假设，按0分处理"


def unit_for(feature_name: str) -> str:
    if "brent" in feature_name:
        return "美元/桶"
    if "ratio" in feature_name or "utilization" in feature_name or "percentile" in feature_name:
        return "%"
    if "price_window" in feature_name or "adjustment" in feature_name:
        return "元/吨"
    return ""


def feature_rule_text(feature: dict[str, Any]) -> str:
    method = feature.get("method")
    if method == "bucket_score":
        label = feature.get("matched_label") or ""
        return f"分档打分，当前命中：{label}。详见YAML规则；表中展示当前命中值和分数。"
    if method == "enum_score":
        return f"标签映射打分，当前标签：{feature.get('matched_label') or feature.get('value')}"
    if method == "bounded_numeric":
        return "数值修正项，按上下限截断；缺失时按0分。"
    return f"{method or '规则'}；缺失时按0分。"


def data_status_text(status: str) -> str:
    return "缺失，按0分" if status == "missing" else "已取数"


def write_rows(ws, rows: list[list[Any]], start_row: int = 1) -> None:
    for row_index, row in enumerate(rows, start=start_row):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index, clean(value))


def title(ws, text: str, width: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = ws.cell(1, 1, text)
    cell.font = Font(size=15, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="C2410C")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def add_table(ws, headers: list[str], rows: list[list[Any]], start_row: int) -> int:
    write_rows(ws, [headers], start_row=start_row)
    write_rows(ws, rows, start_row=start_row + 1)
    return start_row + len(rows) + 2


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    subheader_fill = PatternFill("solid", fgColor="FED7AA")
    header_font = Font(color="FFFFFF", bold=True)
    border_side = Side(style="thin", color="D1D5DB")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for ws in wb.worksheets:
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row == 1:
                    continue
                if cell.row in {3, 10, 18, 26} and cell.value:
                    cell.fill = subheader_fill
                    cell.font = Font(bold=True, color="7C2D12")
        for row in ws.iter_rows():
            first_value = str(row[0].value or "")
            if first_value.startswith("一、") or first_value.startswith("二、") or first_value.startswith("三、"):
                for cell in row:
                    cell.fill = subheader_fill
                    cell.font = Font(bold=True, color="7C2D12")
        for cell in ws[2]:
            if cell.value:
                cell.fill = PatternFill("solid", fgColor="FFF7ED")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            is_header = all(cell.value for cell in row[: min(ws.max_column, 4)]) and row[0].row != 1
            if is_header and row[0].row not in {2}:
                values = [str(cell.value or "") for cell in row]
                if any(text in values[0] for text in ("周期", "智能体", "模块", "区域", "字段", "项目")):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for item in cell:
                    value = str(item.value or "")
                    max_len = max(max_len, min(len(value), 60))
            width = min(max(max_len + 3, 12), 48)
            if col_idx in {1, 2, 3}:
                width = min(max(width, 16), 30)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 24


def mapping_rows(mapping: dict[str, Any], current_price: float, point_value: float) -> list[list[Any]]:
    return [
        ["点位换算方法", "打分进入状态桶，取历史同状态真实涨跌分布；预测涨跌取P50中位数，区间用P25/P75形成半宽"],
        ["当前分数", mapping.get("score_value")],
        ["换算分数点", mapping.get("score_points")],
        ["状态桶", mapping.get("bucket")],
        ["状态桶区间", mapping.get("bucket_range")],
        ["实际使用桶", "、".join(mapping.get("selected_buckets") or [])],
        ["历史样本数", mapping.get("history_sample_size")],
        ["本次桶样本数", mapping.get("sample_size")],
        ["P10涨跌", mapping.get("p10_delta")],
        ["P25涨跌", mapping.get("p25_delta")],
        ["P50涨跌/预测涨跌", mapping.get("p50_delta")],
        ["P75涨跌", mapping.get("p75_delta")],
        ["P90涨跌", mapping.get("p90_delta")],
        ["当前价", current_price],
        ["预测点位", point_value],
        ["公式", f"预测点位 = 当前价 {current_price} + 预测涨跌 {mapping.get('predicted_delta')} = {point_value}"],
        ["状态说明", mapping.get("reason")],
        ["方向约束", "已触发" if mapping.get("semantic_constraint_applied") else "未触发"],
        ["同向保守分位", (mapping.get("directional_fallback") or {}).get("reason")],
    ]


def build_workbook(payload: dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    predictions = payload["predictions"]
    context = payload["context"]
    row = payload["row"]
    d1 = predictions[0]

    ws = wb.create_sheet("总览")
    title(ws, "两种预测逻辑用数及换算说明（当前系统规则版）", 10)
    write_rows(
        ws,
        [
            ["说明", "本工作簿使用当前系统真实预测输出重建，展示用2026-06-05数据预测2026-06-08山东92#市场现汇价的两套逻辑。"],
            ["关键结论", "业务基准打分模型独立输出点位，不参与智能体综合分加权；智能体综合预测使用专家先验权重形成综合分，再进入状态桶换算点位。"],
        ],
        start_row=2,
    )
    overview_rows = [
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["输入日期", context.get("as_of_date")],
        ["价格锚点日期", context.get("price_anchor_date")],
        ["当前山东92#市场现汇价", context.get("sd_gas92_market")],
        ["实际2026-06-08价格", ACTUAL_20260608_PRICE],
        ["数据模式", context.get("market_data_mode")],
        ["数据原因", context.get("market_data_reason")],
        ["ETA状态", "已禁用；历史eta_market_snapshot数据已删除，价格锚点优先使用隆众/钢联/手工导入等现汇价来源。"],
    ]
    next_row = add_table(ws, ["项目", "内容"], overview_rows, 5)
    comparison_rows = []
    for item in predictions:
        rc = item["raw_context"]
        business = rc.get("business_scorecard_prediction") or {}
        comparison_rows.append(
            [
                item["horizon"],
                HORIZON_LABELS.get(item["horizon"], item["horizon"]),
                item["target_date"],
                round_number(item["point_value"]),
                round_number(float(item["point_value"]) - ACTUAL_20260608_PRICE) if item["horizon"] == "D1" else "",
                DIRECTION_LABELS.get(item["direction_label"], item["direction_label"]),
                f"{round_number(item['range_lower'])} ~ {round_number(item['range_upper'])}",
                round_number(item["score_value"], 4),
                round_number(business.get("point_value")),
                round_number(float(business.get("point_value")) - ACTUAL_20260608_PRICE) if item["horizon"] == "D1" and business.get("point_value") is not None else "",
                DIRECTION_LABELS.get(business.get("direction_label"), business.get("direction_label")),
                business.get("score"),
            ]
        )
    add_table(
        ws,
        ["周期", "周期含义", "目标日", "智能体预测点位", "D1误差", "智能体趋势", "智能体区间", "综合分", "业务预测点位", "D1误差", "业务趋势", "业务总分"],
        comparison_rows,
        next_row,
    )

    ws = wb.create_sheet("智能体综合预测")
    title(ws, "智能体综合预测：用数、赋值、换算", 14)
    summary_rows = []
    for item in predictions:
        mapping = item["raw_context"].get("point_mapping") or {}
        summary_rows.append(
            [
                item["horizon"],
                HORIZON_LABELS.get(item["horizon"], item["horizon"]),
                item["as_of_date"],
                item["target_date"],
                context.get("sd_gas92_market"),
                item["raw_context"].get("predicted_delta"),
                item["point_value"],
                f"{round_number(item['range_lower'])} ~ {round_number(item['range_upper'])}",
                DIRECTION_LABELS.get(item["direction_label"], item["direction_label"]),
                item["score_value"],
                mapping.get("bucket"),
                mapping.get("sample_size"),
                mapping.get("p50_delta"),
                item["confidence_label"],
            ]
        )
    add_table(
        ws,
        ["周期", "含义", "输入日", "目标日", "当前价", "预测涨跌", "预测点位", "预测区间", "趋势", "综合分", "状态桶", "桶样本数", "P50涨跌", "置信度"],
        summary_rows,
        3,
    )

    ws = wb.create_sheet("智能体逐项打分")
    title(ws, "智能体逐项赋值与打分明细", 15)
    agent_rows = []
    for item in predictions:
        for claim in item["agent_claims"]:
            name = claim["agent_name"]
            if name == "business_scorecard_agent":
                continue
            signals = claim.get("numeric_signals") or {}
            payload_item = claim.get("structured_payload") or {}
            dq = payload_item.get("data_quality") or {}
            runtime = payload_item.get("runtime_control") or {}
            values = {k: v for k, v in payload_item.items() if k not in {"data_quality", "runtime_control", "llm_labels", "agent_judgement"}}
            if name == "refined_oil_news_agent":
                labels = payload_item.get("llm_labels") or {}
                values = {
                    "成交活跃度": labels.get("deal_activity") or payload_item.get("label"),
                    "贸易商心态": labels.get("trader_mindset") or payload_item.get("trader_mindset"),
                    "报价行为": labels.get("quote_behavior") or payload_item.get("quote_behavior"),
                    "证据": labels.get("evidence"),
                }
            if name == "agent_judge_agent":
                judgement = payload_item.get("agent_judgement") or {}
                values = {
                    "裁判结论": judgement.get("display_label"),
                    "硬数据支持": judgement.get("hard_support"),
                    "硬数据反对": judgement.get("hard_counter"),
                    "软信号支持": judgement.get("soft_support"),
                    "软信号反对": judgement.get("soft_counter"),
                    "点位调整": judgement.get("adjustment_delta"),
                    "区间扩宽": judgement.get("range_extra_width"),
                }
            agent_rows.append(
                [
                    item["horizon"],
                    AGENT_LABELS.get(name, name),
                    AGENT_RULES.get(name, ""),
                    round_number(signals.get("raw_score"), 4),
                    round_number(signals.get("max_score"), 4),
                    round_number(signals.get("normalized_score"), 4),
                    round_number(signals.get("expert_prior_weight") if signals.get("expert_prior_weight") is not None else runtime.get("weight"), 4),
                    round_number(signals.get("weighted_score"), 4),
                    DIRECTION_LABELS.get(claim.get("direction"), claim.get("direction")),
                    claim.get("summary"),
                    "；".join(claim.get("evidence") or []),
                    clean(values),
                    dq.get("coverage_ratio"),
                    "、".join(dq.get("missing_fields") or []),
                ]
            )
    add_table(
        ws,
        ["周期", "智能体", "原版规则/赋值逻辑", "原始分", "满分口径", "归一化分", "专家先验权重", "加权贡献", "方向", "结论", "证据", "本次使用数据", "覆盖率", "缺失字段"],
        agent_rows,
        3,
    )

    ws = wb.create_sheet("业务基准预测")
    title(ws, "业务基准打分模型：独立预测结果", 14)
    business_rows = []
    for item in predictions:
        business = item["raw_context"].get("business_scorecard_prediction") or {}
        mapping = business.get("mapping") or {}
        business_rows.append(
            [
                item["horizon"],
                HORIZON_LABELS.get(item["horizon"], item["horizon"]),
                item["as_of_date"],
                item["target_date"],
                business.get("current_price"),
                business.get("score"),
                business.get("predicted_delta"),
                business.get("point_value"),
                f"{business.get('range_lower')} ~ {business.get('range_upper')}",
                DIRECTION_LABELS.get(business.get("direction_label"), business.get("direction_label")),
                mapping.get("bucket"),
                mapping.get("selected_buckets"),
                mapping.get("sample_size"),
                mapping.get("p50_delta"),
                business.get("basis"),
            ]
        )
    add_table(
        ws,
        ["周期", "含义", "输入日", "目标日", "当前价", "业务总分", "预测涨跌", "预测点位", "预测区间", "趋势", "状态桶", "实际使用桶", "样本数", "P50涨跌", "应用说明"],
        business_rows,
        3,
    )

    ws = wb.create_sheet("业务逐项打分")
    title(ws, "业务基准模型逐项用数与打分", 13)
    business_detail_rows = []
    for item in predictions:
        scorecard = item["raw_context"].get("business_scorecard") or {}
        for group in scorecard.get("groups") or []:
            for feature in group.get("features") or []:
                feature_name = feature.get("feature_name")
                business_detail_rows.append(
                    [
                        item["horizon"],
                        GROUP_LABELS.get(group.get("group_code"), group.get("display_name")),
                        FEATURE_LABELS.get(feature_name, feature_name),
                        feature_name,
                        round_number(feature.get("value"), 4),
                        unit_for(feature_name),
                        feature.get("method"),
                        feature_rule_text(feature),
                        feature.get("matched_label"),
                        feature.get("score"),
                        group.get("score"),
                        group.get("score_cap"),
                        data_status_text(feature.get("status")),
                        source_hint(feature_name),
                    ]
                )
    add_table(
        ws,
        ["周期", "业务模块", "数据项", "字段编码", "本次取值", "单位", "打分方法", "如何判定", "命中规则", "本项得分", "模块得分", "模块封顶", "取数状态", "数据来源/口径"],
        business_detail_rows,
        3,
    )

    ws = wb.create_sheet("点位换算说明")
    title(ws, "分数如何换算成点位、区间和趋势", 9)
    write_rows(
        ws,
        [
            ["核心原则", "两套逻辑都不使用截距/斜率，不使用人为周期点值上限；均使用状态桶历史分布把分数换成涨跌额。"],
            ["P50含义", "P50是当前状态桶历史真实涨跌额的中位数。本次预测涨跌优先取P50；若P50与状态桶方向冲突，则使用同向样本的保守分位数。"],
            ["P25/P75含义", "P25和P75分别是状态桶历史涨跌额的25%和75%分位，用于生成区间半宽。"],
            ["样本不足", "智能体综合可逐步合并更多状态桶；业务基准只允许合并相邻状态桶，仍不足则进入冷启动专家桶。"],
            ["趋势", "用预测涨跌与周期方向阈值比较得出上涨/下跌/震荡；点位是当前价+预测涨跌。"],
        ],
        start_row=2,
    )
    row_idx = 9
    for item in predictions:
        ws.cell(row_idx, 1, f"{item['horizon']} 智能体综合点位换算")
        row_idx += 1
        row_idx = add_table(
            ws,
            ["项目", "值"],
            mapping_rows(item["raw_context"].get("point_mapping") or {}, float(context.get("sd_gas92_market")), float(item["point_value"])),
            row_idx,
        )
        business = item["raw_context"].get("business_scorecard_prediction") or {}
        ws.cell(row_idx, 1, f"{item['horizon']} 业务基准点位换算")
        row_idx += 1
        row_idx = add_table(
            ws,
            ["项目", "值"],
            mapping_rows(business.get("mapping") or {}, float(context.get("sd_gas92_market")), float(business.get("point_value") or 0.0)),
            row_idx,
        )

    ws = wb.create_sheet("区域价差预测逻辑")
    title(ws, "区域价差预测逻辑（当前说明）", 8)
    write_rows(
        ws,
        [
            ["定位", "区域价差预测已与山东92#价格预测分开：一屏看清展示真实区域价差；研究台/晨报展示预测区域价差。"],
            ["价差口径", "目标区域价格 - 山东价格。区域预测先预测目标区域单价，再用预测目标区域价格 - 预测山东价格得到区域价差。"],
            ["两套逻辑", "区域智能体综合预测与区域业务基准预测并列展示，均不使用历史拟合回归系数。"],
            ["核心用数", "当前区域价差、1日/3日价差变化、区域库存、运费、净回款价差、经营上下边界。"],
            ["经营边界", "区域价差必须受运费、必要毛利和风险缓冲约束，避免给出经营上不可执行的点位。"],
            ["库存影响", "区域库存高于常态压制区域价格，库存低于常态支撑区域价格；库存缺失时不打方向分，只降低置信度。"],
        ],
        start_row=2,
    )

    ws = wb.create_sheet("字段口径清单")
    title(ws, "预测使用字段口径清单", 8)
    field_rows = [
        ["山东92#市场现汇价", context.get("sd_gas92_market"), "元/吨", context.get("price_anchor_date"), "隆众/钢联/手工导入优先；不再使用ETA"],
        ["Brent特征结算价", row.get("brent_active_settlement"), "美元/桶", context.get("as_of_date"), "Wind历史结算价/实时接口"],
        ["D1 Brent预测涨跌", (d1["raw_context"].get("brent_forecast_basis") or {}).get("scorecard_change_usd"), "美元/桶", context.get("as_of_date"), "Brent日报预测点位 - Brent特征结算价"],
        ["山东地炼汽油产销率", row.get("sales_production_ratio_d1"), "%", context.get("as_of_date"), "隆众/手工导入，需求端主评分项"],
        ["山东地炼常减压开工率", row.get("sd_crude_run_weekly"), "%", context.get("as_of_date"), "业务基准D1供给端使用原始开工率，不用分位"],
        ["山东炼油利润", row.get("sd_refining_profit"), "元/吨", context.get("as_of_date"), "供给库存智能体辅助项"],
        ["汽油裂解价差分位", row.get("gasoline_crack_percentile"), "%", context.get("as_of_date"), "保留展示，当前不参与打分"],
        ["库存合计分位", row.get("shandong_product_inventory_percentile_weekly"), "%", context.get("as_of_date"), "保留字段；当前缺失时按0分或不参与"],
    ]
    add_table(ws, ["字段", "本次取值", "单位", "日期", "业务口径"], field_rows, 3)

    style_workbook(wb)
    return wb


def main() -> None:
    payload = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    wb = build_workbook(payload)
    wb.save(OUTPUT)
    loaded = load_workbook(OUTPUT, data_only=False)
    print(f"saved={OUTPUT}")
    print("sheets=" + ",".join(loaded.sheetnames))
    for name in loaded.sheetnames:
        ws = loaded[name]
        print(f"{name}: rows={ws.max_row}, cols={ws.max_column}")


if __name__ == "__main__":
    main()
