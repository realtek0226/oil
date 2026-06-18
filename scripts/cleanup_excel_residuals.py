
from pathlib import Path
from openpyxl import load_workbook

MAIN = Path('outputs') / '逻辑说明.xlsx'
wb=load_workbook(MAIN)
if 'RegionalSpread' in wb.sheetnames:
    ws=wb['RegionalSpread']
    ws.cell(1,1).value='区域价差'
    ws.cell(2,1).value='口径'
    ws.cell(2,2).value='区域价差=区域价格-山东价格；净回款=区域价差-运费。'
    headers=['区域','当前区域价','当前山东价','当前价差','运费','当前净回款']
    for c,h in enumerate(headers,1): ws.cell(3,c).value=h
    regions=['华东','华北','华南','华中','西北','西南','东北']
    for i,name in enumerate(regions,4): ws.cell(i,1).value=name
if 'PolicyTheory' in wb.sheetnames:
    ws=wb['PolicyTheory']
    ws.cell(2,2).value='调价预测理论有效区间应为 -82 ~ 89.5。'
# ensure key notes are present even if generated rows changed
ws=wb['BusinessScoring']
notes={
 'shandong_product_inventory_percentile_weekly':'库存=贸易商+主营+独立炼厂可用项求和；不是三项必须齐全。若某组件上期有数本期缺失，提示缺失并按0；分位从2024-01-01起算。',
 'restocking_rhythm_monthly':'山东M1备货节奏=比较预测月前一月与前二月山东独立炼厂汽油产销率>90%的天数；前一月更多+5，减少-5，持平0。',
 'price_window_expectation_weekly':'调价预测来自卓创成品油调价专栏。',
 'price_window_expectation_monthly':'调价预测来自卓创成品油调价专栏。',
}
for r in range(4, ws.max_row+1):
    field=ws.cell(r,4).value
    if field in notes:
        ws.cell(r,12).value=notes[field]
wb.save(MAIN)
print(MAIN)
