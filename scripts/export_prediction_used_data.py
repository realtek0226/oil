
from __future__ import annotations

import math
import sys
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
warnings.filterwarnings("ignore")

from app.core.container import get_dataset_service, get_predictor
from app.services.predictors.horizons import DEFAULT_HORIZONS

T = {
    "agent_sheet": "\u667a\u80fd\u4f53\u9884\u6d4b\u7528\u6570",
    "business_sheet": "\u4e1a\u52a1\u6253\u5206\u7528\u6570",
    "gas": "\u6c7d\u6cb9",
    "diesel": "\u67f4\u6cb9",
    "d1": "\u6b21\u65e5",
    "w1": "\u4e00\u5468",
    "m1": "\u4e00\u6708",
    "product": "\u54c1\u79cd",
    "horizon": "\u5468\u671f",
    "agent": "\u667a\u80fd\u4f53",
    "module": "\u6a21\u5757",
    "data_item": "\u6570\u636e\u9879",
    "field": "\u5b57\u6bb5\u540d",
    "value": "\u672c\u6b21\u53d6\u503c",
    "raw_value": "\u672c\u6b21\u539f\u59cb\u53d6\u503c",
    "date": "\u6570\u636e\u5bf9\u5e94\u65e5\u671f",
    "score": "\u5f97\u5206/\u8d21\u732e",
    "cap": "\u6ee1\u5206/\u4e0a\u9650",
    "result": "\u7ed3\u8bba/\u547d\u4e2d",
    "how": "\u5177\u4f53\u600e\u4e48\u7528",
    "evidence": "\u8bc1\u636e/\u53e3\u5f84",
    "source": "\u6765\u6e90/\u8bf4\u660e",
    "brent": "Brent\u9884\u6d4b\u4e0e\u7ed3\u7b97\u4ef7",
    "report_date": "Brent\u65e5\u62a5\u65e5\u671f",
    "event_gate": "\u4e8b\u4ef6\u98ce\u9669\u95e8\u63a7",
    "trade_sentiment": "\u6210\u4ea4/\u8d38\u6613\u5546\u60c5\u7eea",
    "monthly_sentiment": "\u6708\u5ea6\u5e02\u573a\u60c5\u7eea",
    "maintenance": "\u68c0\u4fee\u8ba1\u5212",
    "inventory_note": "\u5c71\u4e1c\u5e93\u5b58\u53e3\u5f84\u4e0d\u5305\u542b\u8d38\u6613\u5546\u5e93\u5b58\uff1b\u5c71\u4e1c\u53ea\u7528\u4e3b\u8425\u9500\u552e\u516c\u53f8\u5e93\u5b58+\u72ec\u7acb\u70bc\u5382\u5382\u5185\u5e93\u5b58\u3002",
}
HORIZON = {"D1": T["d1"], "W1": T["w1"], "M1": T["m1"]}
PRODUCTS = [("GASOLINE_92", T["gas"]), ("DIESEL_0", T["diesel"])]
AGENT_LABELS = {
    "business_scorecard_agent": "\u4e1a\u52a1\u57fa\u51c6\u6253\u5206\u6a21\u578b",
    "crude_cost_agent": "\u539f\u6cb9\u6210\u672c\u667a\u80fd\u4f53",
    "market_structure_agent": "\u5e02\u573a\u7ed3\u6784\u667a\u80fd\u4f53",
    "supply_inventory_agent": "\u4f9b\u7ed9\u5e93\u5b58\u667a\u80fd\u4f53",
    "demand_seasonality_agent": "\u9700\u6c42\u5b63\u8282\u667a\u80fd\u4f53",
    "refined_oil_news_agent": "\u6210\u54c1\u6cb9\u8d44\u8baf\u667a\u80fd\u4f53",
    "shandong_spot_jump_agent": "\u5c71\u4e1c\u73b0\u8d27\u8df3\u53d8\u8bc6\u522b\u667a\u80fd\u4f53",
    "policy_cycle_agent": "\u653f\u7b56\u5468\u671f\u667a\u80fd\u4f53",
    "event_risk_agent": "\u4e8b\u4ef6\u98ce\u9669\u667a\u80fd\u4f53",
    "agent_judge_agent": "\u667a\u80fd\u4f53\u88c1\u5224",
}
FEATURE_LABELS = {
    "brent_change_usd_d1": "Brent\u65e5\u62a5D1\u9884\u6d4b\u53d8\u5316",
    "brent_change_usd_w1": "Brent\u5468\u62a5W1\u9884\u6d4b\u53d8\u5316",
    "brent_change_usd_mom": "Brent\u5468\u62a5W4/M1\u9884\u6d4b\u53d8\u5316",
    "shandong_cdu_utilization_weekly": "\u5c71\u4e1c\u5730\u70bc\u5e38\u51cf\u538b\u5f00\u5de5\u7387",
    "shandong_product_inventory_percentile_weekly": "\u5c71\u4e1c\u5e93\u5b58\u5408\u8ba1\u5206\u4f4d",
    "refinery_inventory_monthly": "\u5c71\u4e1c\u72ec\u7acb\u70bc\u5382\u5e93\u5b58\u5206\u4f4d",
    "main_company_inventory_monthly": "\u5c71\u4e1c\u4e3b\u8425\u9500\u552e\u516c\u53f8\u5e93\u5b58\u5206\u4f4d",
    "sales_production_ratio_d1": "\u5c71\u4e1c\u72ec\u7acb\u70bc\u5382\u4ea7\u9500\u7387",
    "sales_production_ratio_w1_avg": "\u5c71\u4e1c\u72ec\u7acb\u70bc\u5382\u4ea7\u9500\u73877\u65e5\u5747\u503c",
    "trader_sentiment_label_d1": "\u6210\u4ea4\u6d3b\u8dc3\u5ea6/\u8d38\u6613\u5546\u5fc3\u6001",
    "market_sentiment_monthly": "\u6708\u5ea6\u5e02\u573a\u60c5\u7eea",
    "next_month_maintenance_plan": "\u4e0b\u6708\u68c0\u4fee\u8ba1\u5212",
    "refinery_maintenance_plan_adjustment_d1": "\u70bc\u5382\u68c0\u4fee/\u8d1f\u8377\u4fee\u6b63",
    "refinery_maintenance_plan_adjustment_w1": "\u70bc\u5382\u68c0\u4fee/\u8d1f\u8377\u4fee\u6b63",
    "price_window_expectation_weekly": "\u8c03\u4ef7\u7a97\u53e3\u9884\u671f",
    "price_window_expectation_monthly": "\u8c03\u4ef7\u7a97\u53e3\u9884\u671f",
}

SOURCE_CODE_ALIASES = {
    "shandong_cdu_utilization_weekly": [("ganglian_excel_import", ["shandong_cdu_utilization_weekly", "ganglian_id01374956"]), ("zhonglu_excel_archive", ["shandong_cdu_utilization_weekly", "zhonglu_id01374956"])],
    "shandong_product_inventory_percentile_weekly": [("ganglian_excel_import", ["shandong_independent_refinery_inventory", "ganglian_id01374817"]), ("zhonglu_excel_archive", ["shandong_independent_refinery_inventory", "zhonglu_id01374817"])],
    "shandong_diesel_product_inventory_percentile_weekly": [("ganglian_excel_import", ["shandong_diesel_inventory", "ganglian_id01374828"]), ("zhonglu_excel_archive", ["shandong_diesel_inventory", "zhonglu_id01374828"])],
    "refinery_inventory_monthly": [("ganglian_excel_import", ["shandong_independent_refinery_inventory", "ganglian_id01374817"]), ("zhonglu_excel_archive", ["shandong_independent_refinery_inventory", "zhonglu_id01374817"])],
    "diesel_refinery_inventory_monthly": [("ganglian_excel_import", ["shandong_diesel_inventory", "ganglian_id01374828"]), ("zhonglu_excel_archive", ["shandong_diesel_inventory", "zhonglu_id01374828"])],
    "main_company_inventory_monthly": [("oilchem_openapi_inventory", ["oilchem_openapi_inventory_12887"])],
    "diesel_main_company_inventory_monthly": [("oilchem_openapi_inventory", ["oilchem_openapi_inventory_12891"])],
    "sales_production_ratio_d1": [("oilchem_production_sales_ratio", ["oilchem_sd_gasoline_production_sales_ratio"])],
    "diesel_sales_production_ratio_d1": [("oilchem_production_sales_ratio", ["oilchem_sd_diesel_production_sales_ratio"])],
}


def source_latest_date(dataset: Any, key: str, as_of: date) -> str:
    repo = getattr(dataset, "snapshot_repository", None)
    if repo is None:
        return ""
    rows_all = []
    for source_code, codes in SOURCE_CODE_ALIASES.get(key, []):
        try:
            rows = repo.load_market_timeseries_values(source_code=source_code, indicator_codes=list(codes), start_date=date(2024, 1, 1), end_date=as_of)
        except Exception:
            rows = []
        rows_all.extend([row for row in rows if row.get("value_num") is not None and row.get("dt") is not None])
    if not rows_all:
        return ""
    return max(pd.Timestamp(row["dt"]).date() for row in rows_all).isoformat()


def clean(v: Any) -> Any:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if hasattr(v, "item"):
        try:
            return clean(v.item())
        except Exception:
            pass
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return ""
        return round(v, 4)
    return v


def latest_date(frame: pd.DataFrame, as_of: date, cols: list[str]) -> str:
    if frame.empty or "date" not in frame.columns:
        return ""
    data = frame[pd.to_datetime(frame["date"]).dt.date <= as_of]
    for col in cols:
        if col not in data.columns:
            continue
        subset = data[data[col].notna()]
        if not subset.empty:
            return pd.to_datetime(subset.iloc[-1]["date"]).date().isoformat()
    return ""


def value_date(feature_name: str, product_code: str, context: Any, pred: Any, dataset: Any | None = None) -> str:
    if feature_name.startswith("brent_change_usd") or "Brent" in feature_name:
        return str((context.report_payload or {}).get("report_date") or "")
    if feature_name == "shandong_cdu_utilization_weekly" or "\u5e38\u51cf\u538b" in feature_name or "\u5f00\u5de5\u7387" in feature_name:
        obs = context.current_row.get("shandong_cdu_utilization_observation_date")
        if obs is not None:
            try:
                obs_float = float(obs)
                if abs(obs_float) < 10000000000:
                    return pd.to_datetime(obs_float, unit="s").date().isoformat()
                return pd.Timestamp(obs).date().isoformat()
            except Exception:
                return str(obs)[:10]
    if "maintenance" in feature_name or "\u68c0\u4fee" in feature_name:
        plan = (context.metadata or {}).get("oilchem_maintenance_plan") or {}
        return str(plan.get("observation_date") or plan.get("report_date") or "")[:10]
    if "sentiment" in feature_name or "\u60c5\u7eea" in feature_name:
        items = context.refined_news_items or []
        ds = [str((x or {}).get("publish_time") or (x or {}).get("date") or "")[:10] for x in items if isinstance(x, dict)]
        return max([d for d in ds if d], default="")
    if "\u4ea7\u9500\u7387" in feature_name:
        source_key_for_ratio = "diesel_sales_production_ratio_d1" if product_code == "DIESEL_0" else "sales_production_ratio_d1"
        if dataset is not None:
            source_date = source_latest_date(dataset, source_key_for_ratio, pred.as_of_date)
            if source_date:
                return source_date
        return latest_date(context.feature_frame, pred.as_of_date, [source_key_for_ratio])
    source_key = feature_name
    if product_code == "DIESEL_0" and feature_name == "shandong_product_inventory_percentile_weekly":
        source_key = "shandong_diesel_product_inventory_percentile_weekly"
    elif product_code == "DIESEL_0" and feature_name == "refinery_inventory_monthly":
        source_key = "diesel_refinery_inventory_monthly"
    elif product_code == "DIESEL_0" and feature_name == "main_company_inventory_monthly":
        source_key = "diesel_main_company_inventory_monthly"
    elif product_code == "DIESEL_0" and feature_name in {"sales_production_ratio_d1", "sales_production_ratio_w1_avg"}:
        source_key = "diesel_sales_production_ratio_d1"
    if dataset is not None:
        source_date = source_latest_date(dataset, source_key, pred.as_of_date)
        if source_date:
            return source_date
    mapping = {
        "shandong_cdu_utilization_weekly": ["shandong_cdu_utilization_weekly"],
        "shandong_product_inventory_percentile_weekly": ["shandong_diesel_product_inventory_total_formal"] if product_code == "DIESEL_0" else ["shandong_product_inventory_total_formal"],
        "refinery_inventory_monthly": ["shandong_diesel_inventory"] if product_code == "DIESEL_0" else ["shandong_independent_refinery_inventory"],
        "main_company_inventory_monthly": ["shandong_main_company_diesel_inventory"] if product_code == "DIESEL_0" else ["shandong_main_company_inventory"],
        "sales_production_ratio_d1": ["diesel_sales_production_ratio_d1"] if product_code == "DIESEL_0" else ["sales_production_ratio_d1"],
        "sales_production_ratio_w1_avg": ["diesel_sales_production_ratio_d1"] if product_code == "DIESEL_0" else ["sales_production_ratio_d1"],
    }
    return latest_date(context.feature_frame, pred.as_of_date, mapping.get(feature_name, [feature_name]))


def raw_value(feature_name: str, product_code: str, context: Any, pred: Any, feature: dict[str, Any] | None = None) -> str:
    feature = feature or {}
    row = context.current_row
    if feature_name.startswith("brent_change_usd"):
        basis = (pred.raw_context or {}).get("brent_forecast_basis") or {}
        return f"{T['report_date']}={basis.get('report_date')}; forecast={basis.get('forecast_point_usd')}; settlement={basis.get('brent_settlement_usd')}; change={basis.get('scorecard_change_usd')}"
    if feature_name == "shandong_cdu_utilization_weekly":
        return f"{clean(row.get('shandong_cdu_utilization_weekly'))}%; percentile={clean(row.get('shandong_cdu_utilization_percentile_weekly'))}%"
    if feature_name == "shandong_product_inventory_percentile_weekly":
        if product_code == "DIESEL_0":
            return f"{clean(row.get('shandong_diesel_product_inventory_total_formal'))}\u4e07\u5428; percentile={clean(row.get('shandong_diesel_product_inventory_percentile_weekly'))}%"
        return f"{clean(row.get('shandong_product_inventory_total_formal'))}\u4e07\u5428; percentile={clean(row.get('shandong_product_inventory_percentile_weekly'))}%; {T['inventory_note']}"
    if feature_name == "refinery_inventory_monthly":
        if product_code == "DIESEL_0":
            return f"{clean(row.get('shandong_diesel_inventory'))}\u4e07\u5428; percentile={clean(row.get('shandong_diesel_refinery_inventory_percentile_monthly'))}%"
        return f"{clean(row.get('shandong_independent_refinery_inventory'))}\u4e07\u5428; percentile={clean(row.get('shandong_refinery_inventory_percentile_monthly'))}%"
    if feature_name == "main_company_inventory_monthly":
        if product_code == "DIESEL_0":
            return f"{clean(row.get('shandong_main_company_diesel_inventory'))}\u4e07\u5428; percentile={clean(row.get('shandong_diesel_main_company_inventory_percentile_monthly'))}%"
        return f"{clean(row.get('shandong_main_company_inventory'))}\u4e07\u5428; percentile={clean(row.get('shandong_main_company_inventory_percentile_monthly'))}%"
    return str(clean(feature.get("value") if feature else ""))


def business_rows(pred: Any, context: Any, product_code: str, product_label: str, dataset: Any) -> list[list[Any]]:
    sc = (pred.raw_context or {}).get("business_scorecard") or {}
    rows = []
    for group in sc.get("groups") or []:
        for feature in group.get("features") or []:
            name = str(feature.get("feature_name") or "")
            rows.append([
                product_label,
                HORIZON.get(str(sc.get("horizon") or pred.horizon), pred.horizon),
                group.get("display_name") or group.get("group_code"),
                FEATURE_LABELS.get(name, feature.get("display_name") or name),
                name,
                raw_value(name, product_code, context, pred, feature),
                value_date(name, product_code, context, pred, dataset),
                clean(feature.get("value")),
                clean(feature.get("score")),
                clean(feature.get("score_cap") or group.get("score_cap")),
                feature.get("matched_label") or feature.get("status") or "",
                feature.get("raw_rule") or feature.get("rule_text") or "",
                "\n".join(str(x) for x in (feature.get("evidence") or []) if x),
            ])
    return rows


def agent_rows(pred: Any, context: Any, product_code: str, product_label: str, dataset: Any) -> list[list[Any]]:
    raw = pred.raw_context or {}
    rows = []
    for item in ((raw.get("agent_judgement") or {}).get("review_items") or []):
        agent_name = str(item.get("agent_name") or "")
        evidence = item.get("evidence") or []
        evidence_text = "\n".join(str(x) for x in evidence if x)
        rows.append([
            product_label,
            HORIZON.get(pred.horizon, pred.horizon),
            AGENT_LABELS.get(agent_name, agent_name),
            item.get("summary") or item.get("evidence_type") or "",
            clean(item.get("contribution")),
            value_date(evidence_text, product_code, context, pred, dataset),
            item.get("direction") or "",
            item.get("relation") or "",
            evidence_text,
            "\u7528\u8be5\u667a\u80fd\u4f53\u7684\u786c\u6570\u636e/\u8f6f\u4fe1\u53f7\u5224\u65ad\u65b9\u5411\uff0c\u8fdb\u5165\u667a\u80fd\u4f53\u88c1\u5224\u5c42\u505a\u51b2\u7a81\u68c0\u67e5\u3001\u7f6e\u4fe1\u5ea6\u548c\u533a\u95f4\u590d\u6838\u3002",
        ])
    # Add hard feature snapshot rows so business users can see concrete values.
    snap = raw.get("agent_business_feature_snapshot") or {}
    for k, v in (snap.get("\u786c\u6570\u636e") or {}).items():
        rows.append([
            product_label,
            HORIZON.get(pred.horizon, pred.horizon),
            "\u667a\u80fd\u4f53\u786c\u6570\u636e\u5f15\u7528",
            k,
            clean(v),
            value_date(str(k), product_code, context, pred, dataset),
            "",
            "",
            "\u667a\u80fd\u4f53\u89e3\u91ca\u94fe\u5f15\u7528\u7684\u786c\u6570\u636e\u3002",
            "\u4f5c\u4e3a\u667a\u80fd\u4f53\u65b9\u5411\u5224\u65ad\u548c\u88c1\u5224\u5c42\u51b2\u7a81\u68c0\u67e5\u7684\u8bc1\u636e\uff0c\u4e0d\u7b49\u540c\u4e8e\u4e1a\u52a1\u5206\u76f4\u63a5\u52a0\u6743\u3002",
        ])
    # Add explicit event gate if triggered.
    gate = raw.get("event_gate") or {}
    if gate:
        rows.append([
            product_label, HORIZON.get(pred.horizon, pred.horizon), T["event_gate"], T["event_gate"],
            f"level={gate.get('level')}; direction={(gate.get('llm_risk_gate') or {}).get('direction')}; overlay={(raw.get('point_adjustments') or {}).get('event_cost_overlay')}",
            str(raw.get("event_report_date") or raw.get("prediction_news_cutoff") or "")[:10],
            "", "", str(gate),
            "\u9ad8\u7b49\u7ea7\u4e8b\u4ef6\u4f5c\u4e3a\u667a\u80fd\u4f53\u95e8\u63a7/\u533a\u95f4\u4e0e\u7f6e\u4fe1\u5ea6\u590d\u6838\u4f9d\u636e\uff0c\u4e0d\u76f4\u63a5\u6539\u4e1a\u52a1\u6253\u5206\u70b9\u4f4d\u3002",
        ])
    return rows


def style(wb: Workbook) -> None:
    fill = PatternFill("solid", fgColor="F97316")
    font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for c in ws[1]:
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for c in row:
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for i in range(1, ws.max_column + 1):
            width = 16
            if i in {4, 5, 6, 9, 10, 12, 13}:
                width = 36
            if i in {11, 12, 13}:
                width = 46
            ws.column_dimensions[get_column_letter(i)].width = width


def main() -> None:
    dataset = get_dataset_service()
    predictor = get_predictor()
    as_of = dataset.resolve_default_prediction_as_of(date.today())
    context = dataset.build_context(as_of)
    horizons = [h for h in DEFAULT_HORIZONS if h != "D3"]
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = T["agent_sheet"]
    ws_b = wb.create_sheet(T["business_sheet"])
    ws_a.append([T["product"], T["horizon"], T["agent"], T["data_item"], T["value"], T["date"], T["result"], T["score"], T["evidence"], T["how"]])
    ws_b.append([T["product"], T["horizon"], T["module"], T["data_item"], T["field"], T["raw_value"], T["date"], "\u5165\u6876/\u6807\u7b7e\u503c", T["score"], T["cap"], T["result"], T["how"], T["evidence"]])
    for product_code, product_label in PRODUCTS:
        for h in horizons:
            if product_code == "DIESEL_0":
                pred = predictor.run_diesel0_prediction_from_context(context, as_of_date=as_of, horizon=h, use_llm_explainer=False)
            else:
                pred = predictor.run_prediction_from_context(context, as_of_date=as_of, horizon=h, use_llm_explainer=False)
            for r in agent_rows(pred, context, product_code, product_label, dataset):
                ws_a.append([clean(x) for x in r])
            for r in business_rows(pred, context, product_code, product_label, dataset):
                ws_b.append([clean(x) for x in r])
    style(wb)
    out = ROOT / "outputs" / ("\u9884\u6d4b\u7528\u6570\u660e\u7ec6_" + as_of.strftime("%Y%m%d") + ".xlsx")
    try:
        wb.save(out)
    except PermissionError:
        out = ROOT / "outputs" / ("\u9884\u6d4b\u7528\u6570\u660e\u7ec6_" + as_of.strftime("%Y%m%d") + "_new.xlsx")
        wb.save(out)
    print(out)

if __name__ == "__main__":
    main()
