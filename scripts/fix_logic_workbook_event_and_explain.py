
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
p=Path('outputs')/'逻辑说明.xlsx'
wb=load_workbook(p)
# Explain / rename detail sheet
if 'BusinessScoring_Detail' in wb.sheetnames:
    ws=wb['BusinessScoring_Detail']
    ws.title='业务打分明细_可删'
    ws.cell(1,1).value='业务打分明细（可删）'
    ws.cell(2,1).value='说明'
    ws.cell(2,2).value='BusinessScoring 是给业务看的合并表：同一周期同一模块合并为一行；本表是字段级追溯明细，检查用，业务汇报可删除。'
if 'BusinessScoring' in wb.sheetnames:
    ws=wb['BusinessScoring']
    ws.cell(2,2).value='给业务看的合并表：同一周期同一模块合并为一行；字段级追溯见“业务打分明细_可删”。'
# Ensure event risk raw rule and visible evidence
ws=wb['AgentScoring']
headers=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
raw_col=headers.index('原始规则')+1 if '原始规则' in headers else None
for r in range(4,ws.max_row+1):
    if ws.cell(r,2).value=='事件风险智能体':
        ws.cell(r,3).value='美伊停战宣布霍尔木兹海峡解封，航运恢复；event_type=supply_relief；direction=down；severity=high'
        ws.cell(r,9).value='供应风险缓和，事件门控利空'
        if raw_col:
            ws.cell(r,raw_col).value='原始规则：事件风险为门控/裁判层输入，权重=0；命中“停战/霍尔木兹解封/通航恢复/解除封锁”时强制识别为supply_relief、direction=down、severity=high，并要求人工复核，不允许被裁判层反向抵消。'
        ws.cell(r,ws.max_column).value='识别到美伊停战、霍尔木兹海峡解封、航运恢复，按供应风险缓和处理，对油价方向为利空。'
# add explanation sheet
if '表格说明' in wb.sheetnames:
    del wb['表格说明']
info=wb.create_sheet('表格说明',0)
info.append(['表格说明'])
info.append(['BusinessScoring','业务阅读用合并表：同一周期同一模块一行，方便看模块如何打分。'])
info.append(['业务打分明细_可删','字段级追溯表：保留每个字段的取值、命中规则和得分，检查用；如果只给业务看，可以删除。'])
info.append(['事件风险修正','已强制识别“美伊停战/霍尔木兹海峡解封/航运恢复”为供应风险缓和：event_type=supply_relief，direction=down，severity=high。'])
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(vertical='top', wrap_text=True)
    if ws.max_row>=3:
        for cell in ws[3]:
            cell.font=Font(bold=True,color='FFFFFF')
            cell.fill=PatternFill('solid',fgColor='1F4E78')
wb.save(p)
print(p)
