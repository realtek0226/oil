
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, time, timezone
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import bindparam, text

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.settings import DatabaseSettings, get_settings
from app.services.postgres_snapshot_repository import PostgresSnapshotRepository

SOURCE_CODE = "ganglian_excel_import"
DEFAULT_EXCEL_NAME = "\u6a21\u578b\u9884\u6d4b\u57fa\u7840\u6570\u636e6.12.xlsx"

GASOLINE_SHEET = "\u6c7d\u6cb9\u4ef7\u683c"
DIESEL_SHEET = "\u67f4\u6cb9\u4ef7\u683c"
DATE_HEADER = "\u65e5\u671f"

PRICE_COLUMNS: dict[str, dict[str, tuple[str, str, str, str, str]]] = {
    GASOLINE_SHEET: {
        "\u5c71\u4e1c\u5e02\u573a\u4ef7": ("sd_gas92_market", "\u5c71\u4e1c92#\u6c7d\u6cb9\u5e02\u573a\u73b0\u6c47\u4ef7", "SHANDONG", "\u5c71\u4e1c", "province"),
        "\u897f\u5357\u5e02\u573a\u4ef7": ("southwest_gas92_market", "\u897f\u535792#\u6c7d\u6cb9\u5e02\u573a\u4ef7", "SOUTHWEST", "\u897f\u5357", "macro_region"),
        "\u534e\u4e2d\u5e02\u573a\u4ef7": ("central_china_gas92_market", "\u534e\u4e2d92#\u6c7d\u6cb9\u5e02\u573a\u4ef7", "CENTRAL_CHINA", "\u534e\u4e2d", "macro_region"),
        "\u534e\u5317\u5e02\u573a\u4ef7": ("north_china_gas92_market", "\u534e\u531792#\u6c7d\u6cb9\u5e02\u573a\u4ef7", "NORTH_CHINA", "\u534e\u5317", "macro_region"),
        "\u534e\u4e1c\u5e02\u573a\u4ef7": ("east_china_gas92_market", "\u534e\u4e1c92#\u6c7d\u6cb9\u5e02\u573a\u4ef7", "EAST_CHINA", "\u534e\u4e1c", "macro_region"),
        "\u4e1c\u5317\u5e02\u573a\u4ef7": ("northeast_gas92_market", "\u4e1c\u531792#\u6c7d\u6cb9\u5e02\u573a\u4ef7", "NORTHEAST", "\u4e1c\u5317", "macro_region"),
        "\u897f\u5317\u5e02\u573a\u4ef7": ("northwest_gas92_market", "\u897f\u531792#\u6c7d\u6cb9\u5e02\u573a\u4ef7", "NORTHWEST", "\u897f\u5317", "macro_region"),
        "\u534e\u5357\u5e02\u573a\u4ef7": ("south_china_gas92_market", "\u534e\u535792#\u6c7d\u6cb9\u5e02\u573a\u4ef7", "SOUTH_CHINA", "\u534e\u5357", "macro_region"),
    },
    DIESEL_SHEET: {
        "\u5c71\u4e1c\u5e02\u573a\u4ef7": ("sd_diesel0_market", "\u5c71\u4e1c0#\u67f4\u6cb9\u5e02\u573a\u73b0\u6c47\u4ef7", "SHANDONG", "\u5c71\u4e1c", "province"),
        "\u897f\u5357\u5e02\u573a\u4ef7": ("southwest_diesel0_market", "\u897f\u53570#\u67f4\u6cb9\u5e02\u573a\u4ef7", "SOUTHWEST", "\u897f\u5357", "macro_region"),
        "\u534e\u4e2d\u5e02\u573a\u4ef7": ("central_china_diesel0_market", "\u534e\u4e2d0#\u67f4\u6cb9\u5e02\u573a\u4ef7", "CENTRAL_CHINA", "\u534e\u4e2d", "macro_region"),
        "\u534e\u5317\u5e02\u573a\u4ef7": ("north_china_diesel0_market", "\u534e\u53170#\u67f4\u6cb9\u5e02\u573a\u4ef7", "NORTH_CHINA", "\u534e\u5317", "macro_region"),
        "\u534e\u4e1c\u5e02\u573a\u4ef7": ("east_china_diesel0_market", "\u534e\u4e1c0#\u67f4\u6cb9\u5e02\u573a\u4ef7", "EAST_CHINA", "\u534e\u4e1c", "macro_region"),
        "\u897f\u5317\u5e02\u573a\u4ef7": ("northwest_diesel0_market", "\u897f\u53170#\u67f4\u6cb9\u5e02\u573a\u4ef7", "NORTHWEST", "\u897f\u5317", "macro_region"),
        "\u4e1c\u5317\u5e02\u573a\u4ef7": ("northeast_diesel0_market", "\u4e1c\u53170#\u67f4\u6cb9\u5e02\u573a\u4ef7", "NORTHEAST", "\u4e1c\u5317", "macro_region"),
        "\u534e\u5357\u5e02\u573a\u4ef7": ("south_china_diesel0_market", "\u534e\u53570#\u67f4\u6cb9\u5e02\u573a\u4ef7", "SOUTH_CHINA", "\u534e\u5357", "macro_region"),
    },
}

PRODUCT_FAMILY_BY_SHEET = {GASOLINE_SHEET: "GASOLINE_92", DIESEL_SHEET: "DIESEL"}
INDICATOR_CODES = [item[0] for sheet in PRICE_COLUMNS.values() for item in sheet.values()]


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value).strip()).date()
    except Exception:
        return None


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def payload_hash(payload: dict[str, Any]) -> str:
    return sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def source_record_id(payload: dict[str, Any]) -> str:
    return f"{SOURCE_CODE}:price:{payload_hash(payload)[:24]}"


def extract_price_rows(path: Path, *, start_date: date | None, end_date: date | None) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    indicator_rows: dict[str, dict[str, Any]] = {}
    data_rows: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    for sheet_name, column_map in PRICE_COLUMNS.items():
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Missing sheet: {sheet_name}")
        ws = wb[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        if not headers or headers[0] != DATE_HEADER:
            raise RuntimeError(f"Invalid header in sheet: {sheet_name}")
        column_indexes = {header: idx + 1 for idx, header in enumerate(headers) if header in column_map}
        missing = sorted(set(column_map) - set(column_indexes))
        if missing:
            raise RuntimeError(f"Missing columns in {sheet_name}: {missing}")
        rows_before = len(data_rows)
        for header, (indicator_code, indicator_name, entity_code, entity_name, region_level) in column_map.items():
            indicator_rows[indicator_code] = {
                "indicator_code": indicator_code,
                "indicator_name": indicator_name,
                "category": "refined_oil",
                "sub_category": "ganglian_excel_price_612",
                "unit": "\u5143/\u5428",
                "freq": "daily",
                "value_type": "number",
                "fill_policy_default": "latest_available",
                "description": f"\u6a21\u578b\u9884\u6d4b\u57fa\u7840\u6570\u636e6.12.xlsx {sheet_name} {header}",
                "entity_type": "market_region",
                "entity_code": entity_code,
                "entity_name": entity_name,
                "region_level": region_level,
                "product_family": PRODUCT_FAMILY_BY_SHEET[sheet_name],
            }
        for row_number, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            observation_date = parse_date(row_values[0] if row_values else None)
            if observation_date is None:
                continue
            if start_date and observation_date < start_date:
                continue
            if end_date and observation_date > end_date:
                continue
            for header, col_idx in column_indexes.items():
                value_num = parse_number(row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None)
                if value_num is None:
                    continue
                indicator_code, _, entity_code, _, _ = column_map[header]
                raw_payload = {
                    "file_name": path.name,
                    "sheet_name": sheet_name,
                    "row_number": row_number,
                    "column_number": col_idx,
                    "observation_date": observation_date.isoformat(),
                    "indicator_code": indicator_code,
                    "header": header,
                    "value_num": value_num,
                    "unit": "\u5143/\u5428",
                }
                data_rows.append({
                    "indicator_code": indicator_code,
                    "entity_code": entity_code,
                    "observation_date": observation_date,
                    "value_num": value_num,
                    "unit": "\u5143/\u5428",
                    "freq": "daily",
                    "source_record_id": source_record_id(raw_payload),
                    "raw_payload": raw_payload,
                })
        summaries.append({
            "sheet_name": sheet_name,
            "indicator_count": len(column_map),
            "timeseries_rows": len(data_rows) - rows_before,
        })
    summary = {
        "source_code": SOURCE_CODE,
        "excel_path": str(path),
        "indicator_count": len(indicator_rows),
        "timeseries_row_count": len(data_rows),
        "date_min": min((r["observation_date"] for r in data_rows), default=None),
        "date_max": max((r["observation_date"] for r in data_rows), default=None),
        "sheets": summaries,
    }
    return list(indicator_rows.values()), data_rows, summary


def resolve_database_settings(args: argparse.Namespace) -> DatabaseSettings:
    if args.database_url:
        return DatabaseSettings(url=args.database_url, schema=args.schema or "oil_research", echo=False)
    settings = get_settings()
    if args.schema and args.schema != settings.database.schema:
        return DatabaseSettings(url=settings.database.url, schema=args.schema, echo=settings.database.echo)
    return settings.database


def replace_price_timeseries(repository: PostgresSnapshotRepository, indicator_rows: list[dict[str, Any]], data_rows: list[dict[str, Any]], summary: dict[str, Any]) -> dict[str, int]:
    if not repository.engine:
        return {"deleted": 0, "inserted": 0, "raw_inserted": 0}
    now = datetime.now(timezone.utc)
    indicator_codes = [row["indicator_code"] for row in indicator_rows]
    with repository.engine.begin() as connection:
        source_id = repository._ensure_source_ids(connection, {SOURCE_CODE})[SOURCE_CODE]
        repository._ensure_indicators(connection, indicator_rows)
        repository._ensure_entities(connection, indicator_rows)
        delete_params = {
            "source_id": source_id,
            "indicator_codes": indicator_codes,
            "date_min": summary.get("date_min"),
            "date_max": summary.get("date_max"),
        }
        deleted = connection.execute(
            text(f"""
                delete from {repository._fqtn('fact_market_timeseries')} f
                using {repository._fqtn('dim_indicator')} i
                where f.indicator_id = i.indicator_id
                  and f.source_id = :source_id
                  and i.indicator_code in :indicator_codes
                  and f.dt between :date_min and :date_max
            """).bindparams(bindparam("indicator_codes", expanding=True)),
            delete_params,
        ).rowcount or 0
        raw_payload = {"summary": summary, "indicator_codes": sorted(indicator_codes)}
        raw_result = connection.execute(
            text(f"""
                insert into {repository._fqtn('ods_raw_market')} (
                    source_id, source_record_id, topic, source_event_time,
                    payload, payload_hash, dt
                ) values (
                    :source_id, :source_record_id, :topic, :source_event_time,
                    cast(:payload as jsonb), :payload_hash, :dt
                ) on conflict (source_id, source_record_id, payload_hash) do nothing
            """),
            {
                "source_id": source_id,
                "source_record_id": f"{SOURCE_CODE}:price_replace:{summary['date_min']}:{summary['date_max']}",
                "topic": "ganglian_excel_price_replace_612",
                "source_event_time": now,
                "payload": json.dumps(raw_payload, ensure_ascii=False, default=str),
                "payload_hash": payload_hash(raw_payload),
                "dt": date.today(),
            },
        )
        rows = list(connection.execute(
            text(f"""
                select i.indicator_id, i.indicator_code, e.entity_id, e.entity_code
                from {repository._fqtn('dim_indicator')} i
                join {repository._fqtn('dim_entity')} e on 1 = 1
                where i.indicator_code in :indicator_codes
                  and e.entity_code in :entity_codes
            """).bindparams(bindparam("indicator_codes", expanding=True), bindparam("entity_codes", expanding=True)),
            {
                "indicator_codes": indicator_codes,
                "entity_codes": sorted({row["entity_code"] for row in data_rows}),
            },
        ).mappings())
        indicator_id_map = {str(row["indicator_code"]): int(row["indicator_id"]) for row in rows}
        entity_id_map = {str(row["entity_code"]): int(row["entity_id"]) for row in rows}
        insert_rows: list[dict[str, Any]] = []
        for row in data_rows:
            indicator_id = indicator_id_map.get(row["indicator_code"])
            entity_id = entity_id_map.get(row["entity_code"])
            if not indicator_id or not entity_id:
                continue
            observation_time = datetime.combine(row["observation_date"], time(0, 0))
            insert_rows.append({
                "indicator_id": indicator_id,
                "entity_id": entity_id,
                "observation_time": observation_time,
                "publish_time": now,
                "freq": row["freq"],
                "value_num": row["value_num"],
                "value_text": None,
                "unit": row["unit"],
                "currency": "CNY",
                "source_id": source_id,
                "source_record_id": row["source_record_id"],
                "is_final": True,
                "revision_no": 1,
                "quality_flag": "ok",
                "effective_from": now,
                "effective_to": None,
                "dt": row["observation_date"],
            })
        statement = text(f"""
            insert into {repository._fqtn('fact_market_timeseries')} (
                indicator_id, entity_id, observation_time, publish_time, freq, value_num,
                value_text, unit, currency, source_id, source_record_id, is_final,
                revision_no, quality_flag, effective_from, effective_to, dt
            ) values (
                :indicator_id, :entity_id, :observation_time, :publish_time, :freq, :value_num,
                :value_text, :unit, :currency, :source_id, :source_record_id, :is_final,
                :revision_no, :quality_flag, :effective_from, :effective_to, :dt
            )
        """)
        inserted = 0
        for start in range(0, len(insert_rows), 5000):
            result = connection.execute(statement, insert_rows[start:start+5000])
            inserted += int(result.rowcount or 0)
    return {"deleted": int(deleted), "inserted": inserted, "raw_inserted": int(raw_result.rowcount or 0)}


def main() -> None:
    parser = argparse.ArgumentParser(description="Replace regional gasoline/diesel market prices from 6.12 workbook.")
    parser.add_argument("--excel", default=str(ROOT / DEFAULT_EXCEL_NAME))
    parser.add_argument("--database-url", default="")
    parser.add_argument("--schema", default="")
    parser.add_argument("--start-date", default="")
    parser.add_argument("--end-date", default="")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--summary-output", default="artifacts/replace_region_market_prices_612_summary.json")
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    start_date = parse_date(args.start_date) if args.start_date else None
    end_date = parse_date(args.end_date) if args.end_date else None
    indicator_rows, data_rows, summary = extract_price_rows(excel_path, start_date=start_date, end_date=end_date)
    summary["dry_run"] = bool(args.dry_run)
    summary["indicator_codes"] = sorted({row["indicator_code"] for row in indicator_rows})
    if args.dry_run:
        summary["db_result"] = {"deleted": 0, "inserted": 0, "raw_inserted": 0}
    else:
        settings = resolve_database_settings(args)
        repository = PostgresSnapshotRepository(settings)
        repository.ensure_schema()
        summary["db_result"] = replace_price_timeseries(repository, indicator_rows, data_rows, summary)
    output_path = (ROOT / args.summary_output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
