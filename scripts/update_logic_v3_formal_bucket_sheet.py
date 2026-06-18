from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SUMMARY_JSON = ROOT / "artifacts" / "hard_data_score_bucket_calibration_20240101_20260612.json"
WORKBOOK_PATH = ROOT / "outputs" / "逻辑说明V3.xlsx"
SHEET_NAME = "正式状态桶校准"

BUCKET_LABELS = ["强空", "偏空", "弱空", "震荡", "弱多", "偏多", "强多"]


def fmt_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        numeric = float(value)
    except Exception:
        return str(value)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.2f}".rstrip("0").rstrip(".")


def fmt_rate(value: Any) -> str:
    if value is None:
        return ""
    return f"{float(value) * 100:.1f}%"


def threshold_ranges(thresholds: list[float]) -> list[str]:
    bounds: list[float | None] = [None, *[float(v) for v in thresholds], None]
    ranges: list[str] = []
    for idx in range(7):
        lower = bounds[idx]
        upper = bounds[idx + 1]
        lower_text = "-∞" if lower is None else fmt_value(lower)
        upper_text = "+∞" if upper is None else fmt_value(upper)
        ranges.append(f"{BUCKET_LABELS[idx]}:{lower_text}~{upper_text}")
    return ranges


def bucket_explain(row: dict[str, Any]) -> str:
    bucket = str(row.get("bucket") or "")
    n = int(row.get("sample_size") or 0)
    p50 = row.get("p50")
    up_rate = row.get("up_rate")
    if n <= 0:
        return "该桶暂无完整历史样本，不单独解释，预测时需合并相邻桶或提示谨慎。"
    base = (
        f"历史落入{bucket}的完整样本{n}条，后续实际价格变动中位数{fmt_value(p50)}元/吨，"
        f"上涨概率{fmt_rate(up_rate)}。"
    )
    if n < 12:
        return base + "样本低于12条，只作方向参考，不单独硬套。"
    return base + "该统计用于说明该分数段在历史上对应的价格弹性。"


def best_candidate(item: dict[str, Any], model: str) -> dict[str, Any]:
    key = "agent_threshold_candidates" if model == "智能体" else "business_threshold_candidates"
    candidates = item.get(key) or []
    return candidates[0] if candidates else {}


def write_sheet(summary: dict[str, Any]) -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = WORKBOOK_PATH.with_name(f"{WORKBOOK_PATH.stem}.backup_{timestamp}{WORKBOOK_PATH.suffix}")
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    if SHEET_NAME in wb.sheetnames:
        old = wb[SHEET_NAME]
        index = wb.sheetnames.index(SHEET_NAME)
        wb.remove(old)
        ws = wb.create_sheet(SHEET_NAME, index)
    else:
        ws = wb.create_sheet(SHEET_NAME)

    ws.append(["正式状态桶校准（硬数据历史回放）"] + [None] * 12)
    ws.append(["项目", "口径/逻辑"] + [None] * 11)
    ws.append(["校准范围", f"{summary['start_date']} 至 {summary['as_of']}；覆盖山东及华东/华北/华南/华中/西北/西南/东北汽油和柴油。"] + [None] * 11)
    ws.append(["山东周末口径", "山东汽油使用库内日度市场价；周末行如果只是缺少Brent交易价格，不作为剔除条件，仍进入历史打分。"] + [None] * 11)
    ws.append(["其他区域价格口径", "其他区域按本区域真实市场价日期序列计算当前价和目标价；周末没有本区域市场价的行不进入打分。"] + [None] * 11)
    ws.append(["Brent口径", "历史打分用已发生真实值，不用历史Brent预测点位：D1=1日真实变动，W1=5日真实变动，M1=20日真实变动；山东汽油仅Brent缺失时忽略该缺失。"] + [None] * 11)
    ws.append(["产销率口径", "使用《产销率终版.xlsx》；汽油/柴油分列读取，1.09按109%处理。"] + [None] * 11)
    ws.append(["M1库存口径", "取当月已发布的最后一次统计库存分位；汽油用山东独立炼厂和主营销售公司库存分位，柴油用柴油对应库存分位。"] + [None] * 11)
    ws.append(["软信号", "资讯、事件、政策情绪本次暂不纳入状态桶历史校准，避免因历史结构化数据不全而产生缺失兜底。"] + [None] * 11)
    ws.append(["缺失原则", "除山东汽油仅Brent缺失外，必需字段缺失的历史样本直接剔除。"] + [None] * 11)
    ws.append(["更新说明", "本页已按新口径重跑并替换旧版状态桶统计；下方“正式阈值汇总”和“状态桶明细”均来自新一轮硬数据历史回放。"] + [None] * 11)
    ws.append([])

    ws.append(["一、正式阈值汇总（按品种区域校准）"] + [None] * 12)
    ws.append([
        "品种区域",
        "品种",
        "周期",
        "模型",
        "完整样本数",
        "是否可用于校准",
        "正式校准阈值",
        "状态桶区间解释",
        "MAE",
        "RMSE",
        "惩罚",
        "目标函数",
        "逻辑说明",
    ])
    for target in summary.get("targets") or []:
        for horizon, item in (target.get("horizons") or {}).items():
            for model in ["智能体", "业务"]:
                best = best_candidate(item, model)
                thresholds = best.get("thresholds") or []
                ws.append([
                    target.get("target_label"),
                    target.get("product_code"),
                    horizon,
                    model,
                    item.get("complete_samples"),
                    "是" if item.get("threshold_update_allowed") else "否",
                    json.dumps(thresholds, ensure_ascii=False) if thresholds else "",
                    "；".join(threshold_ranges(thresholds)) if thresholds else "",
                    best.get("mae"),
                    best.get("rmse"),
                    best.get("penalty"),
                    best.get("objective"),
                    "该阈值为本品种区域在候选阈值中的最优历史回放结果；优先考虑误差和方向性惩罚。",
                ])

    ws.append([])
    ws.append(["二、状态桶明细（用历史真实价格变动解释每个打分档位）"] + [None] * 12)
    ws.append([
        "品种区域",
        "品种",
        "周期",
        "模型",
        "状态桶",
        "分数区间",
        "样本数",
        "历史P25",
        "历史P50",
        "历史P75",
        "上涨概率",
        "样本是否充足",
        "业务解释",
    ])
    for target in summary.get("targets") or []:
        for horizon, item in (target.get("horizons") or {}).items():
            for model in ["智能体", "业务"]:
                best = best_candidate(item, model)
                for row in best.get("bucket_stats") or []:
                    ws.append([
                        target.get("target_label"),
                        target.get("product_code"),
                        horizon,
                        model,
                        row.get("bucket"),
                        row.get("score_range"),
                        row.get("sample_size"),
                        row.get("p25"),
                        row.get("p50"),
                        row.get("p75"),
                        row.get("up_rate"),
                        "是" if row.get("sample_sufficient") else "否",
                        bucket_explain(row),
                    ])

    style_sheet(ws)
    wb.save(WORKBOOK_PATH)
    print(json.dumps({"updated": str(WORKBOOK_PATH), "backup": str(backup_path), "rows": ws.max_row, "cols": ws.max_column}, ensure_ascii=False))


def style_sheet(ws: Any) -> None:
    thin = Side(style="thin", color="D9E2EC")
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="D9EAD3")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    note_fill = PatternFill("solid", fgColor="F8FBFD")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx in [2, 14]:
        for cell in ws[row_idx]:
            cell.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF" if row_idx == 14 else "000000")
            cell.fill = header_fill if row_idx == 14 else note_fill

    for row_idx in range(2, 12):
        ws.cell(row_idx, 1).font = Font(name="微软雅黑", bold=True, size=10)
        for col_idx in range(1, 14):
            ws.cell(row_idx, col_idx).fill = note_fill

    for row_idx in range(1, ws.max_row + 1):
        first = ws.cell(row_idx, 1).value
        if isinstance(first, str) and first.startswith(("一、", "二、")):
            for col_idx in range(1, 14):
                cell = ws.cell(row_idx, col_idx)
                cell.font = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
                cell.fill = section_fill
        if row_idx > 1 and first in {"品种区域"}:
            for col_idx in range(1, 14):
                cell = ws.cell(row_idx, col_idx)
                cell.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
                cell.fill = header_fill

    widths = [16, 14, 8, 10, 12, 14, 28, 62, 10, 10, 10, 12, 68]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for idx in range(1, ws.max_row + 1):
        ws.row_dimensions[idx].height = 24
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A15"
    ws.auto_filter.ref = f"A14:M{ws.max_row}"


def main() -> None:
    summary = json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))
    write_sheet(summary)


if __name__ == "__main__":
    main()
