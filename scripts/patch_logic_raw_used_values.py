from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

path = Path('outputs') / '逻辑说明.xlsx'
wb = load_workbook(path)

NOTE = '列含义修正：打分参照规则=具体如何算分/阈值/公式；原始规则=本次用了哪些数据、各打多少分、合计多少；口径/证据=数据来源、日期、取数口径和必要解释。'
SOURCE_8013 = '截图中的当前山东92#=8013来自旧缓存预测 artifacts/prediction_runs/sdgas92-c3cce8c86ac0.json：as_of_date=2026-06-07，raw_context.current_price=8013，current_price_column=sd_gas92_market，market_data_reason=eta_unavailable;local_market_overrides=9591;cash_price_overlay=4644;brent_daily_report_stale；其原始Excel对应《模型预测基础数据.xlsx》价格类C15，指标“汽油：国六：92#：市场现汇价：山东（日）”，日期2026-06-05。当前重新构建上下文的2026-06-12山东92#为8042，不应继续用8013作为当前价。'

BUSINESS_OVERRIDE = {
    ('D1','成本侧'): '本次用数：Brent日变化=-2.06美元/桶，得-30分；汽油裂解价差分位=66.25%，得-5分；模块合计=-35分，组内封顶±40后仍为-35分。',
    ('D1','供给侧'): '本次用数：山东地炼开工率=51.09%，命中45~55区间，得0分；炼厂负荷新闻修正=0，得0分；模块合计=0分。',
    ('D1','需求侧'): '本次用数：山东独立炼厂汽油产销率=108%，命中100~110区间，得+15分；模块合计=+15分。',
    ('D1','情绪侧'): '本次用数：成交/贸易商情绪=bullish_active，得+15分；模块合计=+15分。',
    ('D3','成本侧'): '本次用数：Brent三日变化缺失，按0分；汽油裂解价差分位=66.25%，得-5分；模块合计=-5分。',
    ('D3','供给侧'): '本次用数：山东地炼开工率=51.09%，得0分；炼厂负荷新闻修正=0，得0分；模块合计=0分。',
    ('D3','需求侧'): '本次用数：三日/短周期产销率按当前可用产销率口径参与，命中对应分桶后计入模块；缺失项按0分，不用其他周期冒充。',
    ('D3','情绪侧'): '本次用数：成交/贸易商情绪按D3可用标签计分；无连续明确方向时按0分，存在活跃补货/抛货信号时按枚举分计入。',
    ('W1','成本侧'): '本次用数：Brent W1变化=88.79-90.38=-1.59美元/桶，按W1成本分桶计分；汽油裂解价差分位=66.25%，按高分位偏利空计分；两项合计后按组内上限截断。',
    ('M1','成本侧'): '本次用数：Brent M1变化=89.53-90.38=-0.85美元/桶，按M1成本分桶计分；汽油裂解价差分位=66.25%，按高分位偏利空计分；两项合计后按组内上限截断。',
}

AGENT_OVERRIDE = {
    '原油成本智能体': '本次用数：Brent预测/结算变化=-2.06美元/桶，对应成本分=-42.0；MTBE三日变化=18.75、石脑油三日变化=8.98作为调油修正；原始分=-37.3783，满分65，归一化=-0.5751，权重0.22，加权贡献=-0.1265。',
    '市场结构智能体': '本次用数：华东-山东=48、华北-山东=94、华南-山东=45、华中-山东=101、西北-山东=142元/吨；按区域均值、全国相对山东和动量折算；原始分=64.7589，满分100，归一化=0.6476，权重0.16，加权贡献=0.1036。',
    '供给库存智能体': '本次用数：开工率分位=40.9%，开工变化=-0.22，山东炼油利润=-315.6元/吨，库存分位已计入；原始分=37.7875，满分62，归一化=0.6095，权重0.20，加权贡献=0.1219。',
    '需求季节智能体': '本次用数：D1产销率=108.0%，月度产销率变化缺失按0，季节性修正=14.0，节假日修正=6.0；原始分=56，满分100，归一化=0.56，权重0.18，加权贡献=0.1008。',
    '成品油资讯智能体': '本次用数：资讯标签未启用/无明确方向，label=neutral_flat；原始分=0，满分100，归一化=0，权重0.12，加权贡献=0。',
    '政策周期智能体': '本次用数：卓创调价预期=-320元/吨，命中<-100分桶，距离调价窗口约4个工作日，理论有效区间=-82~89.5；按窗口系数折算后计入政策周期得分。',
    '山东现货跳变识别': '本次用数：未识别到低价资源扫空、封单惜售、抢货/抛货等足够硬信号；该项只做D1点位修正，不参与综合加权，本次按0处理。',
    '事件风险智能体': '本次用数：事件文本“美伊停战宣布霍尔木兹海峡解封，航运恢复”，识别为event_type=supply_relief、direction=down、severity=high；该项为门控项，权重0，不直接加权，但要求价格方向不能被裁判层反向抵消。',
    '裁判层智能体': '本次用数：检查硬数据、软信号、预测方向和事件风险的一致性；若发现方向冲突，则调整置信度、区间或提示人工复核；裁判层本身不直接加权。',
}


def header_map(ws):
    return {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}

for sheet_name in ['BusinessScoring', 'AgentScoring']:
    ws = wb[sheet_name]
    base = ws.cell(2, 2).value or ''
    first_line = base.split('\n')[0] if base else ''
    ws.cell(2, 2).value = (first_line + '\n' + NOTE).strip()
    ws.cell(2, 2).alignment = Alignment(wrap_text=True, vertical='top')

ws = wb['BusinessScoring']
headers = header_map(ws)
raw_col = headers['原始规则']
for row in range(4, ws.max_row + 1):
    key = (ws.cell(row, 1).value, ws.cell(row, 2).value)
    if key in BUSINESS_OVERRIDE:
        ws.cell(row, raw_col).value = BUSINESS_OVERRIDE[key]
    else:
        data_item = ws.cell(row, 3).value or ''
        value = ws.cell(row, 4).value or ''
        score = ws.cell(row, 7).value or ''
        hit = ws.cell(row, 8).value or ''
        ws.cell(row, raw_col).value = f'本次用数：{data_item}；取值={value}；命中/结论={hit}；模块合计={score}分。'

ws = wb['AgentScoring']
headers = header_map(ws)
raw_col = headers['原始规则']
for row in range(4, ws.max_row + 1):
    agent = ws.cell(row, 2).value
    if agent in AGENT_OVERRIDE:
        ws.cell(row, raw_col).value = AGENT_OVERRIDE[agent]
    else:
        raw_score = ws.cell(row, 3).value
        max_score = ws.cell(row, 4).value
        normalized = ws.cell(row, 5).value
        weight = ws.cell(row, 6).value
        contribution = ws.cell(row, 7).value
        evidence = ws.cell(row, 11).value or ''
        ws.cell(row, raw_col).value = f'本次用数：{evidence}；原始分={raw_score}，满分={max_score}，归一化={normalized}，权重={weight}，加权贡献={contribution}。'

for sheet_name in ['BusinessScoring', 'AgentScoring']:
    ws = wb[sheet_name]
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(3, col)
        if cell.value == '打分参照规则':
            cell.fill = PatternFill('solid', fgColor='FCE4D6')
            cell.font = Font(bold=True)
        elif cell.value == '原始规则':
            cell.fill = PatternFill('solid', fgColor='D9EAD3')
            cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['I'].width = 58
    ws.column_dimensions['J'].width = 62
    ws.column_dimensions['K'].width = 58

if '表格说明' in wb.sheetnames:
    ws = wb['表格说明']
    rows = {ws.cell(r, 1).value: r for r in range(1, ws.max_row + 1)}
    for key, value in [('规则列含义', NOTE), ('8013来源说明', SOURCE_8013)]:
        row = rows.get(key)
        if not row:
            row = ws.max_row + 1
            ws.cell(row, 1).value = key
        ws.cell(row, 2).value = value
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(path)
print(f'updated {path}')
