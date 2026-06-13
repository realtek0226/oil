from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import sys
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.container import get_dataset_service, get_predictor


OUTPUT = Path(r"E:\中鲁燃能\业务打分模型使用数据明细_20260608.xlsx")
RUN_DATE = date(2026, 6, 8)
HORIZONS = ["D1", "D3", "W1", "M1"]

FEATURE_LABELS = {
    "brent_change_usd_d1": "Brent涨跌",
    "brent_change_usd_d3": "Brent涨跌",
    "brent_change_usd_w1": "Brent涨跌",
    "brent_change_usd_m1": "Brent涨跌",
    "shandong_cdu_utilization_percentile_weekly": "山东地炼产能利用率分位",
    "shandong_cdu_utilization_percentile_monthly": "山东地炼产能利用率月度分位",
    "shandong_refinery_load_news_adjustment_d1": "山东炼厂负荷新闻修正",
    "sales_production_ratio_d1": "山东地炼汽油产销率D1",
    "sales_production_ratio_d3_avg": "汽油产销率3日均值",
    "sales_production_ratio_w1_avg": "汽油产销率5日/短周期均值",
    "sales_production_ratio_monthly_avg": "汽油产销率月度均值",
    "trader_sentiment_label_d1": "贸易商/成交情绪标签",
    "trader_sentiment_label_d3": "贸易商/成交情绪标签",
    "market_sentiment_monthly": "月度市场情绪标签",
    "monthly_seasonality_phase": "月度淡旺季阶段",
    "restocking_rhythm_monthly": "补库节奏",
    "holiday_demand_delta_monthly": "节假日需求变化",
    "next_month_maintenance_plan": "次月检修计划",
}

GROUP_LABELS = {
    "cost": "成本侧",
    "supply": "供给侧",
    "demand": "需求侧",
    "sentiment": "情绪侧",
    "seasonality": "季节性",
}

SOURCE_HINTS = {
    "brent_change": "Brent预测日报 + Brent特征结算价",
    "shandong_cdu": "手工采集模板/隆众经营数据",
    "sales_production": "手工采集模板/山东地炼汽油产销率",
    "trader_sentiment": "成品油资讯/山东日评标签",
    "market_sentiment": "成品油资讯月度标签",
    "monthly": "系统日历规则",
    "holiday": "系统节假日统计",
    "maintenance": "隆众检修计划归档",
}


def clean(value: Any) -> Any:
    if value is None:
        return ""
    try:
        if isinstance(value, float) and value != value:
            return ""
    except Exception:
        pass
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def source_hint(feature: str) -> str:
    for key, hint in SOURCE_HINTS.items():
        if key in feature:
            return hint
    return "系统特征/规则计算"


def unit_for(feature: str) -> str:
    if "brent" in feature:
        return "美元/桶"
    if "percentile" in feature or "ratio" in feature:
        return "%"
    if "adjustment" in feature:
        return "分/元吨修正"
    return ""


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    orange_fill = PatternFill("solid", fgColor="F97316")
    light_orange = PatternFill("solid", fgColor="FFEDD5")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, sheet.max_column + 1):
            max_len = 0
            for col_cells in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=sheet.max_row):
                for cell in col_cells:
                    max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)
        for row_idx in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = 22

    ws = wb["说明与摘要"]
    ws["A1"].fill = orange_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A2"].fill = light_orange


def main() -> None:
    dataset_service = get_dataset_service()
    as_of = dataset_service.resolve_default_prediction_as_of(RUN_DATE)
    context = dataset_service.build_context(as_of)
    predictor = get_predictor()
    predictions = predictor.run_multi_horizon_predictions_from_context(
        context,
        as_of,
        horizons=HORIZONS,
        use_llm_explainer=False,
    )
    row = context.current_row

    wb = Workbook()
    ws = wb.active
    ws.title = "说明与摘要"
    ws["A1"] = "业务打分模型使用数据明细"
    ws["A2"] = (
        "本表列出当前系统中“山东成品油市场价预测打分模型”实际读取并计分的数据。"
        "裂解价差分位、库存分位已按要求不参与业务打分。"
    )
    ws.merge_cells("A2:H2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    summary_rows = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("预测运行日", RUN_DATE.isoformat()),
        ("业务输入日", as_of.isoformat()),
        ("D1目标日", str(predictions[0].target_date)),
        ("价格锚点日", context.metadata.get("price_anchor_date")),
        ("山东92#锚点价", row.get("sd_gas92_market")),
        ("Brent口径", "Brent预测点位 - Brent特征结算价"),
        ("未参与项", "裂解价差分位、库存分位"),
    ]
    for idx, (key, value) in enumerate(summary_rows, start=4):
        ws.cell(idx, 1, key)
        ws.cell(idx, 2, clean(value))

    ws2 = wb.create_sheet("各周期预测与总分")
    ws2.append(["周期", "输入日", "目标日", "价格锚点", "业务模型总分", "业务预测涨跌", "业务预测点位", "业务预测区间", "方向", "覆盖率", "缺失字段"])
    for prediction in predictions:
        raw_context = prediction.raw_context or {}
        business = raw_context.get("business_scorecard_prediction") or {}
        scorecard = raw_context.get("business_scorecard") or {}
        data_quality = scorecard.get("data_quality") or {}
        ws2.append(
            [
                prediction.horizon,
                str(prediction.as_of_date),
                str(prediction.target_date),
                business.get("current_price"),
                business.get("score"),
                business.get("predicted_delta"),
                business.get("point_value"),
                f"{business.get('range_lower')} ~ {business.get('range_upper')}",
                business.get("direction_label"),
                data_quality.get("coverage_ratio"),
                ", ".join(data_quality.get("missing_fields") or []),
            ]
        )

    ws3 = wb.create_sheet("打分字段明细")
    ws3.append(["周期", "分组", "字段编码", "字段名称", "实际取值", "单位", "得分", "状态", "数据来源/口径", "备注"])
    for prediction in predictions:
        scorecard = (prediction.raw_context or {}).get("business_scorecard") or {}
        for group in scorecard.get("groups") or []:
            group_code = group.get("group_code") or ""
            group_label = GROUP_LABELS.get(group_code, group.get("display_name") or group_code)
            for feature in group.get("features") or []:
                name = feature.get("feature_name") or ""
                status = feature.get("status") or ""
                ws3.append(
                    [
                        prediction.horizon,
                        group_label,
                        name,
                        FEATURE_LABELS.get(name, name),
                        clean(feature.get("value")),
                        unit_for(name),
                        clean(feature.get("score")),
                        "缺失未计分" if status == "missing" else "已计分",
                        source_hint(name),
                        "缺失字段按0分处理" if status == "missing" else "",
                    ]
                )

    ws4 = wb.create_sheet("当前原始输入")
    ws4.append(["类别", "数据项", "取值", "单位", "日期/截止", "来源/口径"])
    brent_basis = (predictions[0].raw_context or {}).get("brent_forecast_basis") or {}
    trade_sentiment = (predictions[0].raw_context or {}).get("trade_sentiment") or {}
    raw_rows = [
        ("价格锚点", "山东92#市场价", row.get("sd_gas92_market"), "元/吨", context.metadata.get("price_anchor_date"), "手工模板/ETA覆盖"),
        ("价格锚点", "全国92#市场价", row.get("cn_gas92_market"), "元/吨", context.metadata.get("price_anchor_date"), "手工模板"),
        ("成本", "Brent特征结算价", row.get("brent_active_settlement"), "美元/桶", context.metadata.get("price_anchor_date"), "特征序列"),
        ("成本", "Brent预测点位D1", brent_basis.get("forecast_point_usd"), "美元/桶", brent_basis.get("report_date"), "Brent日报"),
        ("成本", "Brent涨跌D1", brent_basis.get("scorecard_change_usd"), "美元/桶", as_of.isoformat(), "预测点位-特征结算价"),
        ("需求", "汽油产销率D1", row.get("sales_production_ratio_d1"), "%", as_of.isoformat(), "手工模板"),
        ("需求", "汽油产销率3日均值", row.get("sales_production_ratio_d3_avg"), "%", as_of.isoformat(), "手工模板"),
        ("需求", "汽油产销率5日均值", row.get("sales_production_ratio_w1_avg"), "%", as_of.isoformat(), "手工模板"),
        ("供给", "山东地炼产能利用率", row.get("sd_crude_run_weekly"), "%", as_of.isoformat(), "手工模板"),
        ("供给", "产能利用率周环比", row.get("shandong_cdu_utilization_wow_pct"), "百分点", as_of.isoformat(), "手工模板"),
        ("供给", "山东炼油利润", row.get("sd_refining_profit"), "元/吨", as_of.isoformat(), "手工模板"),
        ("情绪", "交易/成交标签", trade_sentiment.get("deal_activity"), "", as_of.isoformat(), "成品油资讯标签"),
        ("政策", "调价预测金额", row.get("price_adjustment_expected_yuan"), "元/吨", as_of.isoformat(), "缺失"),
        ("政策", "距调价窗口", row.get("days_to_next_window"), "工作日", as_of.isoformat(), "政策周期计算"),
    ]
    for raw_row in raw_rows:
        ws4.append([clean(value) for value in raw_row])

    ws5 = wb.create_sheet("不参与打分项")
    ws5.append(["数据项", "当前值", "处理方式", "说明"])
    ws5.append(["汽油裂解价差分位", clean(row.get("gasoline_crack_percentile")), "不参与业务打分/综合打分", "按要求先不加；裂解价差原值可保留展示"])
    ws5.append(["库存分位", clean(row.get("shandong_product_inventory_percentile_weekly")), "不参与业务打分/综合打分", "按要求先不加；库存原值可保留展示"])
    ws5.append(["库存合计", clean(row.get("shandong_product_inventory_total_formal")), "仅展示/备查", "贸易商+主营+独立炼厂库存合计"])
    ws5.append(["汽油裂解价差原值", clean(row.get("sd_gas_crack")), "仅展示/备查", "不转成分位得分"])

    style_workbook(wb)
    wb.save(OUTPUT)

    loaded = load_workbook(OUTPUT, data_only=False)
    print(f"saved={OUTPUT}")
    print(f"sheets={loaded.sheetnames}")
    for sheet_name in ["各周期预测与总分", "打分字段明细", "当前原始输入", "不参与打分项"]:
        sheet = loaded[sheet_name]
        print(f"{sheet_name}: rows={sheet.max_row}, cols={sheet.max_column}")


if __name__ == "__main__":
    main()
