from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

path = Path('outputs') / '逻辑说明.xlsx'
wb = load_workbook(path)

BUSINESS_RATIONALE = {
    ('D1','成本侧'): '业务含义：Brent下跌会降低成品油成本支撑，汽油裂解分位偏高说明利润已较充分、继续上行动力减弱，因此两项都按利空处理；D1强调隔夜原油对次日山东汽油价格的直接传导。',
    ('D1','供给侧'): '业务含义：山东地炼开工率越高，短期供应越宽松，对价格偏利空；开工率越低则供应收缩，对价格偏利多。新闻修正只用于补充突发检修、复产、负荷变化等当日信息。',
    ('D1','需求侧'): '业务含义：产销率反映当日炼厂出货强弱，产销率越高说明需求承接越好、库存压力越小，对价格偏利多；产销率低则说明出货偏弱，对价格偏利空。',
    ('D1','情绪侧'): '业务含义：贸易商主动补货、成交活跃通常领先现货报价上调；观望则不改变方向；抛货、让利或看跌心态会压制短期价格。',
    ('D3','成本侧'): '业务含义：D3看三日原油变化对短周期调价的持续传导；如果日报没有对应D3预测点位，不能用D1替代，避免把不同时间维度混在一起。裂解分位用于判断利润端是否还有继续扩张空间。',
    ('D3','供给侧'): '业务含义：三日维度供给更关注检修、复产和负荷变化的连续影响；开工高代表供应压力，开工低代表供应收缩。',
    ('D3','需求侧'): '业务含义：三日平均产销率比单日更能过滤偶发成交，连续高产销率说明采购节奏有支撑，连续低产销率说明终端承接不足。',
    ('D3','情绪侧'): '业务含义：连续几日成交和心态变化会影响贸易商库存策略，活跃补货偏利多，观望中性，抛货让利偏利空。',
    ('W1','成本侧'): '业务含义：W1看一周原油预期对山东汽油价格的趋势牵引；周度Brent明显下行会压低成本锚，裂解分位偏高说明利润修复已较充分。',
    ('W1','供给侧'): '业务含义：周度供给看开工率、库存和检修复产的综合压力；供应偏紧支撑价格，供应宽松压制价格。',
    ('W1','需求侧'): '业务含义：七日平均产销率代表一周内采购持续性，持续高于平衡线说明需求较好，持续偏低说明补货不足。',
    ('W1','库存侧'): '业务含义：库存分位越低，炼厂和社会库存压力越小，价格更容易获得支撑；库存分位越高，去库压力越大，价格更容易承压。',
    ('W1','政策与消息侧'): '业务含义：调价预期、宏观政策和市场消息会影响贸易商提前补货或延后采购，偏利多消息抬升预期，偏利空消息降低采购积极性。',
    ('M1','成本侧'): '业务含义：M1关注月度原油均值和趋势对成品油价格中枢的影响；如果月度Brent预测下行，价格中枢应相应下移。',
    ('M1','供给侧'): '业务含义：月度供给主要看检修计划、开工中枢和复产节奏；检修增加或开工下降会收紧供应，复产增加或开工上升会扩大供应。',
    ('M1','需求侧'): '业务含义：月度需求看备货节奏和季节性，山东用前两个月产销率大于90%的天数比较，其他区域用月均销量/出货变化比较；备货增强偏利多，减弱偏利空。',
    ('M1','库存侧'): '业务含义：月度库存决定价格弹性，低库存下价格更容易被成本或需求拉动，高库存下价格上涨阻力更大。',
    ('M1','政策与消息侧'): '业务含义：月度政策、调价窗口和重大事件影响经营预期，利多提高补库意愿，利空降低采购节奏。',
}

AGENT_RATIONALE = {
    '原油成本智能体': '业务含义：原油是成品油定价的成本锚。Brent预测点位低于结算价，说明成本端边际走弱，应压低价格方向；反之则抬升价格方向。调油品价差只作为成本传导的补充修正。',
    '市场结构智能体': '业务含义：区域价差反映山东货源外流能力和区域套利空间。区域价格高于山东越多，山东资源外流越顺畅、对山东价格越有支撑；价差收窄或倒挂则支撑减弱。',
    '供给库存智能体': '业务含义：开工率、库存和炼油利润共同描述供应压力。低库存、低开工或利润压缩会支撑价格；高库存、高开工或利润丰厚会压制价格。所有分位从2024-01-01以来历史样本计算。',
    '需求季节智能体': '业务含义：产销率、备货节奏、季节性和节假日共同反映终端承接能力。需求越强，炼厂出货越顺，价格越有支撑；需求越弱，价格越容易承压。',
    '政策情绪智能体': '业务含义：政策、调价预期和市场情绪会改变采购节奏。利多政策或上调预期会促使提前补货；利空政策或下调预期会让采购延后。',
    '成品油资讯智能体': '业务含义：日评和资讯用于补充硬数据尚未体现的成交、报价和贸易商心态变化；无明确方向时不强行给分。',
    '裁判层智能体': '业务含义：裁判层不直接加权，而是检查硬数据、软信号、预测方向和事件风险是否冲突；若冲突则降低置信度、收窄或下修区间，并提示人工复核。',
    '事件风险智能体': '业务含义：黑天鹅、地缘和突发事件不能按普通加权项处理，而是风险门控。供应中断类事件抬升油价风险；停战、海峡解封、航运恢复属于供应风险缓和，必须作为利空识别并要求人工复核。',
}

EXPLAIN_NOTE = '列含义修正：打分参照规则=具体如何算分/阈值/公式；原始规则=为什么这样打分的业务逻辑和方向解释；口径/证据=本次实际用数来源和计算过程。'

def header_map(ws):
    return {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}

for sheet_name in ['BusinessScoring', 'AgentScoring']:
    ws = wb[sheet_name]
    base = ws.cell(2, 2).value or ''
    if EXPLAIN_NOTE not in base:
        ws.cell(2, 2).value = (base + '\n' + EXPLAIN_NOTE).strip()
    ws.cell(2, 2).alignment = Alignment(wrap_text=True, vertical='top')

ws = wb['BusinessScoring']
headers = header_map(ws)
raw_col = headers['原始规则']
ref_col = headers['打分参照规则']
for row in range(4, ws.max_row + 1):
    key = (ws.cell(row, 1).value, ws.cell(row, 2).value)
    current_ref = ws.cell(row, ref_col).value or ''
    current_raw = ws.cell(row, raw_col).value or ''
    if key in BUSINESS_RATIONALE:
        ws.cell(row, raw_col).value = BUSINESS_RATIONALE[key]
    elif current_raw.startswith('原始规则：'):
        ws.cell(row, raw_col).value = '业务含义：该模块按当前数据项判断对应周期的供需、成本或情绪方向；正分代表对价格有支撑，负分代表对价格有压制。'
    if '原始规则：' in current_ref:
        ws.cell(row, ref_col).value = current_ref.replace('原始规则：', '计算规则：')

ws = wb['AgentScoring']
headers = header_map(ws)
raw_col = headers['原始规则']
ref_col = headers['打分参照规则']
for row in range(4, ws.max_row + 1):
    agent = ws.cell(row, 2).value
    current_ref = ws.cell(row, ref_col).value or ''
    if agent in AGENT_RATIONALE:
        ws.cell(row, raw_col).value = AGENT_RATIONALE[agent]
    if '原始规则：' in current_ref:
        ws.cell(row, ref_col).value = current_ref.replace('原始规则：', '计算规则：')

for sheet_name in ['BusinessScoring', 'AgentScoring']:
    ws = wb[sheet_name]
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(3, col)
        if cell.value == '打分参照规则':
            cell.fill = PatternFill('solid', fgColor='FCE4D6')
            cell.font = Font(bold=True)
        elif cell.value == '原始规则':
            cell.fill = PatternFill('solid', fgColor='E2F0D9')
            cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['I'].width = 58
    ws.column_dimensions['J'].width = 54
    ws.column_dimensions['K'].width = 54

if '表格说明' in wb.sheetnames:
    ws = wb['表格说明']
    found = False
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == '规则列含义':
            ws.cell(row, 2).value = EXPLAIN_NOTE
            found = True
    if not found:
        next_row = ws.max_row + 1
        ws.cell(next_row, 1).value = '规则列含义'
        ws.cell(next_row, 2).value = EXPLAIN_NOTE
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(path)
print(f'updated {path}')
