
from pathlib import Path
from openpyxl import load_workbook
paths=[x for x in Path('outputs').glob('*20260608*修正版.xlsx') if not x.name.startswith('~')]
p=paths[0]
wb=load_workbook(p,data_only=True)
for sheet in ['AgentScoring','BusinessScoring']:
    ws=wb[sheet]
    print(sheet, ws.max_row, ws.max_column)
    print([ws.cell(3,c).value for c in range(1, ws.max_column+1)])
    for row in ws.iter_rows(min_row=4,max_row=min(ws.max_row,10),values_only=True):
        print(row)
    print('---')
