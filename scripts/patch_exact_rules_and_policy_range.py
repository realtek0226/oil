
from pathlib import Path
import yaml
from openpyxl import load_workbook

MAIN = Path('outputs') / '逻辑说明.xlsx'
YAML = Path('configs/scorecards/shandong_scorecards_v1.yaml')

def rule_text(feature):
    method = feature.get('method')
    if method == 'bucket_score':
        parts=[]
        for r in feature.get('rules') or []:
            lo,hi,score,label = r.get('min'), r.get('max'), r.get('score'), r.get('label') or ''
            if lo is None:
                interval=f'<{hi}'
            elif hi is None:
                interval=f'>={lo}'
            else:
                interval=f'>={lo}且<{hi}'
            parts.append(f'{interval}: {score}（{label}）')
        return '分桶规则：' + '；'.join(parts) + '；未命中/缺失按0分。'
    if method == 'enum_score':
        return '枚举规则：' + '；'.join(f'{k}: {v}' for k,v in (feature.get('rules') or {}).items()) + '；未命中/缺失按0分。'
    if method == 'calendar_month_band':
        return '月度季节规则：按目标月所在季节桶打分；缺失按0分。'
    if method == 'bounded_numeric_adjustment':
        return '有界数值：取值直接作为加减分，按配置上下限封顶；缺失按0分。'
    return f'方法：{method or "未配置"}；缺失按0分。'

config=yaml.safe_load(YAML.read_text(encoding='utf-8'))
fmap={}
for card in config['scorecards']:
    for spec in (card.get('horizons') or {}).values():
        for group in spec.get('factor_groups') or []:
            for feature in group.get('features') or []:
                fmap[feature.get('feature_name')]=feature
notes={
 'brent_change_usd_d1':'Brent D1=88.32-90.38=-2.06；按D1成本侧分桶。',
 'brent_change_usd_d3':'6.12日报未给D3点位，系统标记缺失；不能用D1点位冒充D3。',
 'brent_change_usd_w1':'Brent W1=88.79-90.38=-1.59；按W1成本侧分桶。',
 'brent_change_usd_mom':'Brent M1=89.53-90.38=-0.85；按M1成本侧分桶。',
 'gasoline_crack_percentile':'裂解价差分位从2024-01-01起算；低裂解利多，高裂解利空。',
 'shandong_cdu_utilization_weekly':'山东地炼开工率按原始百分比区间打分；若前后两条数据不变则置0。',
 'shandong_product_inventory_percentile_weekly':'库存=贸易商+主营+独立炼厂可用项求和；不是三项必须齐全。若某组件上期有数本期缺失，提示缺失并按0；分位从2024-01-01起算。',
 'restocking_rhythm_monthly':'山东M1备货节奏=比较预测月前一月与前二月山东独立炼厂汽油产销率>90%的天数；前一月更多+5，减少-5，持平0。',
 'main_company_inventory_monthly':'山东主营销售公司库存来自隆众OpenAPI；按2024-01-01以来分位打分，不变则0。',
 'refinery_inventory_monthly':'山东独立炼厂库存按2024-01-01以来分位打分，不变则0。',
 'price_window_expectation_weekly':'调价预测来自卓创成品油调价专栏；W1按上调/中性/下调枚举打分。',
 'price_window_expectation_monthly':'调价预测来自卓创成品油调价专栏；M1按上调/中性/下调枚举打分。',
}
wb=load_workbook(MAIN)
ws=wb['BusinessScoring']
for r in range(4, ws.max_row+1):
    field=ws.cell(r,4).value
    if field in fmap:
        ws.cell(r,11).value=rule_text(fmap[field])
    if field in notes:
        ws.cell(r,12).value=notes[field]
# add exact policy theoretical interval section
if 'PolicyTheory' in wb.sheetnames:
    del wb['PolicyTheory']
pt=wb.create_sheet('PolicyTheory')
pt.append(['调价预测理论区间修正'])
pt.append(['结论','调价预测理论有效区间应为 -82 ~ 89.5，不再使用旧表里的 -99.5 ~ 99.5。'])
pt.append(['项目','当前口径'])
pt.append(['调价预测金额分桶','>=100:+70；50~100:+45；0~50:+20；-50~0:-20；-100~-50:-45；<-100:-70'])
pt.append(['窗口系数','<=2个工作日*1.25；>=8个工作日*0.75；其余*1'])
pt.append(['理论有效区间','下沿 -82；上沿 89.5'])
pt.append(['使用方式','历史状态桶结果用于校准有效区间；当天预测仍优先服从硬数据、黑天鹅事件和时间维度完整性。'])
# add state bucket complete note
wsb=wb['BucketCalibration']
wsb.cell(2,2).value='历史样本只用于校准状态桶/有效区间，不直接改写当天硬数据；调价预测理论有效区间已修正为 -82 ~ 89.5。'
wb.save(MAIN)
print(MAIN)
