from pathlib import Path

from openpyxl import load_workbook


def safe_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ")[:240]


def main() -> None:
    path = Path("outputs") / "\u903b\u8f91\u8bf4\u660e.xlsx"
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, data_only=False)
    print(f"file={path}")
    print("sheets=" + ", ".join(workbook.sheetnames))

    sheet_names = [
        "\u8868\u683c\u8bf4\u660e",
        "BusinessScoring",
        "\u4e1a\u52a1\u6253\u5206\u660e\u7ec6_\u53ef\u5220",
        "AgentScoring",
    ]
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            print(f"MISSING_SHEET={sheet_name}")
            continue
        sheet = workbook[sheet_name]
        print(f"\n[{sheet_name}] rows={sheet.max_row} cols={sheet.max_column}")
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True):
            print(" | ".join(safe_text(cell) for cell in row[:13]))

    agent_sheet = workbook["AgentScoring"]
    print("\n[event-risk rows]")
    for row_index, row in enumerate(agent_sheet.iter_rows(values_only=True), start=1):
        text = " ".join(safe_text(cell) for cell in row)
        if (
            "\u4e8b\u4ef6" in text
            or "event" in text.lower()
            or "\u505c\u6218" in text
            or "\u970d\u5c14\u6728\u5179" in text
        ):
            print(f"row={row_index} :: {text}")


if __name__ == "__main__":
    main()
