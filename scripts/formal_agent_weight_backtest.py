from __future__ import annotations

import json
import math
import platform
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Any

platform._wmi_query = lambda *a, **k: ['10.0.0', '1', 'Multiprocessor Free', '0', '0']
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.container import get_dataset_service, get_predictor
from app.services.predictors.horizons import resolve_horizon_config
from app.services.predictors.shandong_gas92 import OUTRIGHT_EXPERT_PRIOR_WEIGHTS

START_DATE = date(2024, 1, 1)
AS_OF = date(2026, 6, 12)
HORIZON = 'D1'
OUTPUT_JSON = Path('artifacts/formal_agent_weight_backtest_20240101_20260612.json')
OUTPUT_XLSX = Path('outputs/formal_agent_weight_backtest_20240101_20260612.xlsx')
AGENTS = [
    'crude_cost_agent',
    'supply_inventory_agent',
    'demand_seasonality_agent',
    'market_structure_agent',
    'policy_cycle_agent',
    'refined_oil_news_agent',
]


def safe_corr(x: np.ndarray, y: np.ndarray) -> float | None:
    if len(x) < 3 or float(np.std(x)) == 0.0 or float(np.std(y)) == 0.0:
        return None
    return round(float(np.corrcoef(x, y)[0, 1]), 4)


def direction_accuracy(pred: np.ndarray, actual: np.ndarray) -> float:
    pred_dir = np.sign(pred)
    actual_dir = np.sign(actual)
    mask = actual_dir != 0
    if not bool(mask.any()):
        return 0.0
    return round(float((pred_dir[mask] == actual_dir[mask]).mean()), 4)


def evaluate_weighted(features: np.ndarray, weights: np.ndarray, actual: np.ndarray) -> dict[str, float]:
    score = features @ weights
    # Convert normalized score to yuan/ton using least squares single scale for evaluation only.
    denom = float(np.dot(score, score))
    scale = float(np.dot(score, actual) / denom) if denom > 1e-9 else 0.0
    pred = score * scale
    return {
        'scale_yuan_per_score': round(scale, 4),
        'mae': round(float(np.mean(np.abs(pred - actual))), 4),
        'rmse': round(float(math.sqrt(np.mean((pred - actual) ** 2))), 4),
        'direction_accuracy': direction_accuracy(pred, actual),
    }


def ridge_weights(features: np.ndarray, actual: np.ndarray, alpha: float = 8.0) -> np.ndarray:
    x = features.copy()
    y = actual.copy()
    # Standardize columns for stable coefficient comparison.
    means = x.mean(axis=0)
    stds = x.std(axis=0)
    stds[stds == 0] = 1.0
    xs = (x - means) / stds
    y = y - y.mean()
    eye = np.eye(xs.shape[1])
    coef = np.linalg.solve(xs.T @ xs + alpha * eye, xs.T @ y)
    importance = np.abs(coef)
    if float(importance.sum()) <= 1e-9:
        return np.ones(xs.shape[1]) / xs.shape[1]
    return importance / importance.sum()


def build() -> dict[str, Any]:
    dataset_service = get_dataset_service()
    predictor = get_predictor()
    horizon_config = resolve_horizon_config(HORIZON)
    frame = dataset_service.build_feature_frame(
        start_date=START_DATE - timedelta(days=180),
        end_date=AS_OF,
    ).sort_values('date').copy()
    frame['date'] = pd.to_datetime(frame['date'])
    frame = frame[frame['date'] >= pd.Timestamp(START_DATE)].copy()
    frame['target_date'] = frame['date'].shift(-horizon_config.steps)
    frame['target_price'] = frame['sd_gas92_market'].shift(-horizon_config.steps)
    frame['target_delta'] = frame['target_price'] - frame['sd_gas92_market']
    history = frame[(frame['date'] < pd.Timestamp(AS_OF)) & (frame['target_date'] <= pd.Timestamp(AS_OF))]
    history = history.dropna(subset=['target_delta']).copy()

    rows: list[dict[str, Any]] = []
    for _, row in history.iterrows():
        extra = {
            'as_of_date': row['date'].date() if hasattr(row['date'], 'date') else row['date'],
            'mode': 'calibration_backtest',
            'report_payload': None,
            'news_items': [],
            'refined_news_items': [],
            'policy_items': [],
            'prediction_subject': 'outright',
            'enable_refined_news': False,
            'enable_event_risk': False,
            'horizon': HORIZON,
        }
        claims, composite_score = predictor._score_row(row, extra)
        item: dict[str, Any] = {
            'date': str(row['date'].date()),
            'target_delta': float(row['target_delta']),
            'current_composite_score': float(composite_score),
        }
        for claim in claims:
            if claim.agent_name in AGENTS:
                item[claim.agent_name] = float(claim.numeric_signals.get('normalized_score', claim.numeric_signals.get('score', 0.0)))
        rows.append(item)

    data = pd.DataFrame(rows).dropna(subset=['target_delta'])
    for agent in AGENTS:
        if agent not in data.columns:
            data[agent] = 0.0
        data[agent] = data[agent].fillna(0.0)
    features = data[AGENTS].astype(float).to_numpy()
    actual = data['target_delta'].astype(float).to_numpy()
    current_weights = np.array([OUTRIGHT_EXPERT_PRIOR_WEIGHTS.get(agent, 0.0) for agent in AGENTS], dtype=float)
    current_weights = current_weights / current_weights.sum()
    suggested = ridge_weights(features, actual)

    agent_stats = []
    for idx, agent in enumerate(AGENTS):
        values = features[:, idx]
        agent_stats.append(
            {
                '智能体': agent,
                '当前权重': round(float(current_weights[idx]), 4),
                '正式回测建议权重': round(float(suggested[idx]), 4),
                '与真实次日涨跌相关系数': safe_corr(values, actual),
                '有效波动样本数': int(np.sum(np.abs(values) > 1e-9)),
                '业务解释': explain_agent(agent),
            }
        )

    summary = {
        'start_date': str(START_DATE),
        'as_of': str(AS_OF),
        'horizon': HORIZON,
        'sample_count': int(len(data)),
        'method': '2024-01-01以来逐日重算各智能体标准化分；用真实次日山东92#涨跌做岭回归权重复核；当前生产权重不自动替换。',
        'current_weight_evaluation': evaluate_weighted(features, current_weights, actual),
        'suggested_weight_evaluation': evaluate_weighted(features, suggested, actual),
        'agents': agent_stats,
        'note': '正式回测给出权重复核依据；若建议权重与业务常识冲突或样本特征覆盖不足，先进入专家复核，不直接替换生产参数。',
    }
    return summary


def explain_agent(agent: str) -> str:
    mapping = {
        'crude_cost_agent': '原油成本是成品油价格的底层成本驱动，权重应较高。',
        'supply_inventory_agent': '供给、开工、库存决定现货松紧，是硬数据驱动。',
        'demand_seasonality_agent': '产销率、季节性和备货节奏影响短期成交。',
        'market_structure_agent': '区域价差反映山东外流和跨区套利空间。',
        'policy_cycle_agent': '调价窗口和政策预期影响批发市场节奏。',
        'refined_oil_news_agent': '资讯和成交心态是软信号，样本不全时只作辅助。',
    }
    return mapping.get(agent, '')


def write_excel(summary: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = '正式权重回测'
    ws.append(['项目', '内容'])
    ws.append(['样本区间', f"{summary['start_date']}至{summary['as_of']}"])
    ws.append(['样本数', summary['sample_count']])
    ws.append(['方法', summary['method']])
    ws.append(['当前权重评估', json.dumps(summary['current_weight_evaluation'], ensure_ascii=False)])
    ws.append(['建议权重评估', json.dumps(summary['suggested_weight_evaluation'], ensure_ascii=False)])
    ws.append(['注意', summary['note']])
    ws2 = wb.create_sheet('智能体权重明细')
    ws2.append(['智能体', '当前权重', '正式回测建议权重', '与真实次日涨跌相关系数', '有效波动样本数', '业务解释'])
    for item in summary['agents']:
        ws2.append([item['智能体'], item['当前权重'], item['正式回测建议权重'], item['与真实次日涨跌相关系数'], item['有效波动样本数'], item['业务解释']])
    thin = Side(style='thin', color='D9E2EC')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.font = Font(name='微软雅黑', size=10)
                if cell.row == 1:
                    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
                    cell.fill = header_fill
        for index in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(index)].width = 22 if index < sheet.max_column else 48
    wb.save(OUTPUT_XLSX)


if __name__ == '__main__':
    summary = build()
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    write_excel(summary)
    print(json.dumps({
        'json': str(OUTPUT_JSON),
        'xlsx': str(OUTPUT_XLSX),
        'sample_count': summary['sample_count'],
        'current': summary['current_weight_evaluation'],
        'suggested': summary['suggested_weight_evaluation'],
    }, ensure_ascii=False, indent=2))
