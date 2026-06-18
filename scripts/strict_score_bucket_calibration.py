
from __future__ import annotations

import argparse
import json
import math
import platform
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Any

platform._wmi_query = lambda *a, **k: ['10.0.0', '1', 'Multiprocessor Free', '0', '0']
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.container import get_dataset_service, get_predictor
from app.services.predictors.horizons import DEFAULT_HORIZONS, resolve_horizon_config

DEFAULT_START_DATE = date(2024, 1, 1)
DEFAULT_AS_OF = date(2026, 6, 12)
MIN_BUCKET_SAMPLE_SIZE = 12
OUTPUT_JSON = Path('artifacts/strict_score_bucket_calibration_no_missing_20240101_20260612.json')
OUTPUT_XLSX = Path('outputs/strict_score_bucket_calibration_no_missing_20240101_20260612.xlsx')

REQUIRED_ARCHIVE_COLUMNS = {
    'D1': [
        'sd_gas_crack',
        'shandong_cdu_utilization_weekly',
        'sales_production_ratio_d1',
        'shandong_product_inventory_percentile_weekly',
    ],
    'W1': [
        'sd_gas_crack',
        'shandong_cdu_utilization_weekly',
        'sales_production_ratio_w1_avg',
        'shandong_product_inventory_percentile_weekly',
    ],
    'M1': [
        'brent_change_20d',
        'sd_gas_crack',
        'shandong_cdu_utilization_weekly',
        'sales_production_ratio_monthly_avg',
        'refinery_inventory_monthly',
        'main_company_inventory_monthly',
    ],
}


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, str):
        return value.strip() == '' or value.strip().lower() in {'missing', 'none', 'nan'}
    return False


def q(values: list[float], quantile: float) -> float | None:
    values = [float(value) for value in values if value == value]
    if not values:
        return None
    return round(float(np.quantile(values, quantile)), 2)


def up_rate(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(1 for value in values if value > 0) / len(values), 3)


def score_points(predictor: Any, score_column: str, value: float) -> float:
    return predictor._score_points(score_column=score_column, score_value=float(value))


def bucket_index(score: float, defs: list[dict[str, Any]]) -> int:
    for idx, item in enumerate(defs):
        if float(item['lower']) <= float(score) < float(item['upper']):
            return idx
    return len(defs) - 1


def business_missing_fields(claim: Any) -> list[str]:
    payload = claim.structured_payload or {}
    scorecard = payload.get('scorecard') or {}
    quality = scorecard.get('data_quality') or {}
    return [str(x) for x in (quality.get('missing_fields') or []) if str(x)]


def claims_missing_fields(claims: list[Any]) -> list[str]:
    missing: list[str] = []
    for claim in claims:
        payload = claim.structured_payload or {}
        quality = payload.get('data_quality') or {}
        for field in quality.get('missing_fields') or []:
            if str(field):
                missing.append(f"{claim.agent_name}:{field}")
    return missing


def strict_score_frame(
    predictor: Any,
    frame: pd.DataFrame,
    horizon: str,
    *,
    refined_news_by_date: dict[date, list[dict[str, Any]]] | None = None,
    event_news_by_date: dict[date, list[dict[str, Any]]] | None = None,
    event_report_by_date: dict[date, dict[str, Any] | None] | None = None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    refined_news_by_date = refined_news_by_date or {}
    event_news_by_date = event_news_by_date or {}
    event_report_by_date = event_report_by_date or {}
    required_columns = REQUIRED_ARCHIVE_COLUMNS.get(horizon, [])
    for _, row in frame.iterrows():
        row_date = pd.Timestamp(row['date']).date()
        missing = [col for col in required_columns if col not in row.index or is_missing(row.get(col))]
        refined_items = refined_news_by_date.get(row_date, [])
        event_items = event_news_by_date.get(row_date, [])
        report_payload = event_report_by_date.get(row_date)
        backtest_extra = {
            'as_of_date': row_date,
            'mode': 'strict_calibration_no_missing',
            'report_payload': report_payload,
            'news_items': event_items,
            'refined_news_items': refined_items,
            'policy_items': [],
            'prediction_subject': 'outright',
            'enable_refined_news': True,
            'enable_event_risk': True,
            'horizon': horizon,
        }
        backtest_extra['refined_news_labels'] = predictor._build_refined_news_label_signal(
            as_of_date=row_date,
            mode='backtest',
            refined_news_items=refined_items,
            scenario_text=None,
        )
        backtest_extra['trade_sentiment'] = predictor._build_trade_sentiment_signal(
            as_of_date=row_date,
            mode='backtest',
            refined_news_items=refined_items,
            news_items=event_items,
            scenario_text=None,
        )
        backtest_extra['monthly_market_sentiment'] = predictor._build_monthly_market_sentiment_signal(
            as_of_date=row_date,
            mode='backtest',
            refined_news_items=refined_items,
            news_items=event_items,
            scenario_text=None,
        )
        business_claim = predictor._score_business_scorecard(row, extra=backtest_extra)
        missing.extend(business_missing_fields(business_claim))
        claims, agent_score = predictor._score_row(row, extra=backtest_extra)
        missing.extend(claims_missing_fields(claims))
        item = row.to_dict()
        item['agent_score'] = float(agent_score)
        item['business_scorecard_score'] = float(business_claim.numeric_signals.get('standalone_score', 0.0))
        item['strict_missing_fields'] = sorted(set(missing))
        item['strict_missing_count'] = len(set(missing))
        item['strict_complete'] = len(set(missing)) == 0
        rows.append(item)
    return pd.DataFrame(rows)


def evaluate_buckets(predictor: Any, scored: pd.DataFrame, score_column: str) -> list[dict[str, Any]]:
    defs = predictor._score_bucket_defs(score_column)
    out: list[dict[str, Any]] = []
    if scored.empty:
        for item in defs:
            out.append({
                'bucket': item['label'],
                'score_range': item['range_label'],
                'sample_size': 0,
                'p25': None,
                'p50': None,
                'p75': None,
                'up_rate': None,
                'sample_sufficient': False,
            })
        return out
    work = scored.copy()
    work['_score_points'] = work[score_column].map(lambda value: score_points(predictor, score_column, float(value)))
    work['_bucket_index'] = work['_score_points'].map(lambda value: bucket_index(float(value), defs))
    for idx, item in enumerate(defs):
        vals = work[work['_bucket_index'] == idx]['target_delta'].astype(float).tolist()
        out.append({
            'bucket': item['label'],
            'score_range': item['range_label'],
            'sample_size': len(vals),
            'p25': q(vals, 0.25),
            'p50': q(vals, 0.5),
            'p75': q(vals, 0.75),
            'up_rate': up_rate(vals),
            'sample_sufficient': len(vals) >= MIN_BUCKET_SAMPLE_SIZE,
        })
    return out


def build(start_date: date = DEFAULT_START_DATE, as_of: date = DEFAULT_AS_OF) -> dict[str, Any]:
    dataset_service = get_dataset_service()
    predictor = get_predictor()
    frame_start = start_date - timedelta(days=180)
    frame = dataset_service.build_feature_frame(start_date=frame_start, end_date=as_of).sort_values('date').copy()
    refined_snapshot = dataset_service.build_archived_refined_news_snapshot(start_date=start_date, end_date=as_of)
    event_snapshot = dataset_service.build_archived_event_risk_snapshot(start_date=start_date, end_date=as_of)
    frame['date'] = pd.to_datetime(frame['date'])
    frame = frame[frame['date'] >= pd.Timestamp(start_date)].copy()
    summary: dict[str, Any] = {
        'calibration_policy': 'no_missing_fallback; rows with any required factor missing are excluded; thresholds are not updated if strict samples are insufficient',
        'start_date': start_date.isoformat(),
        'as_of': as_of.isoformat(),
        'frame_rows': int(len(frame)),
        'horizons': {},
    }
    for horizon in DEFAULT_HORIZONS:
        hc = resolve_horizon_config(horizon)
        work = frame.copy()
        work['target_date'] = work['date'].shift(-hc.steps)
        work['target_price'] = work['sd_gas92_market'].shift(-hc.steps)
        work['target_delta'] = work['target_price'] - work['sd_gas92_market']
        history = work[(work['date'] < pd.Timestamp(as_of)) & (work['target_date'] <= pd.Timestamp(as_of))].dropna(subset=['target_delta']).copy()
        scored_all = strict_score_frame(
            predictor,
            history,
            horizon=horizon,
            refined_news_by_date=refined_snapshot.items_by_date,
            event_news_by_date=event_snapshot.news_items_by_date,
            event_report_by_date=event_snapshot.report_by_date,
        )
        if scored_all.empty or 'strict_complete' not in scored_all.columns:
            strict = scored_all.iloc[0:0].copy()
            missing_counts = {}
        else:
            strict = scored_all[scored_all['strict_complete']].dropna(subset=['agent_score', 'business_scorecard_score', 'target_delta']).copy()
            missing_counts = scored_all['strict_missing_count'].value_counts().sort_index().to_dict()
        top_missing: dict[str, int] = {}
        for fields in scored_all.get('strict_missing_fields', pd.Series(dtype=object)):
            for field in fields or []:
                top_missing[field] = top_missing.get(field, 0) + 1
        top_missing = dict(sorted(top_missing.items(), key=lambda item: item[1], reverse=True)[:30])
        summary['horizons'][horizon] = {
            'raw_history_samples': int(len(history)),
            'strict_complete_samples': int(len(strict)),
            'excluded_samples': int(len(scored_all) - len(strict)),
            'missing_count_distribution': {str(k): int(v) for k, v in missing_counts.items()},
            'top_missing_fields': top_missing,
            'threshold_update_allowed': bool(len(strict) >= MIN_BUCKET_SAMPLE_SIZE * 7),
            'agent_buckets': evaluate_buckets(predictor, strict, 'agent_score'),
            'business_buckets': evaluate_buckets(predictor, strict, 'business_scorecard_score'),
        }
    return summary


def write_xlsx_to_path(summary: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'NoMissingSummary'
    headers = ['horizon', 'raw_samples', 'complete_samples', 'excluded_samples', 'threshold_update_allowed', 'missing_distribution', 'top_missing_fields']
    ws.append(headers)
    for horizon, item in summary['horizons'].items():
        ws.append([
            horizon,
            item['raw_history_samples'],
            item['strict_complete_samples'],
            item['excluded_samples'],
            '?' if item['threshold_update_allowed'] else '?',
            json.dumps(item['missing_count_distribution'], ensure_ascii=False),
            json.dumps(item['top_missing_fields'], ensure_ascii=False),
        ])
    for horizon, item in summary['horizons'].items():
        for model_key, title in [('agent_buckets', f'{horizon}_AgentBuckets'), ('business_buckets', f'{horizon}_BusinessBuckets')]:
            ws2 = wb.create_sheet(title[:31])
            ws2.append(['bucket', 'score_range', 'sample_size', 'P25', 'P50', 'P75', 'up_rate', 'sample_sufficient'])
            for row in item[model_key]:
                ws2.append([row['sample_size'], row['up_rate'], row['sample_size'], row['p25'], row['p50'], row['p75'], row['up_rate'], 'yes' if row['sample_sufficient'] else 'no'])
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18 if col < 7 else 60
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='FCE4D6')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_iso_date(value: str, fallback: date) -> date:
    if not value:
        return fallback
    return date.fromisoformat(value)


def main() -> None:
    parser = argparse.ArgumentParser(description='Strict no-missing score bucket calibration.')
    parser.add_argument('--start-date', default=DEFAULT_START_DATE.isoformat())
    parser.add_argument('--as-of', default=DEFAULT_AS_OF.isoformat())
    parser.add_argument('--output-json', default='')
    parser.add_argument('--output-xlsx', default='')
    args = parser.parse_args()
    start_date = parse_iso_date(args.start_date, DEFAULT_START_DATE)
    as_of = parse_iso_date(args.as_of, DEFAULT_AS_OF)
    summary = build(start_date=start_date, as_of=as_of)
    output_json = Path(args.output_json) if args.output_json else OUTPUT_JSON
    output_xlsx = Path(args.output_xlsx) if args.output_xlsx else Path(OUTPUT_XLSX)
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
    write_xlsx_to_path(summary, output_xlsx)
    print(json.dumps({
        'json': str(output_json),
        'xlsx': str(output_xlsx),
        'policy': summary['calibration_policy'],
        'samples': {h: {'raw': v['raw_history_samples'], 'complete': v['strict_complete_samples'], 'excluded': v['excluded_samples'], 'threshold_update_allowed': v['threshold_update_allowed']} for h, v in summary['horizons'].items()},
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()

