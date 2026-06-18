
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MAIN = Path('outputs') / '逻辑说明.xlsx'

AGENT_RAW_RULES = {
 '原油成本智能体': '原始规则：raw_score=clip(Brent分桶分+调油修正,-100,100)；max_score=65；normalized=raw_score/65；加权贡献=normalized*0.22。Brent分桶原始规则：>=3:+55，2~3:+42，1~2:+28，0.5~1:+14，-0.5~0.5:0，-1~-0.5:-14，-2~-1:-28，-3~-2:-42，<-3:-55。',
 '市场结构智能体': '原始规则：raw_score=clip(区域均值分+全国相对山东分+动量分,-100,100)；max_score=100；normalized=raw_score/100；加权贡献=normalized*0.16。区域均值分=clip(平均(区域价-山东价)/120*65,-65,65)。',
 '供给库存智能体': '原始规则：raw_score=clip(开工分位分+库存分位分+开工变化分+炼油利润分,-100,100)；max_score=62；normalized=raw_score/62；加权贡献=normalized*0.20。开工分位满分40，库存分位满分12。',
 '需求季节智能体': '原始规则：raw_score=clip(产销率分+备货节奏/月度变化分+季节性分+节假日分,-100,100)；max_score=100；normalized=raw_score/100；加权贡献=normalized*0.18。D1/D3/W1分别取当日/3日均值/7日均值。',
 '成品油资讯智能体': '原始规则：raw_score由资讯标签映射；成交活跃/抢货/推涨为正，抛货/让利/看跌为负；max_score=100；normalized=raw_score/100；加权贡献=normalized*0.12。无标签或关闭时0分。',
 '山东现货跳变识别': '原始规则：只做D1点位修正，不参与综合加权；权重=0。硬信号包括低价资源扫空、封单惜售、抢货/抛货、出货强弱；没有硬信号则0。',
 '政策周期智能体': '原始规则：raw_score=调价金额分桶分*窗口系数+上轮调价余波；max_score=100；normalized=raw_score/100；加权贡献=normalized*0.12。本次-320元/吨命中<-100分桶，窗口约4个工作日，理论有效区间修正为-82~89.5。',
 '事件风险智能体': '原始规则：事件风险为门控/裁判层输入，权重=0；供应缓和类黑天鹅（停战、海峡解封）必须压低油价方向，不允许被裁判层反向抵消。',
}
MODULE_RAW_RULES = {
 ('D1','成本侧'): '原始规则：模块得分=Brent日变化分+汽油裂解价差分，组内封顶±40。本次Brent=-2.06命中-30，裂解分位66.25命中-5，模块得分=-35。',
 ('D1','供给侧'): '原始规则：模块得分=山东地炼开工率原始百分比分+炼厂负荷新闻修正，组内封顶±30；若前后两条开工数据不变则置0。本次开工率51.09命中0，新闻修正0，模块得分0。',
 ('D1','需求侧'): '原始规则：模块得分=汽油产销率D1分；>=110:+30，100~110:+15，95~100:+5，90~95:0，85~90:-5，70~85:-15，<70:-30。本次108得+15。',
 ('D1','情绪侧'): '原始规则：bullish_active:+15，neutral_flat:0，bearish_selling:-15。本次bullish_active得+15。',
 ('D3','成本侧'): '原始规则：模块得分=Brent三日变化分+裂解分位分，组内封顶±35。6.12日报未给D3点位，Brent项按缺失0；裂解分位66.25得-5，模块得分-5。',
 ('D3','供给侧'): '原始规则：模块得分=山东地炼开工率原始百分比分+负荷新闻修正，组内封顶±35；前后数据不变则0。本次开工率51.09得0，新闻修正0。',
 ('D3','需求侧'): '原始规则：模块得分=汽油产销率3日均值分；>=120:+20，105~120:+10，100~105:+5，95~100:0，85~95:-5，70~85:-10，<70:-20。',
 ('D3','情绪侧'): '原始规则：bullish_active:+10，neutral_flat:0，bearish_selling:-10。',
 ('W1','成本侧'): '原始规则：模块得分=Brent周变化分+裂解分位分，组内封顶±30。本次Brent=88.79-90.38=-1.59，裂解分位66.25。',
 ('W1','供给侧'): '原始规则：模块得分=山东地炼开工率分+负荷新闻修正，组内封顶±30；开工不变则0。',
 ('W1','需求侧'): '原始规则：模块得分=汽油产销率7日均值分；>=120:+20，105~120:+10，100~105:+5，95~100:0，85~95:-5，70~85:-10，<70:-20。',
 ('W1','库存端'): '原始规则：库存=贸易商+主营+独立炼厂可用项求和；若上期有数本期缺失则按缺失0；否则按2024-01-01以来分位打分：<25:+15，25~45:+5，45~55:0，55~75:-5，>=75:-15。本次21.37%得+15。',
 ('W1','调价预期'): '原始规则：up_adjustment_expected:+5，neutral:0，down_adjustment_expected:-5。本次卓创调价预测-320元/吨，命中down_adjustment_expected，得-5。',
 ('M1','成本侧'): '原始规则：模块得分=Brent月变化分+裂解分位分，M1 Brent封顶±20，裂解封顶±10，成本组封顶±30。本次Brent=89.53-90.38=-0.85。',
 ('M1','供给侧'): '原始规则：检修计划+山东地炼开工率组成供给侧，集中检修+15，负荷平稳0，复工过剩-15；开工率按区间分。',
 ('M1','需求侧'): '原始规则：月度季节性+备货节奏+节假日需求。备货节奏比较前一月与前二月山东独立炼厂汽油产销率>90%的天数；前一月更多+5，减少-5，持平0。',
 ('M1','库存端'): '原始规则：独立炼厂库存分位±10+主营库存分位±10，组内封顶±20；分位从2024-01-01起算；库存较前一条不变则置0。',
 ('M1','政策与情绪侧'): '原始规则：调价窗口预期+月度市场情绪；调价上调+5/中性0/下调-5，月度情绪peak_season_bullish:+5、neutral:0、bearish:-5。',
}

def style(ws):
    header_fill=PatternFill('solid', fgColor='1F4E78')
    header_font=Font(color='FFFFFF', bold=True)
    thin=Side(style='thin', color='D9E2F3')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(vertical='top', wrap_text=True)
            cell.border=border
            if cell.row==3:
                cell.fill=header_fill; cell.font=header_font
    for col in range(1, ws.max_column+1):
        width=12
        for cell in ws.iter_cols(min_col=col,max_col=col,min_row=1,max_row=ws.max_row):
            for item in cell:
                width=max(width, min(len(str(item.value or '')), 70))
        ws.column_dimensions[get_column_letter(col)].width=min(max(width+2, 12), 58)
    for r in range(1, ws.max_row+1): ws.row_dimensions[r].height=36 if r>3 else 24

wb=load_workbook(MAIN)
# AgentScoring: insert 原始规则 after 打分参照规则 if missing
ws=wb['AgentScoring']
headers=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
if '原始规则' not in headers:
    idx=headers.index('打分参照规则')+2
    ws.insert_cols(idx)
    ws.cell(3,idx).value='原始规则'
else:
    idx=headers.index('原始规则')+1
for r in range(4, ws.max_row+1):
    name=ws.cell(r,2).value
    ws.cell(r,idx).value=AGENT_RAW_RULES.get(name,'原始规则：按当前代码 numeric_signals 计算 raw_score、max_score、normalized_score、weight 和 weighted_score。')
style(ws)
# Rebuild BusinessScoring aggregated by cycle+module
old=wb['BusinessScoring']
rows=list(old.iter_rows(min_row=4, values_only=True))
if 'BusinessScoring_Detail' not in wb.sheetnames:
    detail=wb.copy_worksheet(old)
    detail.title='BusinessScoring_Detail'
# remove old and recreate at same position
pos=wb.sheetnames.index('BusinessScoring')
del wb['BusinessScoring']
ws=wb.create_sheet('BusinessScoring', pos)
ws.append(['业务用数与打分'])
ws.append(['说明','同一周期同一模块合并为一行；右侧新增“原始规则”，展示业务原始赋值/封顶/缺失口径。明细保留在 BusinessScoring_Detail。'])
headers=['周期','模块','数据项','字段','本次取值','单位','单项/原始分','满分','模块得分','命中/结论','打分参照规则','原始规则','口径/证据']
ws.append(headers)
from collections import OrderedDict
groups=OrderedDict()
for row in rows:
    if not row or not row[0]: continue
    key=(row[0], row[1])
    groups.setdefault(key, []).append(row)
for (h,module), items in groups.items():
    data_items='；'.join(str(x[2] or '') for x in items)
    fields='；'.join(str(x[3] or '') for x in items)
    values='；'.join(f"{x[3]}={x[4]}" for x in items)
    units='；'.join(sorted({str(x[5]) for x in items if x[5]}))
    scores='；'.join(f"{x[3]}:{x[6]}" for x in items)
    caps='；'.join(str(x[7] or '') for x in items)
    module_score=items[-1][8]
    hits='；'.join(f"{x[3]}->{x[9]}" for x in items)
    refs='\\n'.join(f"【{x[2]}】{x[10]}" for x in items)
    raw=MODULE_RAW_RULES.get((h,module),'原始规则：模块得分为本模块各字段原始分求和后按模块满分封顶；字段缺失不做方向假设，按0分。')
    evidence='\\n'.join(f"【{x[2]}】{x[11]}" for x in items)
    ws.append([h,module,data_items,fields,values,units,scores,caps,module_score,hits,refs,raw,evidence])
style(ws)
# improve detail sheet too: insert 原始规则 for traceability
if 'BusinessScoring_Detail' in wb.sheetnames:
    d=wb['BusinessScoring_Detail']
    hs=[d.cell(3,c).value for c in range(1,d.max_column+1)]
    if '原始规则' not in hs and '打分参照规则' in hs:
        insert_at=hs.index('打分参照规则')+2
        d.insert_cols(insert_at)
        d.cell(3,insert_at).value='原始规则'
    else:
        insert_at=hs.index('原始规则')+1 if '原始规则' in hs else None
    if insert_at:
        for r in range(4,d.max_row+1):
            key=(d.cell(r,1).value,d.cell(r,2).value)
            d.cell(r,insert_at).value=MODULE_RAW_RULES.get(key,'原始规则：按当前YAML字段规则计分，模块内求和后封顶。')
    style(d)

try:
    wb.save(MAIN)
    print(MAIN)
except PermissionError:
    ALT = MAIN.with_name(MAIN.stem + '_已更新待替换.xlsx')
    wb.save(ALT)
    print(ALT)

